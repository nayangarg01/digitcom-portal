import openpyxl
import xlsxwriter
from datetime import datetime

master_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/InputBilling/MASTER TRACKER FOR BILLING-AIRFIBER-RJST (1).xlsx'
dc_number = 'DC0111'
output_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/Billing/DC0111_A6B6_Verification.xlsx'

def safe_float(val):
    if val is None or str(val).strip() == "" or str(val).strip().upper() in ["NR", "NA", "-"]:
        return 0.0
    try: return float(val)
    except: return 0.0

def format_date(val):
    if val is None: return ""
    if isinstance(val, datetime):
        return val.strftime('%d-%b-%y')
    return str(val)

try:
    # 1. Load Data
    wb_master = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
    ws_master = wb_master['A6+B6 Billings']
    headers = [str(h).strip() for h in next(ws_master.iter_rows(max_row=1, values_only=True))]
    dc_col_idx = headers.index('BILLING FILE')
    rows = []
    for row in ws_master.iter_rows(min_row=2, values_only=True):
        if row[dc_col_idx] and str(row[dc_col_idx]).strip().upper() == dc_number:
            rows.append(dict(zip(headers, row)))
    
    # 2. Generate Excel
    with xlsxwriter.Workbook(output_path) as wb:
        # Formats
        f_title = wb.add_format({'bold': True, 'font_size': 12, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#DCE6F1', 'border': 1})
        f_cert = wb.add_format({'bold': True, 'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'border': 1})
        f_head_blue = wb.add_format({'bold': True, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#DCE6F1', 'border': 1, 'text_wrap': True})
        f_head_yellow = wb.add_format({'bold': True, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFFF00', 'border': 1, 'text_wrap': True})
        f_cell = wb.add_format({'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'border': 1})
        f_num = wb.add_format({'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'num_format': '0'})
        f_bold_left = wb.add_format({'font_size': 9, 'bold': True, 'align': 'left'})

        # Main WCC (Following A6 Pattern)
        ws_main = wb.add_worksheet('Main WCC')
        ws_main.merge_range('B2:H3', 'Work Completion Certificate', f_title)
        ws_main.write('B5', 'State', f_bold_left)
        ws_main.write('C5', 'RAJASTHAN', f_cell)
        ws_main.write('B20', 'No of Sites', f_bold_left)
        ws_main.write('C20', f"{len(rows)} SITES", f_cell)
        ws_main.write('B7', 'Project Type', f_bold_left)
        ws_main.write('C7', '93 K', f_cell)

        # WCC (A6+B6 Layout)
        ws_wcc = wb.add_worksheet('WCC')
        ws_wcc.merge_range('C3:P4', 'Work Completion Certificate', f_title)
        cert_text = f"This is to certify that below sites pertaining to WO/WCO No.P14/630330726 respect of Digitcom has been successfully completed."
        ws_wcc.merge_range('C6:P6', cert_text, f_cert)

        h1 = ['Sr. No', 'FT ENB SAP ID', 'FT PMP SAP ID', 'FT GIS SECTOR_ID', 'FB-FT HOP ID', 'No of Sectors', 'Tower type ', 'JC', 'WH', 'VEHICLE NO', 'MIN  NO', 'MIN Date', 'Completion Date ', 'REMARKS']
        h2 = ['ACTUAL KM', 'KM IN WO', 'GAP', 'USED KM IN WCC']
        
        for c, h in enumerate(h1):
            ws_wcc.write(8, 2+c, h, f_head_blue)
            ws_wcc.set_column(2+c, 2+c, 15)
        for c, h in enumerate(h2):
            ws_wcc.write(8, 2+len(h1)+2+c, h, f_head_yellow)
            ws_wcc.set_column(2+len(h1)+2+c, 2+len(h1)+2+c, 12)

        r_idx = 9
        for i, data in enumerate(rows):
            act_km = safe_float(data.get('AKTBC(FT)'))
            wo_km = safe_float(data.get('KM IN WO'))
            gap = act_km - wo_km
            used_km = act_km if gap < 0 else wo_km
            
            vals = [
                i+1, data.get('eNBsiteID'), data.get('PMP ID'), data.get('SEC ID'), data.get('FB-FT HOP ID'),
                safe_float(data.get('NO OF SECTOR')), data.get('TOWER'), data.get('JC'), data.get('WAREHOUSE'),
                data.get('VEHICLE NO'), data.get('MIN NO'), format_date(data.get('MIN DATE')),
                format_date(data.get('RFS DATE')), "RFS DONE"
            ]
            for c, v in enumerate(vals):
                ws_wcc.write(r_idx, 2+c, v, f_cell)
            
            ws_wcc.write(r_idx, 18, act_km, f_num)
            ws_wcc.write(r_idx, 19, wo_km, f_num)
            ws_wcc.write(r_idx, 20, gap, f_num)
            ws_wcc.write(r_idx, 21, used_km, f_num)
            r_idx += 1

    print(f"Successfully generated verification file at {output_path}")

except Exception as e:
    import traceback
    print(f"Error: {e}")
    traceback.print_exc()
