import openpyxl

generated_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/Billing/DC0111_A6B6_Billing_Test_Full.xlsx'

try:
    wb = openpyxl.load_workbook(generated_path, data_only=True)
    
    for sheet_name in ['Main WCC', 'WCC']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n--- {sheet_name} Content ---")
            for row in ws.iter_rows(max_row=12, values_only=True):
                # Filter out rows that are entirely None
                if any(x is not None for x in row):
                    print(row)
        else:
            print(f"Sheet {sheet_name} not found")

except Exception as e:
    print(f"Error: {e}")
