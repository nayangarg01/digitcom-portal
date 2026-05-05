import openpyxl
from copy import copy

def copy_sheet(src_ws, dst_ws):
    # Copy values and styles
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.alignment = copy(cell.alignment)
    
    # Copy merged cells
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))
        
    # Copy column dimensions
    for col, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col] = copy(dim)
        
    # Copy row dimensions
    for row, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row] = copy(dim)

def test_copy():
    ref_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/FinaliseBillingFormat/DIGITCOM_ AIRFIBER_DC0105_ JDPR_29-JAN-26_A6 (REJECT) 2.xlsx'
    output_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/scratch/test_template_out.xlsx'
    
    wb_ref = openpyxl.load_workbook(ref_path)
    src_ws = wb_ref['Main WCC']
    
    wb_new = openpyxl.Workbook()
    dst_ws = wb_new.active
    dst_ws.title = 'Main WCC'
    
    copy_sheet(src_ws, dst_ws)
    
    # Update fields
    dst_ws['D32'] = '22 SITES'
    dst_ws['I32'] = 'TEST DATE RANGE'
    dst_ws['D29'] = 'TEST WO'
    
    wb_new.save(output_path)
    print(f"Test output saved to {output_path}")

if __name__ == "__main__":
    test_copy()
