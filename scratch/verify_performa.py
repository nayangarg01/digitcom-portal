import openpyxl
import os

def verify_output(file_path):
    print(f"\n--- Verifying Output: {file_path} ---")
    if not os.path.exists(file_path):
        print("File not found!")
        return
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print("Sheets:", wb.sheetnames)
    
    # Check Sheet "1" (Detail)
    ws1 = wb["1"]
    print(f"\nSheet '1' - Row 1 Headers:")
    headers = [str(ws1.cell(row=1, column=c).value) for c in range(1, 15)]
    print(headers)
    
    print(f"\nSheet '1' - Row 2 Data:")
    row2 = [str(ws1.cell(row=2, column=c).value) for c in range(1, 15)]
    print(row2)
    
    # Check Sheet "Sheet1" (Summary)
    ws_sum = wb["Sheet1"]
    print(f"\nSheet 'Sheet1' - Row 2 Title:")
    print(ws_sum.cell(row=2, column=2).value)
    
    print(f"\nSheet 'Sheet1' - Row 3 Headers:")
    sum_headers = [str(ws_sum.cell(row=3, column=c).value) for c in range(2, 7)]
    print(sum_headers)
    
    print(f"\nSheet 'Sheet1' - Last row:")
    last_row = ws_sum.max_row
    last_vals = [str(ws_sum.cell(row=last_row, column=c).value) for c in range(2, 7)]
    print(last_vals)

output_file = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/scratch/test_performa_output.xlsx'
verify_output(output_file)
