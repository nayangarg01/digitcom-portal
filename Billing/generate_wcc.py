import pandas as pd
import openpyxl
import sys
import argparse
import os
from copy import copy
import datetime
from openpyxl.utils import get_column_letter

def load_master_data(master_path, target_billing_file):
    """Loads the KM SHEET from Master DPR and filters strictly for the target DC code."""
    try:
        df_master = pd.read_excel(master_path, sheet_name='KM SHEET', header=1)
    except Exception as e:
        print(f"Error loading master file {master_path}: {e}")
        return None

    df_master.columns = df_master.columns.astype(str).str.strip()
    
    # Identify BILLING FILE column safely
    bill_col = next((c for c in df_master.columns if 'BILLING FILE' in c.upper()), None)
    if not bill_col:
        print("Could not find 'BILLING FILE' column in master sheet.")
        return None
        
    df_filtered = df_master[df_master[bill_col].astype(str).str.strip().str.upper() == target_billing_file.upper()].copy()
    if df_filtered.empty:
        print(f"Warning: No sites found mapped to {target_billing_file}!")
        return None

    return df_filtered

def generate_wcc_sheet(df_sites, template_path, output_path):
    """Injects the filtered sites specifically into the WCC Template sheet."""
    try:
        wb = openpyxl.load_workbook(template_path)
        if 'WCC' not in wb.sheetnames:
            print("Error: WCC sheet not found in template.")
            return False
        ws = wb['WCC']
    except Exception as e:
        print(f"Error loading template {template_path}: {e}")
        return False
        
    # Target exactly the explicit string the user confirmed
    aktbc_col = next((c for c in df_sites.columns if 'CHRG EXTRA TRANSPORT' in str(c).upper() or 'AKTBC' == str(c).strip().upper()), None)
    
    # Locate headers in WCC
    header_row_idx = None
    cols_map = {}
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and 'GIS SECTOR_ID' in str(row):
            header_row_idx = r_idx
            for c_idx, val in enumerate(row, start=1):
                if val:
                    cols_map[str(val).strip()] = c_idx
            break

    if not header_row_idx:
        print("Error: Could not locate visual headers in the WCC template sheet.")
        return False

    start_row = header_row_idx + 1
    
    # Pre-calculate base styling
    base_styles = {}
    for c_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=start_row, column=c_idx)
        base_styles[c_idx] = {
            'font': copy(cell.font),
            'border': copy(cell.border),
            'fill': copy(cell.fill),
            'number_format': cell.number_format,
            'alignment': copy(cell.alignment)
        }

    # Management of rows
    if len(df_sites) > 22:
        ws.insert_rows(start_row + 22, amount=(len(df_sites) - 22))
    elif len(df_sites) < 22:
        amount_to_delete = 22 - len(df_sites)
        ws.delete_rows(start_row + len(df_sites), amount=amount_to_delete)
    
    # Map and Write rows
    for i, (_, row) in enumerate(df_sites.iterrows()):
        curr_row = start_row + i
        
        def get_val(matcher):
            c_name = next((c for c in df_sites.columns if matcher.upper() in c.upper()), None)
            return row[c_name] if c_name else ""
            
        def get_exact(name):
            return row[name] if name in df_sites.columns else ""

        mapping = [
            ('Sr. No', i + 1),
            ('ENB SITE ID', get_val('ENBSITEID') or get_val('SAP ID') or get_exact('Unnamed: 0')),
            ('PMP SAP ID', get_val('PMP ID')),
            ('GIS SECTOR_ID', get_val('GIS SECTOR')),
            ('No of Sectors', get_val('NO OF SECTOR')),
            ('Tower type', get_val('Tower type')),
            ('JC', get_val('JC')),
            ('WH', get_val('WH')),
            ('VEHICLE NO', get_val('VEHICLE NO')),
            ('MIN  NO', get_val('MIN NO')),
            ('MIN Date', get_val('MIN DATE')),
            ('Completion Date', get_val('Completion Date')),
            ('REMARKS', "RFS DONE" if pd.notna(get_val('Completion Date')) and get_val('Completion Date') != "" else ""),
            ('ACTUAL KM', float(row[aktbc_col]) if aktbc_col and pd.notna(row[aktbc_col]) else 0.0),
            ('KM IN WO', float(get_val('KM IN WO')) if pd.notna(get_val('KM IN WO')) and get_val('KM IN WO') != "" else 0.0)
        ]

        # Calculate Gap and Used KM
        actual_km = next((val for k, val in mapping if k == 'ACTUAL KM'), 0.0)
        km_in_wo = next((val for k, val in mapping if k == 'KM IN WO'), 0.0)
        mapping.append(('GAP', actual_km - km_in_wo))
        mapping.append(('USED KM IN WCC', actual_km if actual_km <= km_in_wo else km_in_wo))

        # Styling and value injection
        for c_idx in range(1, ws.max_column + 1):
            c = ws.cell(row=curr_row, column=c_idx)
            styles = base_styles.get(c_idx)
            if styles:
                c.font = copy(styles['font'])
                c.border = copy(styles['border'])
                c.fill = copy(styles['fill'])
                c.number_format = styles['number_format']
                c.alignment = copy(styles['alignment'])

        for col_name, val in mapping:
            c_target = next((cols_map[k] for k in cols_map if col_name.upper().strip() in k.upper().strip()), None)
            if c_target:
                target_cell = ws.cell(row=curr_row, column=c_target)
                if isinstance(val, pd.Timestamp):
                    target_cell.value = val.to_pydatetime()
                else:
                    target_cell.value = val

    # Formulas
    last_row = start_row + len(df_sites) - 1
    summary_row = last_row + 1
    for sum_col in ['ACTUAL KM', 'USED KM IN WCC']:
        c_idx = next((cols_map[k] for k in cols_map if sum_col.upper().strip() in k.upper().strip()), None)
        if c_idx:
            col_let = get_column_letter(c_idx)
            ws.cell(row=summary_row, column=c_idx).value = f"=SUM({col_let}{start_row}:{col_let}{last_row})"

    try:
        wb.save(output_path)
        print(f"WCC successfully generated: {output_path}")
        return True
    except Exception as e:
        print(f"Error saving WCC file: {e}")
        return False

def main():
    parser = argparse.ArgumentParser(description="Generate JIO WCC Billing Sheet.")
    parser.add_argument("master_file", help="Path to the MASTERDPR.xlsx file")
    parser.add_argument("billing_target", help="Target Billing File code (e.g., DC0105)")
    args = parser.parse_args()

    template_path = 'Billing/DC0105.xlsx' 
    output_path = f'Billing/{args.billing_target.upper()}_WCC_Automated.xlsx'

    print(f"--- Starting Standalone WCC Generation for {args.billing_target.upper()} ---")
    df_sites = load_master_data(args.master_file, args.billing_target)
    
    if df_sites is not None:
        print(f"Loaded {len(df_sites)} sites. Processing...")
        generate_wcc_sheet(df_sites, template_path, output_path)
    else:
        print("Generation aborted due to missing data.")

if __name__ == '__main__':
    main()
