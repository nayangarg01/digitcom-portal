import openpyxl

file_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/InputBilling/MASTER TRACKER FOR BILLING-AIRFIBER-RJST (1).xlsx'
try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    if 'KM SHEET' in wb.sheetnames:
        ws = wb['KM SHEET']
        print("\nRow 1 (SAP Codes) of 'KM SHEET':")
        row1 = next(ws.iter_rows(max_row=1, values_only=True))
        print(row1)
        print("\nRow 2 (Descriptions) of 'KM SHEET':")
        row2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        print(row2)
    else:
        print("Sheet 'KM SHEET' not found")
except Exception as e:
    print(f"Error: {e}")
