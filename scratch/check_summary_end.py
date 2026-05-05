import openpyxl

file_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/PerformaInvoiceWork/Performa_Test_v2.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws_sum = wb['Summary sheet1']
    print("\n--- Summary Sheet Last Rows ---")
    for row in ws_sum.iter_rows(min_row=ws_sum.max_row - 10, values_only=True):
        print(row)
except Exception as e:
    print(f"Error: {e}")
