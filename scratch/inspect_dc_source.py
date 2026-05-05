import openpyxl

file_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/PerformaInvoiceWork/DIGITCOM_ AIRFIBER_DC083_ JDPR_31-OCT-25.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    print(f"Sheets in DC083: {wb.sheetnames}")
    
    if 'Main WCC' in wb.sheetnames:
        ws = wb['Main WCC']
        print("\n--- Main WCC Row 1-10 ---")
        for row in ws.iter_rows(max_row=10, values_only=True):
            print(row)
            
    if 'WCC' in wb.sheetnames:
        ws = wb['WCC']
        print("\n--- WCC Headers ---")
        print(next(ws.iter_rows(max_row=1, values_only=True)))
        
    if 'JMS' in wb.sheetnames:
        ws = wb['JMS']
        print("\n--- JMS Header Area ---")
        for row in ws.iter_rows(max_row=15, values_only=True):
            print(row)
except Exception as e:
    print(f"Error: {e}")
