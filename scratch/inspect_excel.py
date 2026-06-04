import openpyxl

def inspect(file_path):
    print(f"\n=== Inspecting {file_path} ===")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if 'JMS' in wb.sheetnames:
        ws = wb['JMS']
        for r in range(1, 15):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, 8)]
            print(f"Row {r}: {row_vals}")
    else:
        print("JMS sheet not found")

inspect("../Backend_Portal/uploads/DC0122_Clean_Billing.xlsx")
inspect("../Backend_Portal/uploads/DC0128_Clean_Billing.xlsx")
