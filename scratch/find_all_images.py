from openpyxl import load_workbook

ref_file = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/FinaliseBillingFormat/DIGITCOM_ AIRFIBER_DC0105_ JDPR_29-JAN-26_A6 (REJECT) 2.xlsx'

try:
    wb = load_workbook(ref_file)
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        if hasattr(ws, '_images') and len(ws._images) > 0:
            print(f"Sheet '{sheetname}' has {len(ws._images)} images")
            for img in ws._images:
                print(f"  Anchor: {img.anchor}")
        else:
            print(f"Sheet '{sheetname}' has no images")
except Exception as e:
    print(f"Error: {e}")
