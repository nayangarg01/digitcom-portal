import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.cell.cell import MergedCell
import sys
import argparse
import os
from copy import copy

def safe_float(val):
    if pd.isna(val) or str(val).strip() == "" or str(val).strip().upper() in ["NR", "NA", "N.A", "-"]:
        return 0.0
    try:
        return float(val)
    except:
        return 0.0

def load_master_data(master_path, dc_number):
    try:
        df_full = pd.read_excel(master_path, header=None)
        if df_full.empty: return None, None
        
        # Discover DC column
        dc_col_idx = 0
        raw_row1 = df_full.iloc[1].tolist() if len(df_full) > 1 else []
        for i, h in enumerate(raw_row1):
            if "BILLING FILE" in str(h).upper() or "DC NUMBER" in str(h).upper():
                dc_col_idx = i; break

        df_sites = df_full[df_full.iloc[:, dc_col_idx].astype(str).str.strip().str.upper() == dc_number.upper()].copy()
        if df_sites.empty: return None, None
            
        df_sites.columns = [str(h).strip() for h in raw_row1]
        
        # build mapping
        code_to_col_idx = {}
        row0 = df_full.iloc[0].tolist()
        row1 = df_full.iloc[1].tolist()
        for i in range(len(row0)):
            sap = str(row0[i]).split('.')[0].strip()
            if sap and sap != 'nan': code_to_col_idx[sap] = i
            desc = str(row1[i]).strip()
            if desc and desc != 'nan': code_to_col_idx[desc] = i
                
        return df_sites, code_to_col_idx
    except Exception as e:
        print(f"Error: {e}"); return None, None

def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = copy(src_cell.number_format)
        dst_cell.alignment = copy(src_cell.alignment)

def generate_wcc_sheet(df_sites, wb):
    if 'WCC' not in wb.sheetnames: return False
    ws = wb['WCC']
    header_row_idx = None
    cols_map = {}
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and 'GIS SECTOR_ID' in str(row):
            header_row_idx = r_idx
            for c_idx, val in enumerate(row, start=1):
                if val: cols_map[str(val).strip()] = c_idx
            break
    if not header_row_idx: return False

    start_row = header_row_idx + 1
    num_sites = len(df_sites)
    if num_sites > 22: ws.insert_rows(start_row + 22, amount=(num_sites - 22))
    elif num_sites < 22: ws.delete_rows(start_row + num_sites, amount=(22 - num_sites))

    aktbc_col = next((c for c in df_sites.columns if 'CHRG EXTRA TRANSPORT' in c.upper() or 'AKTBC' == c.upper()), None)
    
    # Grab style from original template row (usually start_row)
    base_styles = {}
    for c_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=start_row, column=c_idx)
        base_styles[c_idx] = {'font': copy(cell.font), 'border': copy(cell.border), 'fill': copy(cell.fill), 'number_format': cell.number_format, 'alignment': copy(cell.alignment)}

    for i, (_, row) in enumerate(df_sites.iterrows()):
        curr_row = start_row + i
        def get_v(m):
            name = next((c for c in df_sites.columns if m.upper() in c.upper()), None)
            return row[name] if name else ""

        km_actual = safe_float(row[aktbc_col]) if aktbc_col else 0.0
        km_wo = safe_float(get_v('KM IN WO'))
        mapping = [
            ('Sr. No', i + 1), ('ENB SITE ID', get_v('ENBSITEID')), ('PMP SAP ID', get_v('PMP ID')),
            ('GIS SECTOR_ID', get_v('GIS SECTOR')), ('No of Sectors', get_v('NO OF SECTOR')),
            ('Tower type', get_v('Tower type')), ('JC', get_v('JC')), ('WH', get_v('WH')), 
            ('VEHICLE NO', get_v('VEHICLE NO')), ('MIN  NO', get_v('MIN NO')), ('MIN Date', get_v('MIN DATE')), 
            ('Completion Date', get_v('Completion Date')),
            ('ACTUAL KM', km_actual), ('KM IN WO', km_wo), ('GAP', km_actual - km_wo),
            ('USED KM IN WCC', km_actual if km_actual <= km_wo else km_wo)
        ]

        for c_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=curr_row, column=c_idx); s = base_styles.get(c_idx)
            if s: cell.font, cell.border, cell.fill, cell.number_format, cell.alignment = s['font'], s['border'], s['fill'], s['number_format'], s['alignment']

        for k, v in mapping:
            c_idx = next((cols_map[x] for x in cols_map if k.upper() in x.upper()), None)
            if c_idx:
                cell = ws.cell(row=curr_row, column=c_idx)
                cell.value = v.to_pydatetime() if isinstance(v, pd.Timestamp) else v

    last_row = start_row + num_sites - 1
    for k in ['ACTUAL KM', 'USED KM IN WCC']:
        c_idx = next((cols_map[x] for x in cols_map if k.upper() in x.upper()), None)
        if c_idx: ws.cell(row=last_row + 1, column=c_idx).value = f"=SUM({get_column_letter(c_idx)}{start_row}:{get_column_letter(c_idx)}{last_row})"
    return True

def populate_jms_surgical(df_sites, code_to_col_idx, wb):
    if 'JMS' not in wb.sheetnames: return False
    ws = wb['JMS']
    num_sites = len(df_sites)
    START_COL, C_ROW, S_ROW, T_ROW, SEC_ROW = 4, 11, 12, 13, 14
    MAX_DATA_ROW = 27 # PROTECTED BOUND: Do NOT touch beyond Row 27 for site data
    
    # 0. Safety Unmerge - ONLY DATA AREA
    for merged_range in list(ws.merged_cells.ranges):
        if 11 <= merged_range.min_row <= 27:
            ws.unmerge_cells(str(merged_range))

    # 1. Discover Total Qty Col
    orig_total_col = None
    for c in range(START_COL + 1, 60):
        if "TOTAL QUANTITY" in str(ws.cell(row=12, column=c).value).upper():
            orig_total_col = c; break
    
    # 2. Scale matrix OR Clear Ghost Sites
    num_template = 22 # default in template
    if orig_total_col: num_template = orig_total_col - START_COL

    if num_sites > num_template:
        ws.insert_cols(orig_total_col, num_sites - num_template)
    elif num_sites < num_template:
        # GHOST SITE CLEANUP: Clear unused template columns
        for c in range(START_COL + num_sites, START_COL + num_template):
            for r in range(11, MAX_DATA_ROW + 1 if 'MAX_DATA_ROW' in locals() else 35):
                cell = ws.cell(row=r, column=c)
                cell.value = None
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border()
    
    # 3. Dynamic Discovery of Footers & Columns
    total_qty_col, rate_col, amt_col = None, None, None
    for c in range(START_COL, 70):
        h = str(ws.cell(row=12, column=c).value).upper().strip()
        if "TOTAL QUANTITY" in h: total_qty_col = c
        elif "RATE AS PER SOW" in h: rate_col = c
        elif "AMOUNT" in h: amt_col = c
    
    if not total_qty_col: total_qty_col = START_COL + num_sites
    if not rate_col: rate_col = total_qty_col + 1
    if not amt_col: amt_col = rate_col + 1

    extra_visit_row, pole_mount_row, total_row = None, None, None
    for r in range(25, 45):
        # Scan across first few cols to find keyword
        row_vals = [str(ws.cell(row=r, column=cc).value).upper() for cc in range(1, 5)]
        text = " ".join(row_vals)
        if "TOTAL" in text: total_row = r
        if "EXTRA VISIT" in text: extra_visit_row = r
        if "POLE MOUNT" in text: pole_mount_row = r
        if total_row: break
    
    if not total_row: total_row = 28
    MAX_DATA_ROW = total_row - 1

    # 4. Blanket Zero-Initialization & Surgical Styling
    # We only zero out the columns that BELONG to the current sites
    for r in range(16, MAX_DATA_ROW + 1):
        # PROTECT STYLE DNA: Only mirror Column 4's style to the site area
        # Preserve original styles for A, B, C, Total, Rate, Amount
        src_cell = ws.cell(row=r, column=START_COL)

        for i in range(num_sites):
            curr_col = START_COL + i
            cell = ws.cell(row=r, column=curr_col)
            cell.value = 0.0
            cell.alignment = Alignment(horizontal='center', vertical='center')
            copy_cell_style(src_cell, cell)

    # 5. Inject Sites Data
    for i, (_, row) in enumerate(df_sites.iterrows()):
        curr_col = START_COL + i
        ws.cell(row=C_ROW, column=curr_col).value = i + 1
        cell_id = ws.cell(row=S_ROW, column=curr_col)
        cell_id.value = str(row['PMP ID']).strip()
        cell_id.font = Font(name='Verdana', size=35, bold=True)
        cell_id.alignment = Alignment(text_rotation=90, horizontal='center', vertical='center')
        ws.cell(row=T_ROW, column=curr_col).value = str(row['Tower type']).strip()
        ws.cell(row=SEC_ROW, column=curr_col).value = row['NO OF SECTOR']
        
        for r in range(16, MAX_DATA_ROW + 1):
            # Resolve code
            item_id = ws.cell(row=r, column=1).value
            try: code = str(int(item_id)) if item_id else ""
            except: code = str(item_id).strip() if item_id else ""
            
            # Use dynamic mapping for extra labels if code is empty in Column A
            if not code:
                if r == extra_visit_row: code = "EXTRA VISIT"
                elif r == pole_mount_row: code = "POLE MOUNT"
            
            if code in code_to_col_idx:
                val = row.iloc[code_to_col_idx[code]]
                ws.cell(row=r, column=curr_col).value = safe_float(val)

    # 6. Rates & Summations
    fcl, lcl = get_column_letter(START_COL), get_column_letter(START_COL + num_sites - 1)
    for r in range(16, MAX_DATA_ROW + 1):
        if r == extra_visit_row: ws.cell(row=r, column=rate_col).value = 1000
        elif r == pole_mount_row: ws.cell(row=r, column=rate_col).value = 500
        
        ws.cell(row=r, column=total_qty_col).value = f"=SUM({fcl}{r}:{lcl}{r})"
        ws.cell(row=r, column=amt_col).value = f"={get_column_letter(total_qty_col)}{r}*{get_column_letter(rate_col)}{r}"
    
    # 7. Grand Total Row (Dynamic)
    ws.cell(row=total_row, column=amt_col).value = f"=SUM({get_column_letter(amt_col)}16:{get_column_letter(amt_col)}{total_row-1})"
    
    # Update Sectors Total
    ws.cell(row=SEC_ROW, column=total_qty_col).value = f"=SUM({fcl}{SEC_ROW}:{lcl}{SEC_ROW})"
    return True

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("master_file")
    parser.add_argument("dc_number")
    parser.add_argument("template")
    parser.add_argument("output")
    args = parser.parse_args()

    print(f"--- Verification Run: {args.dc_number} ---")
    df_sites, code_to_col_idx = load_master_data(args.master_file, args.dc_number)
    if df_sites is not None:
        wb = openpyxl.load_workbook(args.template)
        # Swap DC labels everywhere
        for s in wb.sheetnames:
            ws = wb[s]
            for r in range(1, 10):
                for c in range(1, 25):
                    if "DC-CODE" in str(ws.cell(row=r, column=c).value):
                        ws.cell(row=r, column=c).value = str(ws.cell(row=r, column=c).value).replace("DC-CODE", args.dc_number)
        
        print("- Generating WCC...")
        generate_wcc_sheet(df_sites, wb)
        print("- Generating JMS (Surgical)...")
        populate_jms_surgical(df_sites, code_to_col_idx, wb)
        
        wb.save(args.output)
        print(f"VERIFICATION COMPLETE: {args.output}")
    else:
        print("Failed.")

if __name__ == "__main__":
    main()
