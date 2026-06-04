import openpyxl
import os

folder = "../a6b6_essentials"
files = [f"DC0{i}_Unified_Billing.xlsx" for i in range(128, 136)]

print(f"{'DC Number':<12} | {'Annexture-A6 PMPs':<20} | {'Annexture-B6 PMPs':<20} | {'Status':<15}")
print("-" * 75)

for f in files:
    path = os.path.join(folder, f)
    if not os.path.exists(path):
        print(f"{f:<12} | {'File missing':<20} | {'File missing':<20} | {'Error':<15}")
        continue
        
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        
        # A6 columns count
        a6_cols = []
        if 'Annexture-A6' in wb.sheetnames:
            ws_a6 = wb['Annexture-A6']
            # Read row 2 (PMP IDs)
            row2 = [cell for cell in next(ws_a6.iter_rows(min_row=2, max_row=2, values_only=True)) if cell is not None]
            # Exclude 'Row Labels' and 'Grand Total'
            a6_cols = [c for c in row2 if c not in ['Row Labels', 'Grand Total']]
            
        # B6 columns count
        b6_cols = []
        if 'Annexture-B6' in wb.sheetnames:
            ws_b6 = wb['Annexture-B6']
            row2_b = [cell for cell in next(ws_b6.iter_rows(min_row=2, max_row=2, values_only=True)) if cell is not None]
            b6_cols = [c for c in row2_b if c not in ['Row Labels', 'Grand Total']]
            
        dc_name = f.split("_")[0]
        status = "As Expected" if len(b6_cols) >= len(a6_cols) else "Check"
        print(f"{dc_name:<12} | {len(a6_cols):<20} | {len(b6_cols):<20} | {status:<15}")
        print(f"  - A6 PMPs ({len(a6_cols)}): {a6_cols}")
        print(f"  - B6 PMPs ({len(b6_cols)}): {b6_cols}")
        print("-" * 75)
        
    except Exception as e:
        print(f"{f:<12} | Error: {str(e)[:50]}")
