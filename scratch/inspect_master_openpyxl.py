import openpyxl

file_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/InputBilling/MASTER TRACKER FOR BILLING-AIRFIBER-RJST (1).xlsx'
try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    if 'A6+B6 Billings' in wb.sheetnames:
        ws = wb['A6+B6 Billings']
        print("\nFirst 10 rows of 'A6+B6 Billings':")
        for row in ws.iter_rows(max_row=10, values_only=True):
            print(row)
    else:
        print("Sheet 'A6+B6 Billings' not found")
except Exception as e:
    print(f"Error: {e}")
