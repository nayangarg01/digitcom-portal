import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def combine_dpr_min():
    dpr_path = 'RoutingSampleFiles/DPR TRACKER RJ-13 March.xlsx'
    min_path = 'RoutingSampleFiles/MIN DUMP-RJST TILL 31 DEC  25.xlsx'
    output_path = 'RoutingSampleFiles/Combined_Routing_Data.xlsx'
    
    print("Loading DPR Data (AIR FIBER)...")
    dpr_df = pd.read_excel(dpr_path, sheet_name='AIR FIBER')
    
    # Specific columns from DPR
    dpr_cols = {
        'SITE ID': 'SITE ID',
        'NO OF\n SECTOR': 'no_of_sector',
        'JC NAME': 'jc_name',
        'CMP': 'cmp',
        'Team': 'team',
        'LATITUDE': 'latitude',
        'LONGITUDE': 'longitude',
        'ACTIVITY': 'activity',
        'ALLOTMENT \nLOT': 'allotment_lot',
        'INSTALLATION \nDATE': 'installation_date',
        'INTEGRATION\nDATE': 'integration_date',
        'POWER CABLE LENGTH': 'power_cable_length',
        'CPRI LENGTH': 'cpri_length'
    }
    
    # Keep only existing columns and rename
    dpr_clean = dpr_df[list(dpr_cols.keys())].rename(columns=dpr_cols)
    dpr_clean['SITE ID'] = dpr_clean['SITE ID'].astype(str).str.strip()
    dpr_clean = dpr_clean.drop_duplicates(subset=['SITE ID'])
    
    print("Loading MIN Data (A6 DUMP)...")
    min_df = pd.read_excel(min_path, sheet_name='A6 DUMP')
    
    # Specific columns from MIN
    min_cols = {
        'ENB ID': 'enb_id',
        'Site ID': 'pmp_id', # User said Site ID in MIN is PMP ID
        'WBS ID': 'wbs_id',
        'DWG': 'dwg',
        'Work Order': 'work_order',
        'MIN Number': 'min_number',
        'Date': 'min_date',
        'Sloc': 'sloc'
    }
    
    min_subset = min_df[list(min_cols.keys())].rename(columns=min_cols)
    min_subset['enb_id'] = min_subset['enb_id'].astype(str).str.strip()
    
    print("Aggregating MIN data...")
    # Aggregate MIN Number: comma-separated list
    # For others, take the first one
    agg_funcs = {
        'pmp_id': 'first',
        'wbs_id': 'first',
        'dwg': 'first',
        'work_order': 'first',
        'min_number': lambda x: ', '.join(x.dropna().astype(str).unique()),
        'min_date': 'first',
        'sloc': 'first'
    }
    
    min_agg = min_subset.groupby('enb_id').agg(agg_funcs).reset_index()
    
    print(f"Total DPR Sites (AIR FIBER): {len(dpr_clean)}")
    print(f"Total Unique MIN IDs (A6 DUMP): {len(min_agg)}")
    
    # Merge on SITE ID = enb_id
    combined = pd.merge(dpr_clean, min_agg, left_on='SITE ID', right_on='enb_id', how='inner')
    combined = combined.drop(columns=['enb_id'])
    
    print(f"Combined Sites: {len(combined)}")
    
    # Format dates as strings to avoid display issues in different apps (like Numbers)
    date_cols = ['installation_date', 'integration_date', 'min_date']
    for col in date_cols:
        if col in combined.columns:
            combined[col] = pd.to_datetime(combined[col], errors='coerce').dt.strftime('%Y-%m-%d')
            # Handle cases where value was originally 'NR' or NaN
            combined[col] = combined[col].fillna('')
    
    # Save to Excel
    combined.to_excel(output_path, index=False)
    print(f"Saved combined data to {output_path}")
    
    # Apply Formatting using openpyxl
    print("Applying formatting...")
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="006100", end_color="006100", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    alignment = Alignment(horizontal="center", vertical="center")
    
    # Formatting for all cells (including Alignment)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = alignment
            cell.border = border

    # Header formatting (override with bold/fill)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        
    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        
    # Freeze top row
    ws.freeze_panes = 'A2'
    
    wb.save(output_path)
    print(f"Formatted Excel saved to {output_path}")

if __name__ == "__main__":
    combine_dpr_min()
