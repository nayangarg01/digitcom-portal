import pandas as pd
import openpyxl
import sys
import argparse
import os
from copy import copy
from openpyxl.utils import get_column_letter

def load_master_data(master_path, target_billing_file):
    """Loads the KM SHEET from Master DPR and filters strictly for the target DC code."""
    try:
        df_master = pd.read_excel(master_path, sheet_name='KM SHEET', header=1)
    except Exception as e:
        print(f"Error loading master file {master_path}: {e}")
        return None

    df_master.columns = df_master.columns.astype(str).str.strip()
    
    # Identify BILLING FILE column safely
    bill_col = next((c for c in df_master.columns if 'BILLING FILE' in c.upper()), None)
    if not bill_col:
        print("Could not find 'BILLING FILE' column in master sheet.")
        return None
        
    df_filtered = df_master[df_master[bill_col].astype(str).str.strip().str.upper() == target_billing_file.upper()].copy()
    if df_filtered.empty:
        print(f"Warning: No sites found mapped to {target_billing_file}!")
        return None

    return df_filtered

def generate_jms_sheet(df_sites, template_path, output_path):
    """Injects the filtered sites into the horizontal matrix JMS Template sheet."""
    try:
        wb = openpyxl.load_workbook(template_path)
        if 'JMS' not in wb.sheetnames:
            print("Error: JMS sheet not found in template.")
            return False
        ws = wb['JMS']
    except Exception as e:
        print(f"Error loading template {template_path}: {e}")
        return False
        
    # 1. Identify key horizontal and vertical tracking indices
    # Row 10: Site IDs (Col 4 onwards)
    SITE_ROW = 10
    SITE_TYPE_ROW = 11
    SECTOR_ROW = 12
    START_COL = 4 # 'D'
    
    # Site mapping
    towers = [str(t).strip() for t in df_sites['Tower type']] if 'Tower type' in df_sites.columns else ["GBT"]*len(df_sites)
    sectors = [v for v in df_sites['NO OF SECTOR']] if 'NO OF SECTOR' in df_sites.columns else [1]*len(df_sites)
    site_ids = [str(s).strip() for s in (df_sites['PMP ID'] if 'PMP ID' in df_sites.columns else df_sites.iloc[:, 0])]

    # Column Management
    if len(df_sites) > 22:
        ws.insert_cols(START_COL + 22, amount=(len(df_sites) - 22))
    elif len(df_sites) < 22:
        amount_to_delete = 22 - len(df_sites)
        ws.delete_cols(START_COL + len(df_sites), amount=amount_to_delete)

    # Pre-copy styles from Column D (Start column)
    col_styles = {}
    for r_idx in range(1, ws.max_row + 1):
        cell = ws.cell(row=r_idx, column=START_COL)
        col_styles[r_idx] = {
            'font': copy(cell.font), 'border': copy(cell.border), 'fill': copy(cell.fill),
            'number_format': cell.number_format, 'alignment': copy(cell.alignment)
        }

    # Write Site Headers and Apply styles across columns
    for i in range(len(df_sites)):
        curr_col = START_COL + i
        # Write headers
        ws.cell(row=SITE_ROW, column=curr_col).value = site_ids[i]
        ws.cell(row=SITE_TYPE_ROW, column=curr_col).value = towers[i]
        ws.cell(row=SECTOR_ROW, column=curr_col).value = sectors[i]
        
        # Apply style to whole column
        for r_idx, s in col_styles.items():
            if r_idx < 10: continue # Skip top header fluff
            c = ws.cell(row=r_idx, column=curr_col)
            c.font, c.border, c.fill = copy(s['font']), copy(s['border']), copy(s['fill'])
            c.number_format, c.alignment = s['number_format'], copy(s['alignment'])

    # Item Quantity Injection
    # Scan Column B for descriptions and match with df_sites columns
    for r_idx in range(14, ws.max_row + 1):
        desc = str(ws.cell(row=r_idx, column=2).value).strip() if ws.cell(row=r_idx, column=2).value else ""
        if not desc or "TOTAL" in desc.upper(): continue
        
        # Fuzzy match description with MASTERDPR columns
        match_col = next((c for c in df_sites.columns if desc.upper() in str(c).upper() or str(c).upper() in desc.upper()), None)
        if match_col:
            quantities = df_sites[match_col].values
            for i, qty in enumerate(quantities):
                ws.cell(row=r_idx, column=START_COL + i).value = float(qty) if pd.notna(qty) else 0.0

    # Summary Column Logic (Total Qty, Amount)
    summary_start_col = None
    for c_idx in range(START_COL + len(df_sites), ws.max_column + 1):
        if "Total Quantity" in str(ws.cell(row=SITE_TYPE_ROW + 1, column=c_idx).value):
            summary_start_col = c_idx
            break
            
    if summary_start_col:
        qty_col = summary_start_col
        rate_col = qty_col + 1
        amt_col = qty_col + 2
        
        first_let = get_column_letter(START_COL)
        last_let = get_column_letter(START_COL + len(df_sites) - 1)
        
        for r_idx in range(16, ws.max_row):
            desc_val = ws.cell(row=r_idx, column=2).value
            if not desc_val or "TOTAL" in str(desc_val).upper():
                if "TOTAL" in str(desc_val).upper():
                    # Update Grand Total
                    sum_formula = f"=SUM({get_column_letter(amt_col)}16:{get_column_letter(amt_col)}{r_idx-1})"
                    ws.cell(row=r_idx, column=amt_col).value = sum_formula
                continue
                
            # Total Quantity = SUM(Sites)
            ws.cell(row=r_idx, column=qty_col).value = f"=SUM({first_let}{r_idx}:{last_let}{r_idx})"
            # Amount = Qty * Rate
            ws.cell(row=r_idx, column=amt_col).value = f"={get_column_letter(qty_col)}{r_idx}*{get_column_letter(rate_col)}{r_idx}"

    try:
        wb.save(output_path)
        print(f"JMS successfully generated: {output_path}")
        return True
    except Exception as e:
        print(f"Error saving JMS file: {e}")
        return False

def main():
    parser = argparse.ArgumentParser(description="Generate JIO JMS Billing Sheet.")
    parser.add_argument("master_file", help="Path to the MASTERDPR.xlsx file")
    parser.add_argument("billing_target", help="Target Billing File code (e.g., DC0105)")
    parser.add_argument("template_path", help="Path to the DC0105_TEMPLATE.xlsx file")
    parser.add_argument("output_path", help="Path where the generated file should be saved")
    args = parser.parse_args()

    print(f"--- Starting Standalone JMS Generation for {args.billing_target.upper()} ---")
    df_sites = load_master_data(args.master_file, args.billing_target)
    
    if df_sites is not None:
        print(f"Loaded {len(df_sites)} sites. Processing...")
        generate_jms_sheet(df_sites, args.template_path, args.output_path)
    else:
        print("Generation aborted due to missing data.")

if __name__ == '__main__':
    main()
