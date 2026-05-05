import openpyxl

generated_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/Billing/DC0111_A6B6_Billing_Final_Test_v6.xlsx'

try:
    wb = openpyxl.load_workbook(generated_path, data_only=True)
    if 'Annexture-B6' in wb.sheetnames:
        ws = wb['Annexture-B6']
        print("\n--- Annexture-B6 Headers (Row 2) ---")
        row2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        print(row2)
        print("\n--- Annexture-B6 Sample Row ---")
        row3 = next(ws.iter_rows(min_row=3, max_row=3, values_only=True))
        print(row3)
except Exception as e:
    print(f"Error: {e}")
