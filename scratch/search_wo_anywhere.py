import openpyxl

file_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/PerformaInvoiceWork/48/DIGITCOM_ AIRFIBER_DC087_JPUR_28-OCT-25& 31-OCT-25_A6.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['Main WCC']
    found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "W.O.Number" in str(cell.value):
                # WO is usually 2 cells to the right
                target_cell = ws.cell(row=cell.row, column=cell.column + 2)
                print(f"Found at {cell.coordinate}: Label='{cell.value}', Value='{target_cell.value}'")
                found = True
                break
        if found: break
except Exception as e:
    print(f"Error: {e}")
