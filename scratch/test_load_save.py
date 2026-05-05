import openpyxl
import os
from openpyxl.drawing.image import Image as OpenpyxlImage

def test_load_save_logic(output_path, ref_path, logo_path):
    print(f"- Loading template: {os.path.basename(ref_path)}")
    wb = openpyxl.load_workbook(ref_path)
    ws = wb['Main WCC']
    
    # Update fields
    ws['D32'] = "22 SITES"
    ws['I32'] = "30-JAN-26 TO 04-FEB-26"
    ws['D29'] = "630330726"
    
    # Add logo
    if os.path.exists(logo_path):
        try:
            img = OpenpyxlImage(logo_path)
            img.width, img.height = 120, 100
            ws.add_image(img, 'B2')
            print("- Logo added")
        except Exception as e:
            print(f"- Logo failed: {e}")
            
    # Save directly
    wb.save(output_path)
    print(f"- Saved to {output_path}")

if __name__ == "__main__":
    ref = "/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/FinaliseBillingFormat/DIGITCOM_ AIRFIBER_DC0105_ JDPR_29-JAN-26_A6 (REJECT) 2.xlsx"
    logo = "/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/FinaliseBillingFormat/Screenshot 2026-05-04 at 6.45.10 PM.png"
    output = "/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/FinaliseBillingFormat/DC0105_TEMPLATE_LOAD_TEST.xlsx"
    test_load_save_logic(output, ref, logo)
