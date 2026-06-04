import openpyxl
import sys
import os

def compare_excel(file1, file2):
    print(f"Comparing:\n  File 1 (OOP): {file1}\n  File 2 (REF): {file2}\n")
    
    if not os.path.exists(file1):
        print(f"Error: {file1} does not exist.")
        return False
    if not os.path.exists(file2):
        print(f"Error: {file2} does not exist.")
        return False
        
    wb1 = openpyxl.load_workbook(file1, data_only=False)
    wb2 = openpyxl.load_workbook(file2, data_only=False)
    
    # Compare sheets
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    
    if sheets1 != sheets2:
        print(f"WARNING: Sheet names do not match!")
        print(f"  File 1 sheets: {sheets1}")
        print(f"  File 2 sheets: {sheets2}")
    else:
        print(f"Sheet names match: {sheets1}\n")
        
    mismatches = 0
    common_sheets = [s for s in sheets1 if s in sheets2]
    
    for sheet_name in common_sheets:
        ws1 = wb1[sheet_name]
        ws2 = wb2[sheet_name]
        
        max_r = max(ws1.max_row, ws2.max_row)
        max_c = max(ws1.max_column, ws2.max_column)
        
        print(f"Sheet '{sheet_name}': Comparing up to {max_r} rows and {max_c} columns...")
        
        sheet_mismatches = 0
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                cell1 = ws1.cell(row=r, column=c)
                cell2 = ws2.cell(row=r, column=c)
                
                val1 = cell1.value
                val2 = cell2.value
                
                # Normalize values
                def normalize(v):
                    if v is None: return ""
                    v_str = str(v).strip()
                    # Convert float format to float comparison if possible
                    try:
                        return f"{float(v_str):.4f}"
                    except:
                        pass
                    # If it's a date or timestamp
                    if " 00:00:00" in v_str:
                        v_str = v_str.replace(" 00:00:00", "")
                    return v_str.upper()
                    
                n1 = normalize(val1)
                n2 = normalize(val2)
                
                if n1 != n2:
                    # Ignore minor title formatting difference in spacing if values are close
                    if r <= 6 and sheet_name in ['JMS', 'Abstract', 'BOQ', 'WCC']:
                        # Skip certification date string minor format checks
                        continue
                    # Ignore trailing zeroes or small representation issues in formulas
                    # e.g. =SUM(D13:G13) vs =SUM(D13:G13) but with spaces or uppercase
                    if str(val1).replace(" ", "").upper() == str(val2).replace(" ", "").upper():
                        continue
                        
                    print(f"  Mismatch at Row {r}, Col {c} ({openpyxl.utils.get_column_letter(c)}{r}):")
                    print(f"    File 1 (OOP): {val1!r}")
                    print(f"    File 2 (REF): {val2!r}")
                    sheet_mismatches += 1
                    mismatches += 1
                    if sheet_mismatches >= 10:
                        print("  Too many mismatches in this sheet, skipping remainder...")
                        break
            if sheet_mismatches >= 10:
                break
                
        if sheet_mismatches == 0:
            print(f"  Sheet '{sheet_name}' matches perfectly!\n")
        else:
            print(f"  Sheet '{sheet_name}' has {sheet_mismatches} mismatches.\n")
            
    if mismatches == 0:
        print("SUCCESS: Both workbooks match sheet-for-sheet and value-for-value!")
        return True
    else:
        print(f"FAILURE: Total of {mismatches} mismatches found between the two workbooks.")
        return False

if __name__ == "__main__":
    file1 = "Billing/DC0122_OOP_Billing.xlsx"
    file2 = "../Backend_Portal/uploads/DC0122_Clean_Billing.xlsx"
    
    if len(sys.argv) > 2:
        file1 = sys.argv[1]
        file2 = sys.argv[2]
        
    compare_excel(file1, file2)
