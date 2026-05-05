import openpyxl

file_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/InputBilling/DIGITCOM_AIRFIBER_DC0111_JPUR_12-MAR-26 23-MAR-26 & 25-MAR-26_A6+B6.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    if 'WCC' in wb.sheetnames:
        ws = wb['WCC']
        print("\nFirst 15 rows of 'WCC':")
        for row in ws.iter_rows(max_row=15, values_only=True):
            print(row)
    else:
        print("Sheet 'WCC' not found")
        
    if 'Main WCC' in wb.sheetnames:
        ws_main = wb['Main WCC']
        print("\nFirst 10 rows of 'Main WCC':")
        for row in ws_main.iter_rows(max_row=10, values_only=True):
            print(row)
except Exception as e:
    print(f"Error: {e}")
