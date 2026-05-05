import openpyxl
from datetime import datetime

master_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/InputBilling/MASTER TRACKER FOR BILLING-AIRFIBER-RJST (1).xlsx'
dc_number = 'DC0111'

def safe_float(val):
    if val is None or str(val).strip() == "" or str(val).strip().upper() in ["NR", "NA", "-"]:
        return 0.0
    try: return float(val)
    except: return 0.0

try:
    # Use openpyxl to avoid pandas/numpy issue
    wb_master = openpyxl.load_workbook(master_path, read_only=True, data_only=True)
    ws = wb_master['A6+B6 Billings']
    
    headers = [str(h).strip() for h in next(ws.iter_rows(max_row=1, values_only=True))]
    dc_col_idx = headers.index('BILLING FILE')
    
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[dc_col_idx] and str(row[dc_col_idx]).strip().upper() == dc_number:
            rows.append(row)
    
    print(f"Found {len(rows)} sites for {dc_number} in A6+B6 Billings")
    
    # Simulate WCC Mapping for first 3 sites
    print("\n--- WCC Preview for DC0111 (First 3 sites) ---")
    wcc_headers = ['Sr. No', 'FT ENB SAP ID', 'FT PMP SAP ID', 'FT GIS SECTOR_ID', 'FB-FT HOP ID', 'No of Sectors', 'Tower type ', 'JC', 'WH', 'VEHICLE NO', 'MIN  NO', 'MIN Date', 'Completion Date ', 'REMARKS', '|', 'ACTUAL KM', 'KM IN WO', 'GAP', 'USED KM']
    print("\t".join(wcc_headers))
    
    for i, row in enumerate(rows[:3]):
        data = dict(zip(headers, row))
        
        act_km = safe_float(data.get('AKTBC(FT)'))
        wo_km = safe_float(data.get('KM IN WO'))
        gap = act_km - wo_km
        used_km = act_km if gap < 0 else wo_km
        
        wcc_row = [
            i + 1,
            data.get('eNBsiteID'),
            data.get('PMP ID'),
            data.get('SEC ID'),
            data.get('FB-FT HOP ID'),
            data.get('NO OF SECTOR'),
            data.get('TOWER'),
            data.get('JC'),
            data.get('WAREHOUSE'),
            data.get('VEHICLE NO'),
            data.get('MIN NO'),
            data.get('MIN DATE'),
            data.get('RFS DATE'),
            "RFS DONE",
            "|",
            act_km,
            wo_km,
            gap,
            used_km
        ]
        print("\t".join([str(x) for x in wcc_row]))

except Exception as e:
    import traceback
    print(f"Error: {e}")
    traceback.print_exc()
