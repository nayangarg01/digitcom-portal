import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.cell.cell import MergedCell
import sys
import argparse
import os
from copy import copy

def safe_float(val):
    """Safely converts a value to float, handling 'NR', 'NA', or other text by returning 0.0."""
    if pd.isna(val) or str(val).strip() == "" or str(val).strip().upper() in ["NR", "NA", "N.A", "-"]:
        return 0.0
    try:
        return float(val)
    except:
        return 0.0

def load_master_data(master_path, dc_number):
    """Loads KM SHEET and builds a mapping of SAP Codes/Descriptions to column indices."""
    try:
        df_full = pd.read_excel(master_path, header=None)
        if df_full.empty: return None, None
            
        # Discover 'BILLING FILE' column (Usually in Row 2, index 1)
        dc_col_idx = 0
        found_header = False
        raw_row1 = df_full.iloc[1].tolist() if len(df_full) > 1 else []
        for i, h in enumerate(raw_row1):
            h_str = str(h).upper().strip()
            if "BILLING FILE" in h_str or "DC NUMBER" in h_str:
                dc_col_idx = i
                found_header = True
                break
        
        if not found_header:
            for c_idx in range(df_full.shape[1]):
                col_sample = df_full.iloc[:, c_idx].astype(str).str.upper().tolist()
                if dc_number.upper() in col_sample:
                    dc_col_idx = c_idx
                    break

        df_sites = df_full[df_full.iloc[:, dc_col_idx].astype(str).str.strip().str.upper() == dc_number.upper()].copy()
        if df_sites.empty: return None, None
            
        raw_headers = df_full.iloc[1].tolist()
        df_sites.columns = [str(h).strip() for h in raw_headers]
        
        # Build Code mapping (Row 0: SAP, Row 1: Description)
        code_to_col_idx = {}
        row0 = df_full.iloc[0].tolist()
        row1 = df_full.iloc[1].tolist()
        for i in range(len(row0)):
            sap = str(row0[i]).split('.')[0].strip()
            if sap and sap != 'nan': code_to_col_idx[sap] = i
            desc = str(row1[i]).strip()
            if desc and desc != 'nan': code_to_col_idx[desc] = i
                
        return df_sites, code_to_col_idx
    except Exception as e:
        print(f"Error loading Master: {e}")
        return None, None

def copy_cell_style(src_cell, dst_cell):
    """Clones all style properties from src to dst."""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = copy(src_cell.number_format)
        dst_cell.alignment = copy(src_cell.alignment)

def generate_wcc_sheet(df_sites, wb):
    """Populates the WCC sheet with site data and calculates KM totals."""
    if 'WCC' not in wb.sheetnames: return False
    ws = wb['WCC']
    
    # Identify headers and mapping
    header_row_idx = None
    cols_map = {}
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and 'GIS SECTOR_ID' in str(row):
            header_row_idx = r_idx
            for c_idx, val in enumerate(row, start=1):
                if val: cols_map[str(val).strip()] = c_idx
            break
            
    if not header_row_idx: return False

    start_row = header_row_idx + 1
    base_styles = {}
    for c_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=start_row, column=c_idx)
        base_styles[c_idx] = {'font': copy(cell.font), 'border': copy(cell.border), 'fill': copy(cell.fill), 'number_format': cell.number_format, 'alignment': copy(cell.alignment)}

    # Row scaling
    num_sites = len(df_sites)
    if num_sites > 22: ws.insert_rows(start_row + 22, amount=(num_sites - 22))
    elif num_sites < 22: ws.delete_rows(start_row + num_sites, amount=(22 - num_sites))
        
    aktbc_col = next((c for c in df_sites.columns if 'CHRG EXTRA TRANSPORT' in c.upper() or 'AKTBC' == c.upper()), None)

    for i, (_, row) in enumerate(df_sites.iterrows()):
        curr_row = start_row + i
        def get_val(matcher):
            c_name = next((c for c in df_sites.columns if matcher.upper() in c.upper()), None)
            return row[c_name] if c_name else ""
            
        km_actual = safe_float(row[aktbc_col]) if aktbc_col else 0.0
        km_wo = safe_float(get_val('KM IN WO'))
        
        mapping = [
            ('Sr. No', i + 1), ('ENB SITE ID', get_val('ENBSITEID')), ('PMP SAP ID', get_val('PMP ID')),
            ('GIS SECTOR_ID', get_val('GIS SECTOR')), ('No of Sectors', get_val('NO OF SECTOR')),
            ('Tower type', get_val('Tower type')), ('JC', get_val('JC')), ('WH', get_val('WH')), 
            ('VEHICLE NO', get_val('VEHICLE NO')), ('MIN  NO', get_val('MIN NO')), ('MIN Date', get_val('MIN DATE')), 
            ('Completion Date', get_val('Completion Date')),
            ('REMARKS', "RFS DONE" if pd.notna(get_val('Completion Date')) and str(get_val('Completion Date')) != "" else ""),
            ('ACTUAL KM', km_actual), ('KM IN WO', km_wo), ('GAP', km_actual - km_wo),
            ('USED KM IN WCC', km_actual if km_actual <= km_wo else km_wo)
        ]

        for c_idx in range(1, ws.max_column + 1):
            c = ws.cell(row=curr_row, column=c_idx)
            s = base_styles.get(c_idx)
            if s:
                c.font, c.border, c.fill, c.number_format, c.alignment = s['font'], s['border'], s['fill'], s['number_format'], s['alignment']

        for col_name, val in mapping:
            c_idx = next((cols_map[k] for k in cols_map if col_name.upper().strip() in k.upper().strip() or k.upper().strip() in col_name.upper().strip()), None)
            if c_idx:
                cell = ws.cell(row=curr_row, column=c_idx)
                cell.value = val.to_pydatetime() if isinstance(val, pd.Timestamp) else val

    # Formula updates
    last_row = start_row + num_sites - 1
    summary_row = last_row + 1
    for sum_col in ['ACTUAL KM', 'USED KM IN WCC']:
        c_idx = next((cols_map[k] for k in cols_map if sum_col.upper().strip() in k.upper().strip()), None)
        if c_idx:
            ws.cell(row=summary_row, column=c_idx).value = f"=SUM({get_column_letter(c_idx)}{start_row}:{get_column_letter(c_idx)}{last_row})"
    return True

def populate_main_matrix(sheet_name, df_sites, code_to_col_idx, wb):
    """Generates the horizontal site matrix (JMS / Matrix sheets)."""
    if sheet_name not in wb.sheetnames: return False
    ws = wb[sheet_name]
    num_sites = len(df_sites)
    START_COL, C_ROW, S_ROW, T_ROW, SEC_ROW, MAX_R = 4, 11, 12, 13, 14, 28
    
    # 0. Safety: Unmerge cells in the data matrix area to avoid 'MergedCell is read-only' errors
    data_merge_ranges = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= 11 and merged_range.max_row <= 30:
            data_merge_ranges.append(merged_range)
    for m_range in data_merge_ranges:
        ws.unmerge_cells(str(m_range))
        
    # Re-discover column offsets after scaling
    TOTAL_QTY_COL, rate_col, amt_col = None, None, None
    for c in range(START_COL, 65):
        h = str(ws.cell(row=12, column=c).value).upper().strip()
        if "TOTAL QUANTITY" in h: TOTAL_QTY_COL = c
        elif "RATE AS PER SOW" in h: rate_col = c
        elif "AMOUNT" in h: amt_col = c
    
    if not TOTAL_QTY_COL: TOTAL_QTY_COL = START_COL + num_sites
    if not rate_col: rate_col = TOTAL_QTY_COL + 1
    if not amt_col: amt_col = rate_col + 1

    # Inject data
    for i, (_, site_row) in enumerate(df_sites.iterrows()):
        curr_col = START_COL + i
        ws.cell(row=C_ROW, column=curr_col).value = i + 1
        
        # Site Header
        cell_id = ws.cell(row=S_ROW, column=curr_col)
        cell_id.value = str(site_row['PMP ID']).strip()
        cell_id.font = Font(name='Verdana', size=35, bold=True)
        cell_id.alignment = Alignment(text_rotation=90, horizontal='center', vertical='center')
        
        ws.cell(row=T_ROW, column=curr_col).value = str(site_row['Tower type']).strip()
        ws.cell(row=SEC_ROW, column=curr_col).value = site_row['NO OF SECTOR']
        
        for r in range(16, MAX_R):
            item_id = ws.cell(row=r, column=1).value
            try: code = str(int(item_id)) if item_id else ""
            except: code = str(item_id).strip() if item_id else ""
            
            # Mandatory Row Bypass for items that might not have a clean ID in Column A
            if not code and r == 26: code = "EXTRA VISIT"
            if not code and r == 27: code = "POLE MOUNT"
            
            val = site_row.iloc[code_to_col_idx[code]] if code in code_to_col_idx else 0.0
            cell = ws.cell(row=r, column=curr_col)
            cell.value = safe_float(val)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 3. STYLE DNA MIRRORING, RATES & SUMMATIONS (Mirror Row 25 down)
    fcl, lcl = get_column_letter(START_COL), get_column_letter(START_COL + num_sites - 1)
    for r in range(16, MAX_R + 1):
        # Mirror Style from Row 25 for all columns
        for c in range(1, amt_col + 1):
            copy_cell_style(ws.cell(row=25, column=c), ws.cell(row=r, column=c))
            
        if r < MAX_R:
            # Inject Hardcoded Rates for row 26/27 if applicable
            if r == 26: ws.cell(row=r, column=rate_col).value = 1000
            elif r == 27: ws.cell(row=r, column=rate_col).value = 500
            
            # Row-wise Summations and Amount Calculations
            ws.cell(row=r, column=TOTAL_QTY_COL).value = f"=SUM({fcl}{r}:{lcl}{r})"
            ws.cell(row=r, column=amt_col).value = f"={get_column_letter(TOTAL_QTY_COL)}{r}*{get_column_letter(rate_col)}{r}"
    
    ws.cell(row=SEC_ROW, column=TOTAL_QTY_COL).value = f"=SUM({fcl}{SEC_ROW}:{lcl}{SEC_ROW})"
    return True

def main():
    parser = argparse.ArgumentParser(description="Integrated JIO Billing Engine")
    parser.add_argument("master_file")
    parser.add_argument("billing_target")
    parser.add_argument("template_path")
    parser.add_argument("output_path")
    parser.add_argument("--mode", default="WCC", choices=["WCC", "JMS", "BOTH"])
    args = parser.parse_args()

    print(f"--- Launching Billing Engine Phase 2 (Mode: {args.mode}) ---")
    df_sites, code_to_col_idx = load_master_data(args.master_file, args.billing_target)
    
    if df_sites is not None:
        wb = openpyxl.load_workbook(args.template_path)
        
        # Global Header DC-Number Swap
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for r in range(1, 10):
                for c in range(1, 20):
                    if "DC-CODE" in str(ws.cell(row=r, column=c).value):
                        ws.cell(row=r, column=c).value = str(ws.cell(row=r, column=c).value).replace("DC-CODE", args.billing_target.upper())

        if args.mode in ["WCC", "BOTH"]:
            print("- Generating WCC Sheet...")
            generate_wcc_sheet(df_sites, wb)
            
        if args.mode in ["JMS", "BOTH"]:
            print("- Generating JMS Sheet (Matrix)...")
            populate_main_matrix('JMS', df_sites, code_to_col_idx, wb)

        wb.save(args.output_path)
        print(f"COMPLETE: {args.output_path}")
    else:
        print("CRITICAL: Failed to load data for the specified DC number.")
        sys.exit(1)

if __name__ == '__main__':
    main()
