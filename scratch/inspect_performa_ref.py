import openpyxl
import os

def inspect_performa_sheets(file_path):
    print(f"\n--- Inspecting Sheets in: {file_path} ---")
    if not os.path.exists(file_path):
        print("File not found!")
        return
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print("Sheets:", wb.sheetnames)
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\nSheet: {name}")
        for r in range(1, 5):
            row_vals = []
            for c in range(1, 16):
                val = ws.cell(row=r, column=c).value
                row_vals.append(str(val) if val is not None else "")
            print(f"R{r:2} | {' | '.join(row_vals)}")

output_ref = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/PerformaInvoiceWork/48/MATERIAL_REPORT_FOR_PERFORMA_INVOICE_NO_048_DIGITCOM.xlsx'
inspect_performa_sheets(output_ref)
