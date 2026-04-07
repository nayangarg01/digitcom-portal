import pandas as pd
import openpyxl
import sys
import argparse
import os
from copy import copy
import datetime

def load_master_data(master_path, target_billing_file):
    """Loads the KM SHEET from Master DPR and filters strictly for the target DC code."""
    try:
        df_master = pd.read_excel(master_path, sheet_name='KM SHEET', header=1)
    except Exception as e:
        print(f"Error loading master file {master_path}: {e}")
        return None

    df_master.columns = df_master.columns.astype(str).str.strip()
    
    # Identify BILLING FILE column safely
    bill_col = next((c for c in df_master.columns if 'BILLING FILE' in c.upper()), None)
    if not bill_col:
        print("Could not find 'BILLING FILE' column in master sheet.")
        return None
        
    df_filtered = df_master[df_master[bill_col].astype(str).str.strip().str.upper() == target_billing_file.upper()].copy()
    if df_filtered.empty:
        print(f"Warning: No sites found mapped to {target_billing_file}!")
        return None

    return df_filtered

def generate_wcc_sheet(df_sites, template_path, output_path):
    """Injects the filtered sites specifically into the WCC Template sheet."""
    # Target exactly the explicit string the user confirmed
    aktbc_col = next((c for c in df_sites.columns if 'CHRG EXTRA TRANSPORT' in str(c).upper() or 'AKTBC' == str(c).strip().upper()), None)
    if not aktbc_col:
        print("WARNING: Could not find strict CHRG EXTRA TRANSPORT or AKTBC column.")
    
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb['WCC']
    except Exception as e:
        print(f"Error loading template {template_path}: {e}")
        return False

    # Locate headers in WCC
    header_row_idx = None
    cols_map = {}
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and 'GIS SECTOR_ID' in str(row):
            header_row_idx = r_idx
            for c_idx, val in enumerate(row, start=1):
                if val:
                    cols_map[str(val).strip()] = c_idx
            break

    if not header_row_idx:
        print("Error: Could not locate visual headers in the WCC template sheet.")
        return False

    start_row = header_row_idx + 1
    
    # Pre-calculate base styling from the very first formatting row in the template
    # We will copy these styles to every new row so borders and date-formats survive
    base_styles = {}
    for c_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=start_row, column=c_idx)
        base_styles[c_idx] = {
            'font': copy(cell.font),
            'border': copy(cell.border),
            'fill': copy(cell.fill),
            'number_format': cell.number_format,
            'alignment': copy(cell.alignment)
        }

    # If df_sites is larger than 22, we need to inject empty rows safely 
    # to push the signatures down without destroying them. 
    if len(df_sites) > 22:
        ws.insert_rows(start_row + 22, amount=(len(df_sites) - 22))
    elif len(df_sites) < 22:
        # If the input has FEWER than 22 sites, we must actively DELETE the excess rows 
        # from the DC0105 template so we don't accidentally leave ghost sites behind.
        amount_to_delete = 22 - len(df_sites)
        ws.delete_rows(start_row + len(df_sites), amount=amount_to_delete)
    
    # Process and Map rows
    for i, (_, row) in enumerate(df_sites.iterrows()):
        curr_row = start_row + i
        
        def get_val(matcher):
            c_name = next((c for c in df_sites.columns if matcher.upper() in c.upper()), None)
            return row[c_name] if c_name else ""
            
        def get_exact(name):
            return row[name] if name in df_sites.columns else ""

        sr_no = i + 1
        enb_site = get_val('ENBSITEID') or get_val('SAP ID') or get_exact('Unnamed: 0')
        pmp_id = get_val('PMP ID')
        gis_id = get_val('GIS SECTOR')
        no_sec = get_val('NO OF SECTOR')
        tower = get_val('Tower type')
        jc = get_val('JC')
        wh = get_val('WH')
        veh = get_val('VEHICLE NO')
        min_no = get_val('MIN NO')
        
        min_date = get_val('MIN DATE')
        if pd.isna(min_date): min_date = ""
        comp_date = get_val('Completion Date')
        if pd.isna(comp_date): comp_date = ""

        remarks = "RFS DONE" if comp_date != "" else ""

        actual_km = row[aktbc_col] if aktbc_col else 0
        km_in_wo = get_val('KM IN WO')
        
        try:
            actual_km = float(actual_km) if pd.notna(actual_km) else 0.0
        except:
            actual_km = 0.0
            
        try:
            km_in_wo = float(km_in_wo) if pd.notna(km_in_wo) else 0.0
        except:
            km_in_wo = 0.0
            
        gap = actual_km - km_in_wo
        used_km = actual_km if actual_km <= km_in_wo else km_in_wo
        
        mapping = [
            ('Sr. No', sr_no),
            ('ENB SITE ID', enb_site),
            ('PMP SAP ID', pmp_id),
            ('GIS SECTOR_ID', gis_id),
            ('No of Sectors', no_sec),
            ('Tower type', tower),
            ('JC', jc),
            ('WH', wh),
            ('VEHICLE NO', veh),
            ('MIN  NO', min_no),
            ('MIN Date', min_date),
            ('Completion Date', comp_date),
            ('REMARKS', remarks),
            ('ACTUAL KM', actual_km),
            ('KM IN WO', km_in_wo),
            ('GAP', gap),
            ('USED KM IN WCC', used_km)
        ]

        # Apply formatting and values to the entire row spanning max columns
        for c_idx in range(1, ws.max_column + 1):
            c = ws.cell(row=curr_row, column=c_idx)
            styles = base_styles.get(c_idx)
            if styles:
                c.font = copy(styles['font'])
                c.border = copy(styles['border'])
                c.fill = copy(styles['fill'])
                c.number_format = styles['number_format']
                c.alignment = copy(styles['alignment'])

        # Inject Data Values
        for col_name, val in mapping:
            c_idx = next((cols_map[k] for k in cols_map if col_name.upper().strip() in k.upper().strip() or k.upper().strip() in col_name.upper().strip()), None)
            if c_idx:
                c = ws.cell(row=curr_row, column=c_idx)
                # Convert dates to actual datetime objects if they are pandas timestamps
                if isinstance(val, pd.Timestamp):
                    c.value = val.to_pydatetime()
                else:
                    c.value = val

    try:
        from openpyxl.utils import get_column_letter
    except ImportError:
        pass
        
    # After the loop completes, we must forcefully update the SUM formulas at the bottom!
    # Because openpyxl pushes rows down outside the original formula's range,
    # the original formulas like =SUM(U12:U33) fail to stretch.
    last_row = start_row + len(df_sites) - 1
    summary_row = last_row + 1
    
    for sum_col in ['ACTUAL KM', 'USED KM IN WCC']:
        # Find the column index
        c_idx = next((cols_map[k] for k in cols_map if sum_col.upper().strip() in k.upper().strip()), None)
        if c_idx:
            try:
                col_let = get_column_letter(c_idx)
                # Rewrite the exact excel formula correctly capturing all generated rows
                formula = f"=SUM({col_let}{start_row}:{col_let}{last_row})"
                ws.cell(row=summary_row, column=c_idx).value = formula
            except Exception as e:
                print(f"Failed to update sum formula for {sum_col}: {e}")

    try:
        wb.save(output_path)
        print(f"Successfully generated file: {output_path}")
        return True
    except Exception as e:
        print(f"Error saving output file: {e}")
        return False

def main():
    parser = argparse.ArgumentParser(description="Generate JIO Billing Formats from MASTERDPR.")
    parser.add_argument("master_file", help="Path to the MASTERDPR.xlsx file")
    parser.add_argument("billing_target", help="Target Billing File code (e.g., DC0105)")
    parser.add_argument("template_path", help="Path to the DC0105_TEMPLATE.xlsx file")
    parser.add_argument("output_path", help="Path where the generated file should be saved")
    args = parser.parse_args()

    print(f"--- Starting Sequence for {args.billing_target.upper()} ---")
    df_sites = load_master_data(args.master_file, args.billing_target)
    
    if df_sites is not None:
        print(f"Loaded {len(df_sites)} sites for {args.billing_target}.")
        print("Generating WCC Sheet...")
        generate_wcc_sheet(df_sites, args.template_path, args.output_path)
    else:
        print("Generation aborted due to missing data.")

if __name__ == '__main__':
    main()
