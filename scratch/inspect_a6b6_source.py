import openpyxl

file_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/PerformaInvoiceWork/50/DIGITCOM_AIRFIBER_DC062_JDPR 10-SEP-25 12-SEP-25  19-SEP-25_A6+B6.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    
    if 'JMS' in wb.sheetnames:
        ws = wb['JMS']
        print("\n--- JMS Row 12 ---")
        for cell in ws[12]:
            print(f"[{cell.column_letter}12]: {cell.value}", end=" | ")
        print("\n\n--- JMS Row 16 Sample ---")
        for cell in ws[16]:
             print(f"[{cell.column_letter}16]: {cell.value}", end=" | ")
except Exception as e:
    print(f"Error: {e}")
