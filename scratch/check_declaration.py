import openpyxl

generated_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/Billing/DC0111_A6B6_Billing_Final_Test_v6.xlsx'

try:
    wb = openpyxl.load_workbook(generated_path, data_only=True)
    if 'Declaration' in wb.sheetnames:
        ws = wb['Declaration']
        print("\n--- Declaration Sheet Row 13 ---")
        print(ws['A13'].value)
except Exception as e:
    print(f"Error: {e}")
