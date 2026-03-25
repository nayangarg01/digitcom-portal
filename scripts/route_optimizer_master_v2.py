import os
import math
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Warehouse Coords found in codebase
WAREHOUSE_COORDS = {
    'JAIPUR': (26.8139, 75.5450),
    'JODHPUR': (26.1245, 73.0543),
    'DEFAULT': (26.8139, 75.5450)
}

def get_distance(p1, p2):
    """Euclidean distance in KM (approx)"""
    return math.sqrt((p1[0] - p2[0])**2 + (p1[1] - p2[1])**2) * 111.0

def partition_sizes(n):
    if n < 2: return [n]
    if n == 2: return [2]
    if n == 3: return [3]
    if n == 4: return [2, 2]
    if n % 3 == 0: return [3] * (n // 3)
    if n % 3 == 1: return [3] * ((n - 4) // 3) + [2, 2]
    return [3] * ((n - 2) // 3) + [2]

def process_batch_file(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    headers = [str(cell.value) for cell in ws[1] if cell.value]
    h_idx = {h.strip().lower(): i for i, h in enumerate(headers)}
    
    sites = []
    for row in ws.iter_rows(min_row=2):
        row_data = [cell.value for cell in row]
        if not row_data[h_idx['site id']]: continue
        
        wh_name = str(row_data[h_idx['warehouse_name']]).upper().strip()
        wh_coords = WAREHOUSE_COORDS.get(wh_name, WAREHOUSE_COORDS['DEFAULT'])
        
        if row_data[h_idx['latitude']] is None or row_data[h_idx['longitude']] is None:
            print(f"Skipping site {row_data[h_idx['site id']]} due to missing coordinates.")
            continue
            
        sites.append({
            'id': str(row_data[h_idx['site id']]),
            'coords': (float(row_data[h_idx['latitude']]), float(row_data[h_idx['longitude']])),
            'cmp': str(row_data[h_idx['cmp']]).strip(),
            'wh_coords': wh_coords,
            'full_row': row_data
        })
    
    # 1. Group by CMP
    cmp_groups = {}
    for s in sites:
        c = s['cmp']
        if c not in cmp_groups: cmp_groups[c] = []
        cmp_groups[c].append(s)
    
    all_final_routes = []
    for cmp_name, group_sites in cmp_groups.items():
        wh_coords = group_sites[0]['wh_coords']
        
        # 1. Calculate Angle (Bearing) for each site
        for s in group_sites:
            s['angle'] = math.atan2(s['coords'][0] - wh_coords[0], s['coords'][1] - wh_coords[1])
            s['dist_to_wh'] = get_distance(wh_coords, s['coords'])
            
        # 2. Global Angular Sort (Sector-based)
        group_sites.sort(key=lambda x: x['angle'])
        
        sizes = partition_sizes(len(group_sites))
        curr = 0
        for size in sizes:
            chunk = group_sites[curr : curr + size]
            
            # 3. Local Sort by Distance (Inner to Outer) within sector
            chunk.sort(key=lambda x: x['dist_to_wh'])
            
            # 4. Recursive distance-based splitting inside chunk
            # WH -> A -> B -> C
            p_sites = chunk.copy()
            while p_sites:
                if len(p_sites) == 1:
                    all_final_routes.append([p_sites.pop(0)])
                elif len(p_sites) == 2:
                    A, B = p_sites[0], p_sites[1]
                    if get_distance(wh_coords, B['coords']) < get_distance(A['coords'], B['coords']):
                        all_final_routes.append([p_sites.pop(0)]) # Split
                    else:
                        all_final_routes.append([p_sites.pop(0), p_sites.pop(0)]) # Stay
                else: # len == 3
                    A, B, C = p_sites[0], p_sites[1], p_sites[2]
                    if get_distance(wh_coords, B['coords']) < get_distance(A['coords'], B['coords']):
                        all_final_routes.append([p_sites.pop(0)]) # A is separate
                        # then re-check for B, C
                    elif get_distance(wh_coords, C['coords']) < get_distance(B['coords'], C['coords']):
                        all_final_routes.append([p_sites.pop(0), p_sites.pop(0)]) # A, B together
                        # then C is separate
                    else:
                        all_final_routes.append([p_sites.pop(0), p_sites.pop(0), p_sites.pop(0)]) # All 3
            curr += size
            
    return headers, all_final_routes

def main():
    test_dir = 'RoutingSampleFiles/test_data_dates'
    output_path = 'RoutingSampleFiles/Optimized_Dispatch_Final_Master.xlsx'
    
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "Optimized_Routes"
    
    final_headers = []
    all_final_data = []
    
    files = sorted([f for f in os.listdir(test_dir) if f.endswith('.xlsx')])
    
    route_counter = 0
    for file_name in files:
        if file_name == 'No_Date.xlsx': continue
        headers, routes = process_batch_file(os.path.join(test_dir, file_name))
        
        if not final_headers:
            final_headers = headers + ['CLUBBING', 'AKTBC']
            new_ws.append(final_headers)
            
        date_prefix = file_name.replace('.xlsx', '')
        
        for r_idx, route in enumerate(routes):
            route_counter += 1
            # Route labeling: Date_A, Date_B, etc.
            route_label = f"{date_prefix}_{chr(65 + (r_idx % 26))}{ (r_idx // 26) + 1 if r_idx >= 26 else ''}"
            
            wh_coords = route[0]['wh_coords']
            prev_coords = wh_coords
            for s_idx, site in enumerate(route):
                dist = get_distance(prev_coords, site['coords'])
                if s_idx == 0:
                    dist = max(0, dist - 50) # 50km deduction for first leg
                
                clubbing = f"{route_label}-S{s_idx+1}"
                new_row = site['full_row'] + [clubbing, round(dist, 2)]
                new_ws.append(new_row)
                prev_coords = site['coords']

    # Final Styling
    header_fill = PatternFill(start_color="006100", end_color="006100", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    alignment = Alignment(horizontal="center", vertical="center")
    
    for row in new_ws.iter_rows():
        for cell in row:
            cell.alignment = alignment
            cell.border = border
            
    for cell in new_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    for col in new_ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        new_ws.column_dimensions[column].width = min(max_length + 2, 50)
        
    new_ws.freeze_panes = 'A2'
    new_wb.save(output_path)
    print(f"Done! {route_counter} routes generated in {output_path}")

if __name__ == "__main__":
    main()
