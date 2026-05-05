import openpyxl

file_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/InputBilling/DIGITCOM_AIRFIBER_DC0111_JPUR_12-MAR-26 23-MAR-26 & 25-MAR-26_A6+B6.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    if 'JMS' in wb.sheetnames:
        ws = wb['JMS']
        print("\n--- JMS Sheet Rows 1-25 ---")
        for row in ws.iter_rows(max_row=25, values_only=True):
            if any(x is not None for x in row):
                print(row)
    else:
        print("Sheet 'JMS' not found")
except Exception as e:
    print(f"Error: {e}")
