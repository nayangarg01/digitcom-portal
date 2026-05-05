import openpyxl
import os

folder = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/PerformaInvoiceWork/50'
target = "ONE-SECTOR SITE"

for filename in os.listdir(folder):
    if filename.endswith(".xlsx") and not filename.startswith("~"):
        path = os.path.join(folder, filename)
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            if 'JMS' in wb.sheetnames:
                ws = wb['JMS']
                for r in range(1, ws.max_row + 1):
                    for c in range(1, ws.max_column + 1):
                        val = str(ws.cell(row=r, column=c).value or "")
                        if target in val:
                            print(f"Found in {filename} at row {r}, col {c}: {val}")
        except Exception as e:
            print(f"Error reading {filename}: {e}")
