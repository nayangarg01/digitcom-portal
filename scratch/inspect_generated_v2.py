import openpyxl

generated_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/Billing/DC0111_A6B6_Billing_Test_Full_v2.xlsx'

try:
    wb = openpyxl.load_workbook(generated_path, data_only=True)
    ws = wb['Main WCC']
    # Row 20 in Excel is 0-indexed index 19. Col H is index 7.
    val = ws.cell(row=20, column=8).value
    print(f"Completion Date in Main WCC: {val}")
    
    # Also check if RFS DATE was found in df_sites columns (indirectly by checking if date is not N/A)
    print("\n--- Main WCC Row 20 ---")
    row20 = [c.value for c in ws[20]]
    print(row20)

except Exception as e:
    print(f"Error: {e}")
