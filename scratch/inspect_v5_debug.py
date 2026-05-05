import openpyxl

file_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/PerformaInvoiceWork/debug_Performa_Output_v5.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws_sum = wb['Summary sheet1']
    print("\n--- v5 Debug Performa Summary Sheet ---")
    for row in ws_sum.iter_rows(max_row=15, values_only=True):
        print(row)
        
    ws1 = wb['1']
    print("\n--- v5 Debug Performa Sheet '1' Row 1-5 ---")
    for row in ws1.iter_rows(max_row=5, values_only=True):
        print(row)
except Exception as e:
    print(f"Error: {e}")
