import openpyxl

file_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/InputBilling/MASTER TRACKER FOR BILLING-AIRFIBER-RJST (1).xlsx'
try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb['A6+B6 Billings']
    headers = [str(h).strip() for h in next(ws.iter_rows(max_row=1, values_only=True))]
    hop_col = headers.index('FB-FT HOP ID')
    dc_col = headers.index('BILLING FILE')
    
    print("\nHOP IDs for DC0111:")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[dc_col]).strip().upper() == 'DC0111':
            print(row[hop_col])
except Exception as e:
    print(f"Error: {e}")
