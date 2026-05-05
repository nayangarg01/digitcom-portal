import openpyxl
import os
from copy import copy

def inject_main_wcc_template(output_path, ref_path, wo_number):
    print(f"- Testing Main WCC Copy from: {os.path.basename(ref_path)}")
    wb_ref = openpyxl.load_workbook(ref_path)
    src_ws = wb_ref['Main WCC']
    
    wb_new = openpyxl.Workbook()
    dst_ws = wb_new.active
    dst_ws.title = 'Main WCC'
    
    # Copy values and styles
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.alignment = copy(cell.alignment)
    
    # Merged cells
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))
        
    # Dimensions - FIX APPLIED HERE
    for col, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col].width = dim.width
        dst_ws.column_dimensions[col].hidden = dim.hidden
    for row, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row].height = dim.height
        dst_ws.row_dimensions[row].hidden = dim.hidden
    
    dst_ws['D32'] = "22 SITES"
    dst_ws['I32'] = "30-JAN-26 TO 04-FEB-26"
    dst_ws['D29'] = wo_number
    
    wb_new.save(output_path)
    print(f"- SUCCESS. Saved to {output_path}")

if __name__ == "__main__":
    ref = "/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/FinaliseBillingFormat/DIGITCOM_ AIRFIBER_DC0105_ JDPR_29-JAN-26_A6 (REJECT) 2.xlsx"
    output = "/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/FinaliseBillingFormat/DC0105_FIXED_TEST.xlsx"
    inject_main_wcc_template(output, ref, "630330726")
    
    # Verify loadability
    try:
        openpyxl.load_workbook(output)
        print("Verification: File is perfectly loadable.")
    except Exception as e:
        print(f"Verification: FAILED with {e}")
