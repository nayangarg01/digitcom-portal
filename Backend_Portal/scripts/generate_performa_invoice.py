import pandas as pd
import xlsxwriter
import argparse
import os
import openpyxl
import re

def safe_float(val):
    try:
        if val is None or val == '': return 0.0
        return float(val)
    except:
        return 0.0

def generate_performa_invoice(dc_files, mindump_path, iv_number, activity, output_path):
    # 1. Load MINDUMP for WBS lookup (A6 Dump)
    wbs_mapping = {}
    try:
        # Load using pandas for easier site-based lookup
        df_dump = pd.read_excel(mindump_path, sheet_name='A6 DUMP')
        for _, row in df_dump.iterrows():
            site_id = str(row.get('Site ID', '')).strip()
            wbs_id = str(row.get('WBS ID', '')).strip()
            if site_id and wbs_id and site_id.lower() != 'nan':
                wbs_mapping[site_id] = wbs_id
    except Exception as e:
        print(f"Warning: Could not load MINDUMP for WBS mapping: {e}")

    # 2. Process each DC file
    all_rows = []
    
    for dc_file in dc_files:
        if not os.path.exists(dc_file):
            print(f"Warning: File not found: {dc_file}")
            continue
            
        try:
            wb = openpyxl.load_workbook(dc_file, data_only=True)
            
            # --- Nature of Work ---
            nature_of_work = "AIR FIBER INSTALLATION"
            if activity == 'A6_B6':
                nature_of_work = "AIR FIBER INSTALLATION(A6+B6)"
            
            if 'JMS' not in wb.sheetnames:
                print(f"Warning: JMS sheet missing in {dc_file}")
                continue
                
            ws_jms = wb['JMS']
            
            # --- Robustly find markers in JMS ---
            wo_number = "N/A"
            site_row = -1
            item_header_row = -1
            
            for r in range(1, 31):
                for c in range(1, 15):
                    cell_val = str(ws_jms.cell(row=r, column=c).value or "").strip()
                    
                    # 1. Look for Work Order No
                    if "Work Order" in cell_val and wo_number == "N/A":
                        # Value might be in same cell or next
                        if ':' in cell_val:
                            wo_number = cell_val.split(':')[-1].strip()
                        else:
                            # Try to extract the code at the end (e.g. P14/...)
                            parts = cell_val.split()
                            if parts: wo_number = parts[-1].strip()

                    # 2. Look for Site ID header
                    if "Site ID --" in cell_val:
                        site_row = r
                    
                    # 3. Look for Item Header
                    if 'Description' in cell_val and 'Item' in cell_val:
                        item_header_row = r
            
            # 3. Dynamic WO lookup from Main WCC (Standardized cell D29)
            if 'Main WCC' in wb.sheetnames:
                ws_wcc = wb['Main WCC']
                # Check cell D29 directly
                val = str(ws_wcc['D29'].value or "").strip()
                if val and val.upper() != 'NA' and val.upper() != 'TEMPLATE PLACEHOLDER':
                    wo_number = val
                else:
                    # Search if not in D29
                    for r in range(1, 40):
                        for c in range(1, 10):
                            cell_val = str(ws_wcc.cell(row=r, column=c).value or "").strip()
                            if "W.O.NUMBER" in cell_val.upper() or "WORK ORDER NO" in cell_val.upper():
                                # Try to get value from nearby cells
                                potential_val = str(ws_wcc.cell(row=r, column=c+1).value or ws_wcc.cell(row=r, column=c+2).value or "").strip()
                                if potential_val: wo_number = potential_val; break
                        if wo_number != "N/A": break

            if site_row == -1 or item_header_row == -1:
                print(f"Warning: Missing headers in {dc_file} (SiteRow={site_row}, ItemRow={item_header_row})")
                continue
                
            # --- Extract Sites ---
            sites = []
            # Sites usually start from column D (4) or check all columns in site_row
            for col in range(2, ws_jms.max_column + 1):
                val = str(ws_jms.cell(row=site_row, column=col).value or "").strip()
                if val and (val.startswith('I-RJ') or val.startswith('RJ')):
                    sites.append({'id': val, 'col': col})
                elif val == 'Total Quantity' or 'Total' in val:
                    # If we already have sites and hit Total, stop
                    if sites: break

            # Find Rate column
            rate_col = -1
            # Search in site_row or item_header_row
            for r_check in [site_row, item_header_row]:
                for col in range(1, ws_jms.max_column + 1):
                    val = str(ws_jms.cell(row=r_check, column=col).value or "").strip()
                    if 'RATE' in val:
                        rate_col = col
                        break
                if rate_col != -1: break
            
            if rate_col == -1: rate_col = ws_jms.max_column - 2 # Final fallback

            for site in sites:
                site_id = site['id']
                site_col = site['col']
                wbs_id = wbs_mapping.get(site_id, "N/A")
                
                # Items start from item_header_row + 1
                for r in range(item_header_row + 1, ws_jms.max_row + 1):
                    sap_code = str(ws_jms.cell(row=r, column=1).value or "").strip()
                    desc = str(ws_jms.cell(row=r, column=2).value or "").strip()
                    if not desc or desc.upper() == 'TOTAL': break
                    
                    qty_val = ws_jms.cell(row=r, column=site_col).value
                    qty = safe_float(qty_val)
                    
                    if qty > 0:
                        rate = safe_float(ws_jms.cell(row=r, column=rate_col).value)
                        site_amount = qty * rate
                        
                        all_rows.append({
                            'vendor': 'DIGITCOM INDIA TECHNOLOGIES',
                            'scope': 'ISP',
                            'iv_no': f"PERFORMA INVOICE NO. {iv_number}",
                            'wo_no': wo_number,
                            'site': site_id,
                            'wbs': wbs_id,
                            'sap_code': sap_code,
                            'description': desc,
                            'nature': nature_of_work,
                            'qty': qty,
                            'rate': rate,
                            'amount': site_amount
                        })
                        
        except Exception as e:
            print(f"Error processing {dc_file}: {e}")
            import traceback
            traceback.print_exc()

    # 3. Create Final Workbook
    with xlsxwriter.Workbook(output_path) as wb:
        # Formats
        f_header = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#D9D9D9', 'font_color': 'black', 'font_size': 10})
        f_cell = wb.add_format({'border': 1, 'align': 'left', 'valign': 'vcenter', 'font_size': 9})
        f_num = wb.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_size': 9})
        f_amount = wb.add_format({'border': 1, 'align': 'right', 'valign': 'vcenter', 'font_size': 9, 'num_format': '#,##0'})
        f_title = wb.add_format({'bold': True, 'font_size': 14, 'align': 'center', 'valign': 'vcenter'})
        
        f_sum_title = wb.add_format({'bold': True, 'font_size': 16, 'align': 'center', 'valign': 'vcenter', 'border': 1})
        f_sum_head = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#FFFF00', 'font_color': 'black'})
        f_sum_total = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#00B050', 'font_color': 'black'})
        f_sum_total_num = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#00B050', 'font_color': 'black', 'num_format': '#,##0'})

        # Sheet "1"
        ws1 = wb.add_worksheet('1')
        headers = ['VENDOR', 'SCOPE', 'IV NO', 'WO NO', 'SE NO', 'JMS', 'CI', 'SITE', 'WBS', 'DESC-SRIPTION', 'NATURE OF WORK', 'QTY', 'RATE', 'AMOUNT']
        for i, h in enumerate(headers):
            ws1.write(0, i, h, f_header)
        
        ws1.set_column(0, 0, 30) # Vendor
        ws1.set_column(1, 1, 10) # Scope
        ws1.set_column(2, 2, 25) # IV No
        ws1.set_column(3, 3, 20) # WO No
        ws1.set_column(7, 7, 20) # Site
        ws1.set_column(8, 8, 20) # WBS
        ws1.set_column(9, 9, 40) # Desc
        ws1.set_column(10, 10, 25) # Nature
        
        row_idx = 1
        for r in all_rows:
            ws1.write(row_idx, 0, r['vendor'], f_cell)
            ws1.write(row_idx, 1, r['scope'], f_cell)
            ws1.write(row_idx, 2, r['iv_no'], f_cell)
            ws1.write(row_idx, 3, r['wo_no'], f_cell)
            ws1.write(row_idx, 4, '', f_cell) # SE NO
            ws1.write(row_idx, 5, '', f_cell) # JMS
            ws1.write(row_idx, 6, '', f_cell) # CI
            ws1.write(row_idx, 7, r['site'], f_cell)
            ws1.write(row_idx, 8, r['wbs'], f_cell)
            ws1.write(row_idx, 9, r['description'], f_cell)
            ws1.write(row_idx, 10, r['nature'], f_cell)
            ws1.write(row_idx, 11, r['qty'], f_num)
            ws1.write(row_idx, 12, r['rate'], f_amount)
            ws1.write(row_idx, 13, r['amount'], f_amount)
            row_idx += 1

        # Summary Sheet
        ws_sum = wb.add_worksheet('Summary sheet1')
        ws_sum.set_column('A:A', 5)
        ws_sum.set_column('B:B', 50) # Description
        ws_sum.set_column('C:C', 15) # Code
        ws_sum.set_column('D:D', 15) # Rate
        ws_sum.set_column('E:E', 15) # Qty
        ws_sum.set_column('F:F', 20) # Amount
        
        ws_sum.merge_range('B2:F2', f'JMS DATA FOR PERFORMA INVOICE NO. {iv_number}', f_sum_title)
        
        sum_heads = ['DESCRIPTION', 'CODE', 'RATE', 'Sum of QTY', 'Sum of AMOUNT']
        for i, h in enumerate(sum_heads):
            ws_sum.write(2, i + 1, h, f_sum_head)
            
        df_all = pd.DataFrame(all_rows)
        if not df_all.empty:
            # Group by description, code, and rate
            pivot = df_all.groupby(['description', 'sap_code', 'rate']).agg({'qty': 'sum', 'amount': 'sum'}).reset_index()
            # Sort to match template if possible (usually by description)
            pivot = pivot.sort_values('description')
            
            curr_row = 3
            for _, row in pivot.iterrows():
                ws_sum.write(curr_row, 1, row['description'], f_cell)
                ws_sum.write(curr_row, 2, row['sap_code'], f_num)
                ws_sum.write(curr_row, 3, row['rate'], f_num)
                ws_sum.write(curr_row, 4, row['qty'], f_num)
                ws_sum.write(curr_row, 5, row['amount'], f_amount)
                curr_row += 1
            
            # Grand Total Row
            ws_sum.write(curr_row, 1, 'Grand Total', f_sum_total)
            ws_sum.write(curr_row, 2, '', f_sum_total)
            ws_sum.write(curr_row, 3, '', f_sum_total)
            ws_sum.write(curr_row, 4, df_all['qty'].sum(), f_sum_total_num)
            ws_sum.write(curr_row, 5, df_all['amount'].sum(), f_sum_total_num)

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--files", nargs='+', required=True)
    parser.add_argument("--mindump", required=True)
    parser.add_argument("--iv_number", required=True)
    parser.add_argument("--activity", default='A6')
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    
    generate_performa_invoice(args.files, args.mindump, args.iv_number, args.activity, args.output)
