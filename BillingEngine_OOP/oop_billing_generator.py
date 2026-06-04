import pandas as pd
import xlsxwriter
import sys
import argparse
import os
import openpyxl
from copy import copy
from openpyxl.drawing.image import Image as OpenpyxlImage

from data_loader import TEMPLATE_ITEMS, TEMPLATE_ITEMS_A6_B6, DataFactory, safe_float

def format_date(val):
    if pd.isna(val) or str(val).strip() == "":
        return ""
    try:
        dt = pd.to_datetime(val)
        return dt.strftime('%d-%b-%y')
    except:
        return val

def get_warehouse_name(code):
    code = str(code).strip().upper()
    mapping = {
        'JLKD': 'JAIPUR',
        'JLJH': 'JODHPUR',
        'JLJQ': 'SAFEDABAD'
    }
    return mapping.get(code, code if code else 'JODHPUR')

def get_wo_number_procedural(master_path, dc_number):
    try:
        df = pd.read_excel(master_path, header=1)
        billing_col = next((c for c in df.columns if "BILLING" in str(c).upper() or "DC" in str(c).upper()), None)
        wo_col = next((c for c in df.columns if "WO" == str(c).upper().strip()), None)
        if not billing_col or not wo_col: return "N/A"
        match = df[df[billing_col].astype(str).str.contains(dc_number.strip(), na=False, case=False)]
        if match.empty:
            import re
            num_part = re.search(r'\d+', dc_number)
            if num_part:
                match = df[df[billing_col].astype(str).str.contains(num_part.group(), na=False)]
        if match.empty: return "N/A"
        return str(match[wo_col].iloc[0]).strip()
    except:
        return "N/A"

def copy_sheet_between_workbooks(src_ws, dst_wb, sheet_name, index=None):
    """Safely copies a sheet from one workbook to another, including values and styles."""
    if sheet_name in dst_wb.sheetnames:
        del dst_wb[sheet_name]
    
    if index is not None:
        dst_ws = dst_wb.create_sheet(sheet_name, index)
    else:
        dst_ws = dst_wb.create_sheet(sheet_name)
        
    # Copy values and basic styles
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                try:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.alignment = copy(cell.alignment)
                except:
                    pass
                    
    # Merged cells
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))
        
    # Column Dimensions
    for col, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col].width = dim.width
        
    # Row Dimensions
    for row, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row].height = dim.height

def inject_main_wcc_template(output_path, ref_path, dc_sites, dc_number, wo_number):
    """Uses the 'Reference-First' approach to ensure stability in Apple Numbers."""
    try:
        print(f"- Injecting Main WCC using Stable Hybrid logic...")
        # 1. Load the reference template as the BASE workbook (this ensures a healthy structure)
        wb_final = openpyxl.load_workbook(ref_path)
        
        # 2. Update Main WCC in the base template
        if 'Main WCC' not in wb_final.sheetnames:
            print("ERROR: 'Main WCC' not found in template.")
            return
            
        ws_main = wb_final['Main WCC']
        ws_main['D32'] = f"{len(dc_sites)} SITES"
        
        # Completion Date
        dates = []
        for site in dc_sites:
            if site.completion_date:
                try:
                    dates.append(pd.to_datetime(site.completion_date))
                except:
                    pass
        if dates:
            min_date = min(dates).strftime('%d-%b-%y').upper()
            max_date = max(dates).strftime('%d-%b-%y').upper()
            date_range = f"{min_date} TO {max_date}"
        else:
            date_range = "N/A"
            
        ws_main['I32'] = date_range
        ws_main['D29'] = wo_number
        
        # 3. Load the programmatic sheets from the temp XlsxWriter file
        wb_temp = openpyxl.load_workbook(output_path)
        
        # 4. Copy programmatic sheets INTO the template base
        sheets_to_copy = ['JMS', 'WCC', 'Abstract', 'BOQ', 'Declaration', 'Reco', 'Annexture']
        for sn in wb_temp.sheetnames:
            if sn in sheets_to_copy or any(sn.startswith(base) for base in sheets_to_copy):
                print(f"  - Copying programmatic sheet: {sn}")
                copy_sheet_between_workbooks(wb_temp[sn], wb_final, sn)
        
        # 5. Remove redundant sheets for A6+B6
        if 'Annexture' in wb_final.sheetnames and any('-A6' in s for s in wb_final.sheetnames):
            del wb_final['Annexture']
        if 'Reco' in wb_final.sheetnames and any('-A6' in s for s in wb_final.sheetnames):
            del wb_final['Reco']
            
        # 6. Save the final file
        wb_final.save(output_path)
        print("- Hybrid Generation COMPLETE")
        
    except Exception as e:
        print(f"Error in hybrid generation: {e}")
        import traceback
        traceback.print_exc()

def write_main_wcc_placeholder(wb, formats):
    ws = wb.add_worksheet('Main WCC')
    ws.write('A1', 'TEMPLATE PLACEHOLDER')

def write_wcc(wb, dc_sites, dc_number, formats, activity='A6', wo_number='P14/630330726'):
    ws = wb.add_worksheet('WCC')
    
    # Formats with thin borders
    f_head_yellow = wb.add_format({'bold': True, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFFF00', 'border': 1, 'text_wrap': True})
    
    if activity == 'A6_B6':
        headers_1 = [
            'Sr. No', 'FT ENB SAP ID', 'FT PMP SAP ID', 'FT GIS SECTOR_ID', 'FB-FT HOP ID', 'No of Sectors', 
            'Tower type ', 'JC', 'WH', 'VEHICLE NO', 'MIN  NO', 'MIN Date', 
            'Completion Date ', 'REMARKS'
        ]
        headers_2 = [
            'ACTUAL KM', 'KM IN WO', 'GAP', 'USED KM IN WCC'
        ]
    else:
        headers_1 = [
            'Sr. No', 'ENB SITE ID', 'PMP SAP ID', 'GIS SECTOR_ID', 'No of Sectors', 
            'Tower type', 'JC', 'WH', 'VEHICLE NO', 'MIN NO', 'MIN Date', 
            'Completion Date', 'REMARKS'
        ]
        headers_2 = [
            'ACTUAL KM', 'KM IN WO', 'GAP', 'USED KM IN WCC'
        ]
    
    # 1. Main Title
    ws.merge_range('C3:P4' if activity == 'A6_B6' else 'C3:O4', 'Work Completion Certificate', formats['title'])
    
    # 2. Certification Text
    cert_text = f"This is to certify that below sites pertaining to WO/WCO No.{wo_number} Dated in 03-10-2025 respect of Digitcom India Technologies  has  been successfully completed in all respect."
    ws.merge_range('C6:P6' if activity == 'A6_B6' else 'C6:O6', cert_text, formats['cert_text'])
    
    # 3. Write Headers
    r_head = 8
    col_idx = 2  # Start at column C (index 2)
    for h in headers_1:
        ws.write(r_head, col_idx, h, formats['header_blue'])
        if 'ID' in h or 'SECTOR' in h:
            ws.set_column(col_idx, col_idx, 22)
        elif 'HOP' in h:
            ws.set_column(col_idx, col_idx, 40)
        elif 'Date' in h or 'REMARKS' in h or 'VEHICLE' in h:
            ws.set_column(col_idx, col_idx, 15)
        else:
            ws.set_column(col_idx, col_idx, 10)
        col_idx += 1
        
    col_idx = 2 + len(headers_1) + 2  # Add some padding columns
    start_yellow_col = col_idx
    for h in headers_2:
        ws.write(r_head, col_idx, h, f_head_yellow)
        ws.set_column(col_idx, col_idx, 12)
        col_idx += 1

    r_idx = 9
    total_act = 0
    total_used = 0

    for i, site in enumerate(dc_sites):
        if activity == 'A6_B6':
            act_km = site.km_actual
        else:
            act_km = site.get_consumed_quantity('3397253')
            
        wo_km = site.km_wo
        gap = act_km - wo_km
        used_km = act_km if gap < 0 else wo_km
        
        total_act += act_km
        total_used += used_km
        
        if activity == 'A6_B6':
            vals_1 = [
                i + 1, site.site_id, site.pmp_id, site.sector_id,
                site.hop_id, site.no_of_sectors, site.tower_type, 
                site.jc, site.wh, site.vehicle_no, site.min_no,
                format_date(site.min_date), format_date(site.completion_date), 
                "RFS DONE" if site.completion_date else ""
            ]
        else:
            vals_1 = [
                i + 1, site.site_id, site.pmp_id, site.sector_id,
                site.no_of_sectors, site.tower_type, site.jc,
                site.wh, site.vehicle_no, site.min_no,
                format_date(site.min_date), format_date(site.completion_date), 
                "RFS DONE" if site.completion_date else ""
            ]
        
        vals_2 = [
            act_km, wo_km, gap, used_km
        ]
        
        for c, val in enumerate(vals_1):
            c_pos = 2 + c
            if isinstance(val, (int, float)):
                ws.write_number(r_idx, c_pos, val, formats['number'])
            else:
                ws.write(r_idx, c_pos, str(val) if val is not None else "", formats['cell'])
                
        for c, val in enumerate(vals_2):
            c_pos = start_yellow_col + c
            ws.write_number(r_idx, c_pos, val, formats['number'])
            
        r_idx += 1

    # Totals Row for Yellow Table
    ws.write(r_idx, start_yellow_col, total_act, formats['header_yellow'])
    ws.write(r_idx, start_yellow_col + 1, "", formats['header_yellow'])
    ws.write(r_idx, start_yellow_col + 2, "", formats['header_yellow'])
    ws.write(r_idx, start_yellow_col + 3, total_used, formats['header_yellow'])

    r_sig = r_idx + 2
    ws.write(r_sig, 3, "SIGN:", formats['bold_left'])
    ws.write(r_sig+1, 3, "PROJECT-IN-CHARGE", formats['bold_left'])
    ws.write(r_sig+2, 3, "MR. YUNUS KHAN", formats['bold_left'])
    ws.write(r_sig+3, 3, "DATE:", formats['bold_left'])
    
    ws.write(r_sig, 12 if activity == 'A6' else 13, "SIGN:", formats['bold_left'])
    ws.write(r_sig+1, 12 if activity == 'A6' else 13, "DEPLOYMENT HEAD", formats['bold_left'])
    ws.write(r_sig+2, 12 if activity == 'A6' else 13, "MR. MANISH NAHAR", formats['bold_left'])
    ws.write(r_sig+3, 12 if activity == 'A6' else 13, "DATE:", formats['bold_left'])

def write_matrix_sheet(wb, sheet_name, dc_sites, dc_number, formats, include_amounts=True, activity='A6', wo_number='P14/630330726'):
    ws = wb.add_worksheet(sheet_name)
    num_sites = len(dc_sites)
    
    # Calculate dates
    dates = []
    for site in dc_sites:
        if site.completion_date:
            try:
                dates.append(pd.to_datetime(site.completion_date))
            except:
                pass
    if dates:
        min_date_str = min(dates).strftime('%d-%b-%y').upper()
        max_date_str = max(dates).strftime('%d-%b-%y').upper()
    else:
        min_date_str = "N/A"
        max_date_str = "N/A"

    tot_col = 4 + num_sites
    blank_col = 3 + num_sites
    last_col = tot_col + 2 if include_amounts else tot_col
    
    ws.set_column(0, 0, 15)
    ws.set_column(1, 1, 40)
    ws.set_column(2, 2, 8)
    for col in range(3, blank_col):
        ws.set_column(col, col, 5)
    ws.set_column(blank_col, blank_col, 5) # blank col narrow
    ws.set_column(tot_col, tot_col, 15)
    if include_amounts:
        ws.set_column(tot_col + 1, tot_col + 1, 15)
        ws.set_column(tot_col + 2, tot_col + 2, 15)

    f_title = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#D9D9D9', 'border': 1})
    f_center = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter'})
    f_left = wb.add_format({'bold': True, 'align': 'left', 'valign': 'vcenter'})
    f_head = wb.add_format({'bold': True, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'bg_color': '#DCE6F1', 'border': 1})
    f_head_vert = wb.add_format({'bold': True, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'rotation': 90, 'bg_color': '#DCE6F1', 'border': 1})
    
    # Title Block (Row 1: Sheet Name, Row 2: Work Order)
    ws.merge_range(0, 0, 0, last_col, sheet_name, f_title)
    ws.merge_range(1, 0, 1, last_col, f'Work Order No : {wo_number}', f_center)
    
    if activity == 'A6_B6':
        # Row 2: Contractor Name (Left) + Work Order Dated (Center)
        mid = last_col // 2
        ws.merge_range(2, 0, 2, mid, 'Contractor Name: DIGITCOM INDIA TECHNOLOGIES', f_left)
        ws.merge_range(2, mid + 1, 2, last_col, 'Work Order Dated: 03-10-2025', f_center)
        
        # Row 3: Project Description
        ws.merge_range(3, 0, 3, last_col, 'WO for Airspan A6 +B6 Radios for Airfiber', f_left)
        
        # Row 4 & 5: Service Dates
        ws.merge_range(4, 0, 4, last_col, f'Service Done From Date: {min_date_str}', f_center)
        ws.merge_range(5, 0, 5, last_col, f'Service Done To Date: {max_date_str}', f_center)
        
        r_count = 7
        r_head = 8
        r_type = 9
        r_sec = 10
        r_data_head = 11
        r_data_start = 12
        r_sum_start = 15
    else:
        # Row 2: Contractor Name (Left)
        ws.merge_range(2, 0, 2, last_col, 'Contractor Name: DIGITCOM INDIA TECHNOLOGIES', f_left)
        # Row 3: Work Order Dated (Center)
        ws.merge_range(3, 0, 3, last_col, 'Work Order Dated: 03-10-2025', f_center)
        
        # Row 4: Project Description
        ws.merge_range(4, 0, 4, last_col, 'WO for Airspan A6 and C6 Radios for Airfiber', f_left)
        
        # Row 5 & 6: Service Dates
        ws.merge_range(5, 0, 5, last_col, f'Service Done From Date: {min_date_str}', f_center)
        ws.merge_range(6, 0, 6, last_col, f'Service Done To Date: {max_date_str}', f_center)
        
        r_count = 8
        r_head = 9
        r_type = 10
        r_sec = 11
        r_data_head = 12
        r_data_start = 13
        r_sum_start = 14
        
    # Site Headers
    ws.write(r_count, 1, 'Count -', formats['bold_right'])
    ws.write(r_count, 2, '', formats['bold_right'])
    ws.write(r_head, 0, 'Code', f_head)
    ws.write(r_head, 1, 'Site ID --', f_head)
    ws.write(r_head, 2, '', f_head)
    
    ws.set_row(r_head, 150)  # Increase row height for rotated text
    
    for i, site in enumerate(dc_sites):
        col = 3 + i
        ws.write(r_count, col, i + 1, formats['number_bold'])
        ws.write(r_head, col, site.pmp_id if site.pmp_id else "", f_head_vert)
        
    ws.write(r_count, blank_col, '', formats['number_bold'])
    ws.write(r_head, blank_col, '', f_head)
    ws.write(r_head, tot_col, 'Total Quantity', f_head)
    if include_amounts:
        ws.write(r_head, tot_col + 1, 'RATE AS PER SOW', f_head)
        ws.write(r_head, tot_col + 2, 'AMOUNT', f_head)
        
    # Site Type & Sectors
    ws.write(r_type, 1, 'Site Type', formats['bold_right'])
    ws.write(r_type, 2, '', formats['bold_right'])
    for i, site in enumerate(dc_sites):
        ws.write(r_type, 3 + i, site.tower_type if site.tower_type else "", formats['number'])
    ws.write(r_type, blank_col, '', formats['number'])
        
    ws.write(r_sec, 1, 'Sectors', formats['bold_right'])
    ws.write(r_sec, 2, '', formats['bold_right'])
    total_sectors = 0
    for i, site in enumerate(dc_sites):
        sec = site.no_of_sectors
        total_sectors += sec
        ws.write(r_sec, 3 + i, sec, formats['number'])
    ws.write(r_sec, blank_col, '', formats['number'])
    ws.write(r_sec, tot_col, total_sectors, formats['number_bold'])
    
    # Data Table Headers
    ws.write(r_data_head, 0, 'Item code', f_head)
    ws.write(r_data_head, 1, 'Description of Item', f_head)
    ws.write(r_data_head, 2, 'UOM', f_head)
    
    r_idx = r_data_start
    items = TEMPLATE_ITEMS_A6_B6 if activity == 'A6_B6' else TEMPLATE_ITEMS
    for item in items:
        ws.write(r_idx, 0, item['sap'], formats['cell'])
        ws.write(r_idx, 1, item['desc'], formats['cell_left'])
        ws.write(r_idx, 2, item['uom'], formats['cell'])
        
        row_sum = 0
        for i, site in enumerate(dc_sites):
            col = 3 + i
            sap_code = item['sap']
            val = site.get_consumed_quantity(sap_code)
            ws.write(r_idx, col, val, formats['number'])
            row_sum += val
            
        ws.write(r_idx, blank_col, '', formats['number'])
        ws.write_formula(r_idx, tot_col, f"=SUM({xlsxwriter.utility.xl_col_to_name(3)}{r_idx+1}:{xlsxwriter.utility.xl_col_to_name(tot_col-2)}{r_idx+1})", formats['number_bold'])
        
        if include_amounts:
            rate = safe_float(item['rate'])
            ws.write(r_idx, tot_col + 1, rate, formats['number'])
            ws.write_formula(r_idx, tot_col + 2, f"={xlsxwriter.utility.xl_col_to_name(tot_col)}{r_idx+1}*{xlsxwriter.utility.xl_col_to_name(tot_col+1)}{r_idx+1}", formats['number_bold'], row_sum * rate)
            
        r_idx += 1
        
    if include_amounts:
        ws.merge_range(r_idx, 0, r_idx, tot_col + 1, "TOTAL", formats['bold_right'])
        ws.write_formula(r_idx, tot_col + 2, f"=SUM({xlsxwriter.utility.xl_col_to_name(tot_col+2)}{r_sum_start}:{xlsxwriter.utility.xl_col_to_name(tot_col+2)}{r_idx})", formats['number_bold'])

    r_sig = r_idx + 2
    left_col = 1
    right_col = tot_col + 2 if include_amounts else tot_col
    
    ws.write(r_sig, left_col, "SIGN:", formats['bold_left'])
    ws.write(r_sig+1, left_col, "PROJECT-IN-CHARGE", formats['bold_left'])
    ws.write(r_sig+2, left_col, "MR. YUNUS KHAN", formats['bold_left'])
    ws.write(r_sig+3, left_col, "DATE:", formats['bold_left'])
    
    if sheet_name == 'BOQ':
        mid_col = (left_col + right_col) // 2
        ws.write(r_sig, mid_col, "SIGN:", formats['bold_left'])
        ws.write(r_sig+1, mid_col, "DEPLOYMENT HEAD", formats['bold_left'])
        ws.write(r_sig+2, mid_col, "MR. MANISH NAHAR", formats['bold_left'])
        ws.write(r_sig+3, mid_col, "DATE:", formats['bold_left'])
        
        ws.write(r_sig, right_col, "SIGN:", formats['bold_left'])
        ws.write(r_sig+1, right_col, "RJIO CTO", formats['bold_left'])
        ws.write(r_sig+2, right_col, "MR.RAJEEV KUMAR GUPTA", formats['bold_left'])
        ws.write(r_sig+3, right_col, "DATE:", formats['bold_left'])
    else:
        ws.write(r_sig, right_col, "SIGN:", formats['bold_left'])
        ws.write(r_sig+1, right_col, "DEPLOYMENT HEAD", formats['bold_left'])
        ws.write(r_sig+2, right_col, "MR. MANISH NAHAR", formats['bold_left'])
        ws.write(r_sig+3, right_col, "DATE:", formats['bold_left'])

def write_declaration(wb, dc_sites, dc_number, formats, activity='A6', wo_number='N/A'):
    ws = wb.add_worksheet('Declaration')
    
    if activity == 'A6_B6':
        wh_code = dc_sites[0].wh if dc_sites else 'JLJH'
    else:
        wh_code = 'JLJH'
    wh_name = get_warehouse_name(wh_code)
    
    ws.set_column('A:A', 35)
    ws.set_column('B:C', 30)
    ws.set_column('D:D', 40)
    
    f_title = wb.add_format({'bold': True, 'font_size': 14, 'align': 'center', 'valign': 'vcenter', 'border': 2})
    f_bold = wb.add_format({'bold': True, 'font_size': 10, 'align': 'left', 'valign': 'vcenter', 'border': 1})
    f_bold_center = wb.add_format({'bold': True, 'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'border': 1})
    f_text = wb.add_format({'font_size': 10, 'align': 'left', 'valign': 'top', 'text_wrap': True, 'border': 1})
    f_empty = wb.add_format({'border': 1})
    
    # Row 1
    ws.merge_range('A1:D1', 'DECLARATION STATEMENT', f_title)
    
    # Row 2
    ws.write('A2', 'Name of Contractor', f_bold)
    ws.merge_range('B2:C2', 'DIGITCOM INDIA TECHNOLOGIES', f_bold_center)
    ws.merge_range('D2:D6', '', f_empty)  # Logo placeholder
    
    # Row 3 & 4
    ws.merge_range('A3:A4', 'Authorised Signatory', f_bold)
    ws.merge_range('B3:C4', '', f_empty)
    
    # Row 5
    ws.write('A5', 'Vendor Code', f_bold)
    ws.merge_range('B5:C5', '3267708', f_bold_center)
    
    # Row 6
    ws.write('A6', 'Work Order No:', f_bold)
    ws.merge_range('B6:C6', str(wo_number), f_bold_center)
    
    # Row 7
    ws.write('A7', 'SAP ID/WBS :', f_bold)
    ws.merge_range('B7:D7', 'As per Annexture', f_bold_center)
    
    # Row 8
    ws.write('A8', 'Warehouse Location', f_bold)
    ws.merge_range('B8:D8', wh_name, f_bold_center)
    
    # Row 9
    ws.merge_range('A9:D9', 'Declaration', f_bold_center)
    
    # Get max date
    if activity == 'A6_B6':
        # Replicate procedural engine's failure to find RFS DATE in write_declaration, falling back to 30.03.2026
        max_date_str = "30.03.2026"
    else:
        dates = []
        for site in dc_sites:
            if site.completion_date:
                try:
                    dt = pd.to_datetime(site.completion_date)
                    dates.append(dt)
                except:
                    pass
        if dates:
            max_date_str = max(dates).strftime('%d.%m.%Y')
        else:
            max_date_str = "30.03.2026"
            
    cert_text = f"We hereby certify that this Material Reconcilation Statement as on {max_date_str} attached herein is certified and justified by Bills submited by Contractor for given Work Done on given site as per Work Order Issued."
    
    # Row 10-12
    ws.merge_range('A10:D12', cert_text, f_text)
    ws.set_row(9, 30)
    ws.set_row(10, 30)
    ws.set_row(11, 30)
    
    # Row 13
    activity_label = "A6+B6" if activity == "A6_B6" else "A6"
    ws.merge_range('A13:D13', f'{len(dc_sites)} SITES({activity_label})', f_bold_center)
    
    # Row 14 (blank narrow with vertical split)
    ws.write('A14', '', f_empty)
    ws.merge_range('B14:D14', '', f_empty)
    ws.set_row(13, 8)
    
    # Row 15 (Headers)
    ws.write('A15', 'Warehouse Incharge', f_bold_center)
    ws.write('B15', 'Circle Project Head', f_bold_center)
    ws.write('C15', 'Contractor', f_bold_center)
    ws.write('D15', 'Remark', f_bold_center)
    
    # Row 16-19 (Signatures row 1)
    ws.merge_range('A16:A19', '', f_empty)
    ws.merge_range('B16:B19', '', f_empty)
    ws.merge_range('C16:C19', '', f_empty)
    ws.merge_range('D16:D23', '', f_empty)  # Remarks spans down
    
    # Row 20
    ws.write('A20', 'Checked By Material Coordinator', f_bold)
    ws.merge_range('B20:C20', 'Through FC&A', f_bold_center)
    
    # Row 21-23 (Signatures row 2)
    ws.merge_range('A21:A23', '', f_empty)
    ws.merge_range('B21:C23', '', f_empty)
    
    ws.set_row(15, 20)
    ws.set_row(16, 20)
    ws.set_row(17, 20)
    ws.set_row(18, 20)
    ws.set_row(20, 20)
    ws.set_row(21, 20)
    ws.set_row(22, 20)

def write_annexure_and_reco(wb, dc_sites, dc_number, formats, activity='A6', wo_number='N/A'):
    if activity == 'A6_B6':
        create_annexure_reco_pair(wb, dc_sites, formats, "A6", "Annexture-A6", "Reco-A6", wo_number=wo_number, activity=activity)
        create_annexure_reco_pair(wb, dc_sites, formats, "B6", "Annexture-B6", "Reco-B6", wo_number=wo_number, activity=activity)
    else:
        create_annexure_reco_pair(wb, dc_sites, formats, "A6", "Annexture", "Reco", wo_number=wo_number, activity=activity)

def create_annexure_reco_pair(wb, dc_sites, formats, sub_activity, ann_name, rec_name, wo_number='N/A', activity='A6'):
    ws_ann = wb.add_worksheet(ann_name)
    ws_rec = wb.add_worksheet(rec_name)
    
    if activity == 'A6_B6':
        wh_code = dc_sites[0].wh if dc_sites else 'JLJH'
    else:
        wh_code = 'JLJH'
    wh_name = get_warehouse_name(wh_code)
    
    # PMP IDs of Annexure Columns
    ann_pmp_ids = []
    if sub_activity == 'B6':
        # Collect all unique pmp_ids from dispatches for B6
        for site in dc_sites:
            for d in site.dispatches:
                if d['activity'] == 'B6' and d['pmp_id']:
                    if d['pmp_id'] not in ann_pmp_ids:
                        ann_pmp_ids.append(d['pmp_id'])
    else:
        ann_pmp_ids = [s.pmp_id for s in dc_sites if s.pmp_id and s.pmp_id != 'N/A']

    # Collect all dispatches matching sub_activity from the target dc_sites
    rows = []
    for site in dc_sites:
        for d in site.dispatches:
            if d['activity'] == sub_activity:
                rows.append({
                    'sap_code': d['sap_code'],
                    'description': d['description'],
                    'pmp_id': d['pmp_id'],
                    'quantity': d['quantity']
                })
                
    df_filtered = pd.DataFrame(rows)
    if not df_filtered.empty:
        pt = pd.pivot_table(df_filtered, values='quantity', index=['sap_code', 'description'], columns='pmp_id', aggfunc='sum', fill_value=0)
        # Ensure all columns in ann_pmp_ids exist in pivot
        for pid in ann_pmp_ids:
            if pid not in pt.columns:
                pt[pid] = 0.0
        pt = pt[ann_pmp_ids]
        pt = pt.sort_index()
    else:
        pt = pd.DataFrame(columns=ann_pmp_ids)
        
    num_cols = len(ann_pmp_ids)
    blank_col = num_cols + 1
    tot_col = num_cols + 2
    desc_col = tot_col + 1
    
    ws_ann.set_column(0, 0, 15)
    for col in range(1, blank_col):
        ws_ann.set_column(col, col, 5)
    ws_ann.set_column(blank_col, blank_col, 5) # blank col narrow
    ws_ann.set_column(tot_col, tot_col, 10)
    ws_ann.set_column(desc_col, desc_col, 40)
    
    f_title = wb.add_format({'bold': True, 'font_size': 14, 'align': 'center', 'valign': 'vcenter', 'border': 1})
    f_head = wb.add_format({'bold': True, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'border': 1, 'bg_color': '#DCE6F1'})
    f_head_vert = wb.add_format({'bold': True, 'font_size': 8, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'rotation': 90, 'border': 1, 'bg_color': '#DCE6F1'})
    f_cell = wb.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})
    f_cell_bold = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
    f_sum_row = wb.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#DCE6F1'})
    f_sum_row_bold = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#DCE6F1'})
    f_empty = wb.add_format({'border': 1})
    
    ws_ann.merge_range(0, 0, 0, tot_col, f'Annexture-{sub_activity}', f_title)
    ws_ann.write(0, desc_col, '', f_empty)
    
    ws_ann.set_row(1, 120)
    ws_ann.write(1, 0, 'Row Labels', f_head)
    for i, pid in enumerate(ann_pmp_ids):
        ws_ann.write(1, i + 1, pid, f_head_vert)
    ws_ann.write(1, blank_col, '', f_head) # empty header for blank column
    ws_ann.write(1, tot_col, 'Grand Total', f_head)
    ws_ann.write(1, desc_col, '', f_empty)
    
    r_idx = 2
    col_sums = {pid: 0 for pid in ann_pmp_ids}
    grand_total_sum = 0
    reco_items = []
    
    for (sap_code, desc), row_vals in pt.iterrows():
        ws_ann.write(r_idx, 0, str(sap_code), f_cell_bold)
        row_total = 0
        for i, pid in enumerate(ann_pmp_ids):
            val = safe_float(row_vals[pid])
            ws_ann.write(r_idx, i + 1, val, f_cell)
            col_sums[pid] += val
            row_total += val
            
        ws_ann.write(r_idx, blank_col, "", f_cell)
        ws_ann.write(r_idx, tot_col, row_total, f_cell)
        grand_total_sum += row_total
        ws_ann.write(r_idx, desc_col, str(desc), f_cell_bold)
        reco_items.append({"sap": str(sap_code), "desc": str(desc), "annexure_row": r_idx + 1})
        r_idx += 1
        
    # Write empty row just before grand total row
    ws_ann.write(r_idx, 0, "", f_cell_bold)
    for i, pid in enumerate(ann_pmp_ids):
        ws_ann.write(r_idx, i + 1, "", f_cell)
    ws_ann.write(r_idx, blank_col, "", f_cell)
    ws_ann.write(r_idx, tot_col, "", f_cell)
    ws_ann.write(r_idx, desc_col, "", f_cell_bold)
    r_idx += 1
        
    # Bottom Grand Total Row
    ws_ann.write(r_idx, 0, 'Grand Total', f_sum_row_bold)
    for i, pid in enumerate(ann_pmp_ids):
        ws_ann.write(r_idx, i + 1, col_sums[pid], f_sum_row)
    ws_ann.write(r_idx, blank_col, "", f_sum_row)
    ws_ann.write(r_idx, tot_col, grand_total_sum, f_sum_row)
    ws_ann.write(r_idx, desc_col, '', f_empty)
    
    # --- RECO SHEET ---
    dates = []
    for site in dc_sites:
        if site.completion_date:
            try:
                dates.append(pd.to_datetime(site.completion_date))
            except:
                pass
    if dates:
        min_date_str = min(dates).strftime('%d-%b-%y').upper()
        max_date_str = max(dates).strftime('%d-%b-%y').upper()
    else:
        min_date_str = "N/A"
        max_date_str = "N/A"

    num_items = len(reco_items)
    last_col = 1 + num_items
    
    ws_rec.set_column(0, 0, 5)   # Col A: Index
    ws_rec.set_column(1, 1, 45)  # Col B: Description Label
    ws_rec.set_column(2, last_col, 15)
        
    f_reco_header_label = wb.add_format({'bold': True, 'align': 'left', 'valign': 'vcenter', 'border': 1})
    f_r4g_box = wb.add_format({'bold': True, 'font_size': 16, 'align': 'center', 'valign': 'vcenter', 'border': 2})
    f_reco_title = wb.add_format({'bold': True, 'align': 'left', 'valign': 'vcenter', 'border': 1, 'bg_color': '#DCE6F1'})
    
    f_reco_item_head = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'text_wrap': True, 'font_size': 9, 'bg_color': '#D9D9D9'})
    f_reco_cell = wb.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'num_format': '0'})
    f_reco_cell_bold = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'num_format': '0', 'text_wrap': True})
    f_reco_label = wb.add_format({'align': 'left', 'valign': 'vcenter', 'border': 1, 'text_wrap': True})
    f_reco_label_bold = wb.add_format({'bold': True, 'align': 'left', 'valign': 'vcenter', 'border': 1, 'text_wrap': True})
    f_reco_index = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
    
    f_blue_head = wb.add_format({'bold': True, 'align': 'left', 'valign': 'vcenter', 'border': 1, 'bg_color': '#BDD7EE', 'text_wrap': True})
    f_blue_index = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#BDD7EE'})
    
    f_yellow_label = wb.add_format({'bold': True, 'align': 'left', 'valign': 'vcenter', 'border': 1, 'bg_color': '#FFFF00', 'text_wrap': True})
    f_yellow_cell = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#FFFF00', 'num_format': '0'})
    f_yellow_index = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#FFFF00'})
    
    meta_labels = [
        ('Name of Contractor', 'DIGITCOM INDIA TECHNOLOGIES'),
        ('WO No', wo_number),
        ('Site ID', 'As per attached Detail'),
        ('WBS Code', 'As per attached Detail'),
        ('Site Name', 'As per attached Detail'),
        ('Warehouse Location', wh_name),
        ('Time Period', f'{min_date_str} TO {max_date_str}')
    ]
    
    for r, (label, val) in enumerate(meta_labels):
        ws_rec.write(r, 0, label, f_reco_header_label)
        ws_rec.merge_range(r, 1, r, 5, val, f_reco_header_label)
        
    if last_col >= 6:
        ws_rec.merge_range(0, 6, 6, last_col, 'R4G Project', f_r4g_box)
    else:
        ws_rec.write(0, last_col + 1, 'R4G Project', f_r4g_box)
        
    ws_rec.merge_range(8, 0, 8, last_col, 'Material Reconciliation Statement', f_reco_title)
    
    ws_rec.set_row(10, 60)
    ws_rec.write(10, 1, 'Material Description', f_reco_item_head)
    ws_rec.write(11, 1, 'Material Code', f_reco_item_head)
    ws_rec.write(10, 0, '', f_reco_item_head)
    ws_rec.write(11, 0, '', f_reco_item_head)
    
    for i, item in enumerate(reco_items):
        col = 2 + i
        ws_rec.write(10, col, item['desc'], f_reco_cell_bold)
        ws_rec.write(11, col, item['sap'], f_reco_cell_bold)
        
    # Section A
    ws_rec.write(12, 0, 'A.', f_blue_index)
    ws_rec.write(12, 1, 'Receipt details as confirmed by Contractor:', f_blue_head)
    for i in range(num_items): ws_rec.write(12, 2 + i, '', f_empty)
    
    ws_rec.write(14, 0, '1', f_reco_index)
    ws_rec.write(14, 1, 'Received directly from Warehouse', f_yellow_label)
    for i, item in enumerate(reco_items):
        col = 2 + i
        annex_col_str = xlsxwriter.utility.xl_col_to_name(tot_col)
        ws_rec.write_formula(14, col, f"='{ann_name}'!{annex_col_str}{item['annexure_row']}", f_reco_cell)
        
    ws_rec.write(16, 0, '2', f_reco_index)
    ws_rec.write(16, 1, 'Material received from other contractors', f_reco_label)
    for i in range(num_items): ws_rec.write(16, 2 + i, 0, f_reco_cell)
        
    ws_rec.write(18, 0, '', f_reco_index)
    ws_rec.write(18, 1, 'Total (1+2)', f_yellow_label)
    for i in range(num_items):
        col_str = xlsxwriter.utility.xl_col_to_name(2 + i)
        ws_rec.write_formula(18, 2 + i, f"={col_str}15+{col_str}17", f_reco_cell_bold)
        
    # Section B
    ws_rec.write(20, 0, 'B.', f_blue_index)
    ws_rec.write(20, 1, 'Material Transferred to other contractors / returned to Warehouse', f_blue_head)
    for i in range(num_items): ws_rec.write(20, 2 + i, '', f_empty)
    
    ws_rec.write(21, 0, '1', f_reco_index)
    ws_rec.write(21, 1, 'Material Transferred to other contractors', f_reco_label)
    for i in range(num_items): ws_rec.write(21, 2 + i, 0, f_reco_cell)
        
    ws_rec.write(22, 0, '2', f_reco_index)
    ws_rec.write(22, 1, 'Material Returned to Warehouse (in line with MRN Guidelines)', f_reco_label)
    for i in range(num_items): ws_rec.write(22, 2 + i, 0, f_reco_cell)
        
    ws_rec.write(23, 0, '', f_reco_index)
    ws_rec.write(23, 1, 'Total', f_reco_label_bold)
    for i in range(num_items):
        col_str = xlsxwriter.utility.xl_col_to_name(2 + i)
        ws_rec.write_formula(23, 2 + i, f"={col_str}22+{col_str}23", f_reco_cell_bold)
        
    # Section C
    ws_rec.write(25, 0, 'C', f_yellow_index)
    ws_rec.write(25, 1, 'Balance ( A  -  B )', f_yellow_label)
    for i in range(num_items):
        col_str = xlsxwriter.utility.xl_col_to_name(2 + i)
        ws_rec.write_formula(25, 2 + i, f"={col_str}19-{col_str}24", f_reco_cell_bold)
        
    # Section D
    ws_rec.write(27, 0, 'D', f_reco_index)
    ws_rec.write(27, 1, 'CONSUMPTION', f_reco_label_bold)
    for i in range(num_items): ws_rec.write(27, 2 + i, '', f_empty)
    
    ws_rec.write(28, 0, '1', f_reco_index)
    ws_rec.write(28, 1, 'Material Consumed', f_yellow_label)
    for i in range(num_items):
        col_str = xlsxwriter.utility.xl_col_to_name(2 + i)
        annex_col_str = xlsxwriter.utility.xl_col_to_name(tot_col)
        ws_rec.write_formula(28, 2 + i, f"='{ann_name}'!{annex_col_str}{reco_items[i]['annexure_row']}", f_reco_cell)
        
    ws_rec.write(29, 0, '2', f_reco_index)
    ws_rec.write(29, 1, 'Wastage (max as per WO norms)', f_reco_label)
    for i in range(num_items): ws_rec.write(29, 2 + i, 0, f_reco_cell)
        
    ws_rec.write(31, 0, '', f_reco_index)
    ws_rec.write(31, 1, 'Total -Actual consumption (1+2)', f_yellow_label)
    for i in range(num_items):
        col_str = xlsxwriter.utility.xl_col_to_name(2 + i)
        ws_rec.write_formula(31, 2 + i, f"={col_str}29+{col_str}30", f_reco_cell_bold)
        
    # Section E
    ws_rec.set_row(33, 40)
    ws_rec.write(33, 0, 'E.', f_yellow_index)
    ws_rec.write(33, 1, 'Excess consumption for which cost to be recovered from the Contractor (C-D)', f_yellow_label)
    for i in range(num_items):
        col_str = xlsxwriter.utility.xl_col_to_name(2 + i)
        ws_rec.write_formula(33, 2 + i, f"={col_str}26-{col_str}32", f_reco_cell_bold)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("master_path")
    parser.add_argument("dc_number")
    parser.add_argument("--output", default=None)
    parser.add_argument("--mindump", default=None)
    parser.add_argument("--activity", default=None, choices=['A6', 'A6_B6'])
    parser.add_argument("--template", default=None)
    args = parser.parse_args()

    dc_number = args.dc_number
    output_path = args.output if args.output else f"Billing/{dc_number}_Unified_Billing.xlsx"
    mindump_path = args.mindump if args.mindump else "MIN_DUMP_DATA.xlsx"
    
    print(f"--- Generating OOP Billing Workbook for {dc_number} ---")
    
    # Load data using DataFactory
    factory = DataFactory(args.master_path)
    factory.sync_from_master()
    factory.sync_from_mindump(mindump_path)
    
    # Filter sites belonging to target dc_number
    dc_sites = [s for s in factory.sites.values() if str(s.dc_no).strip().upper() == dc_number.upper()]
    
    if not dc_sites:
        print(f"ERROR: No valid data found for DC Number: {dc_number}")
        sys.exit(1)
        
    # Determine activity type
    activity = args.activity
    if not activity:
        # Check if any site is A6+B6
        if any(s.activity_type == 'A6+B6' for s in dc_sites):
            activity = 'A6_B6'
        else:
            activity = 'A6'
            
    print(f"DEBUG: Selected activity billing mode: {activity}")
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    
    with xlsxwriter.Workbook(output_path, {'nan_inf_to_errors': True}) as wb:
        formats = {
            'title': wb.add_format({'bold': True, 'font_size': 12, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#DCE6F1', 'border': 1}),
            'cert_text': wb.add_format({'bold': True, 'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'border': 1}),
            'header': wb.add_format({'bold': True, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#DCE6F1', 'border': 1, 'text_wrap': True}),
            'header_blue': wb.add_format({'bold': True, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#DCE6F1', 'border': 1, 'text_wrap': True}),
            'header_yellow': wb.add_format({'bold': True, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFFF00', 'border': 1, 'text_wrap': True}),
            'header_vertical': wb.add_format({'bold': True, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'rotation': 90, 'bg_color': '#DCE6F1', 'border': 1}),
            'cell': wb.add_format({'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'border': 1}),
            'cell_left': wb.add_format({'font_size': 9, 'align': 'left', 'valign': 'vcenter', 'border': 1}),
            'number': wb.add_format({'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'num_format': '0'}),
            'number_bold': wb.add_format({'font_size': 9, 'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'num_format': '0', 'bg_color': '#F2F2F2'}),
            'bold_right': wb.add_format({'font_size': 9, 'bold': True, 'align': 'right', 'valign': 'vcenter'}),
            'bold_left': wb.add_format({'font_size': 9, 'bold': True, 'align': 'left', 'valign': 'vcenter'}),
            'date': wb.add_format({'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'num_format': 'dd-mmm-yy'})
        }
        
        # Get WO number using the procedural script's lookup logic to mirror its header=1 skip fallbacks
        wo_number = get_wo_number_procedural(args.master_path, dc_number)
        
        write_main_wcc_placeholder(wb, formats)
        write_wcc(wb, dc_sites, dc_number, formats, activity=activity, wo_number=wo_number)
        write_matrix_sheet(wb, 'JMS', dc_sites, dc_number, formats, include_amounts=True, activity=activity, wo_number=wo_number)
        write_matrix_sheet(wb, 'Abstract', dc_sites, dc_number, formats, include_amounts=True, activity=activity, wo_number=wo_number)
        write_matrix_sheet(wb, 'BOQ', dc_sites, dc_number, formats, include_amounts=True, activity=activity, wo_number=wo_number)
        write_declaration(wb, dc_sites, dc_number, formats, activity=activity, wo_number=wo_number)
        write_annexure_and_reco(wb, dc_sites, dc_number, formats, activity=activity, wo_number=wo_number)

    # Copy template's Main WCC
    ref_template = args.template
    if not ref_template or not os.path.exists(ref_template):
        script_dir = os.path.dirname(os.path.abspath(__file__))
        ref_template = os.path.join(script_dir, '..', 'Backend_Portal', 'templates', 'billing_template.xlsx')
        if not os.path.exists(ref_template):
            ref_template = os.path.abspath(os.path.join(script_dir, 'billing_template.xlsx'))
        if not os.path.exists(ref_template):
            ref_template = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/Backend_Portal/templates/billing_template.xlsx'
        
    inject_main_wcc_template(output_path, ref_template, dc_sites, dc_number, wo_number)
    print(f"SUCCESS: Generated OOP Billing sheet at: {output_path}")

if __name__ == "__main__":
    main()
