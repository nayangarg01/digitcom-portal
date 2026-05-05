import openpyxl

generated_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/Billing/DC0111_A6B6_Billing_Final_Test.xlsx'

try:
    wb = openpyxl.load_workbook(generated_path, data_only=True)
    print(f"Sheets in generated file: {wb.sheetnames}")
    
    for sheet_name in ['Annexture-A6', 'Annexture-B6', 'Reco-A6', 'Reco-B6']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n--- {sheet_name} Headers ---")
            row1 = next(ws.iter_rows(max_row=1, values_only=True))
            print(row1)
            # Check row labels
            row2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
            print(row2[:5]) # Show first 5 columns of labels
        else:
            print(f"Sheet {sheet_name} NOT FOUND")
except Exception as e:
    print(f"Error: {e}")
