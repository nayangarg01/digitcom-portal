import openpyxl
import json

wb = openpyxl.load_workbook('Backend_Portal/scripts/MASTER_JMS_TEMPLATE.xlsx', data_only=True)
ws = wb['JMS']

items = []
for r in range(16, 28):
    item_id = str(ws.cell(row=r, column=1).value or "").strip()
    desc = str(ws.cell(row=r, column=2).value or "").strip()
    uom = str(ws.cell(row=r, column=3).value or "").strip()
    rate = str(ws.cell(row=r, column=28).value or "").strip()
    if item_id:
        items.append({"sap": item_id, "desc": desc, "uom": uom, "rate": rate})

# Check rows 26 and 27 which are hardcoded in python script
items.append({"sap": "EXTRA VISIT", "desc": "EXTRA VISIT", "uom": "EA", "rate": "1000"})
items.append({"sap": "POLE MOUNT", "desc": "POLE MOUNT", "uom": "EA", "rate": "500"})

print(json.dumps(items, indent=2))
