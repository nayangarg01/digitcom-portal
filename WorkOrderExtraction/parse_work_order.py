import pdfplumber
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import argparse
import sys
import os

def parse_work_order(pdf_path, output_excel):
    print(f"Parsing {pdf_path}...")
    
    records = []
    
    current_site_index = None
    current_site_id = None
    current_site_value = None
    current_line_item = None
    
    # Global WO Summary
    wo_summary = {
        "Vendor Name": None,
        "Work Order No.": None,
        "Work Order Date": None,
        "WO Period From": None,
        "WO Period To": None,
        "Value of Work (INR)": None,
        "CGST (INR)": None,
        "SGST (INR)": None,
        "IGST (INR)": None,
        "Total Order Value (INR)": None
    }
    
    # Global Regex Patterns
    vendor_name_pattern = re.compile(r'^(.*?)\s+Date\s*:')
    wo_no_pattern = re.compile(r'Work OrderNo\.\s*:\s*(\S+)')
    wo_date_pattern = re.compile(r'Date\s*:\s*([0-9]{2}\.[0-9]{2}\.[0-9]{4})')
    wo_from_pattern = re.compile(r'WOPeriod From DT\s*:\s*([0-9]{2}\.[0-9]{2}\.[0-9]{4})')
    wo_to_pattern = re.compile(r'To DT\s*:\s*([0-9]{2}\.[0-9]{2}\.[0-9]{4})')
    val_of_work_pattern = re.compile(r'(TotalValueofWork|ValueofWork|RevisedTotalWorkOrderPrice)\s+INR\s+([0-9,.]+)')
    cgst_pattern = re.compile(r'(TotalCGST|CGST)\s+INR\s+([0-9,.]+)')
    sgst_pattern = re.compile(r'(TotalSGST|SGST)\s+INR\s+([0-9,.]+)')
    igst_pattern = re.compile(r'(TotalIGST|IGST)\s+INR\s+([0-9,.]+)')
    total_val_pattern = re.compile(r'(RevisedTotalWorkOrderPrice|TOTALORDERVALUE)\s+INR\s+([0-9,.]+)')

    # Line Item Regex Patterns
    site_pattern = re.compile(r'^([0-9]+)\s+(\S+)\s+([0-9]+)\s+AU')
    site_val_pattern = re.compile(r'ValueofWork\s+INR/AU\s+([0-9,.]+)')
    item_pattern = re.compile(r'^([0-9]+)\s+([0-9]{7})\s+(.*?)\s+([0-9,.]+)\s*([A-Za-z]+)\s*-\s*(.*)')
    item_val_pattern = re.compile(r'Netvalueofitem\s+([0-9,.]+)\s+([0-9,.]+)')
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            print(f"Total Pages to process: {total_pages}")
            
            for i in range(total_pages):
                page = pdf.pages[i]
                text = page.extract_text()
                
                # Progress logging
                if (i + 1) % 10 == 0 or (i + 1) == total_pages:
                    print(f"Processing Page {i + 1}/{total_pages}...")

                if not text:
                    continue
                
                lines = text.split('\n')
                for line in lines:
                    line = line.strip()
                    
                    # Extract Summary Data
                    if not wo_summary["Vendor Name"]:
                        m = vendor_name_pattern.search(line)
                        if m: wo_summary["Vendor Name"] = m.group(1).strip()
                    if not wo_summary["Work Order No."]:
                        m = wo_no_pattern.search(line)
                        if m: wo_summary["Work Order No."] = m.group(1)
                    if not wo_summary["Work Order Date"]:
                        m = wo_date_pattern.search(line)
                        if m: wo_summary["Work Order Date"] = m.group(1)
                    if not wo_summary["WO Period From"]:
                        m = wo_from_pattern.search(line)
                        if m: wo_summary["WO Period From"] = m.group(1)
                    if not wo_summary["WO Period To"]:
                        m = wo_to_pattern.search(line)
                        if m: wo_summary["WO Period To"] = m.group(1)

                    # 0. Match Summary Header Info
                    if not wo_summary["Value of Work (INR)"] or "TotalValueofWork" in line or "ValueofWork" in line:
                        m = val_of_work_pattern.search(line)
                        if m: 
                            val = float(m.group(2).replace(',', ''))
                            # If we find TotalValueofWork, it's definitely the subtotal we want
                            if "TotalValueofWork" in m.group(1) or "ValueofWork" in m.group(1):
                                wo_summary["Value of Work (INR)"] = val
                            elif not wo_summary["Value of Work (INR)"]:
                                wo_summary["Value of Work (INR)"] = val

                    if not wo_summary["CGST (INR)"] or "TotalCGST" in line:
                        m = cgst_pattern.search(line)
                        if m: wo_summary["CGST (INR)"] = float(m.group(2).replace(',', ''))

                    if not wo_summary["SGST (INR)"] or "TotalSGST" in line:
                        m = sgst_pattern.search(line)
                        if m: wo_summary["SGST (INR)"] = float(m.group(2).replace(',', ''))

                    if not wo_summary["IGST (INR)"] or "TotalIGST" in line:
                        m = igst_pattern.search(line)
                        if m: wo_summary["IGST (INR)"] = float(m.group(2).replace(',', ''))

                    if not wo_summary["Total Order Value (INR)"] or "RevisedTotalWorkOrderPrice" in line:
                        m = total_val_pattern.search(line)
                        if m: wo_summary["Total Order Value (INR)"] = float(m.group(2).replace(',', ''))

                    # 1. Match Site
                    site_match = site_pattern.match(line)
                    if site_match:
                        current_site_index = site_match.group(1)
                        raw_id = site_match.group(2)
                        current_raw_site_id = raw_id # Keep the full string for the matrix sheet
                        current_site_id = raw_id.rsplit('_', 1)[-1]
                        current_site_value = None
                        continue
                        
                    # 2. Match Site Value
                    if current_site_id and not current_site_value:
                        val_match = site_val_pattern.search(line)
                        if val_match:
                            current_site_value = val_match.group(1).replace(',', '')
                            continue
                            
                    # 3. Match Line Item
                    item_match = item_pattern.match(line)
                    if item_match:
                        current_line_item = {
                            "Site Index": current_site_index,
                            "Site ID": current_site_id,
                            "Raw Site ID": current_raw_site_id,
                            "Site Base Value (INR)": float(current_site_value) if current_site_value else 0.0,
                            "Line Item No": item_match.group(1),
                            "Item Code": item_match.group(2),
                            "Item Description": item_match.group(3).strip(),
                            "Quantity": float(item_match.group(4).replace(',', '')),
                            "Unit": item_match.group(5),
                            "Unit Rate": None,
                            "Total Amount": None
                        }
                        continue
                        
                    # 4. Match Line Item Value
                    if current_line_item and current_line_item["Unit Rate"] is None:
                        item_val_match = item_val_pattern.search(line)
                        if item_val_match:
                            current_line_item["Unit Rate"] = float(item_val_match.group(1).replace(',', ''))
                            current_line_item["Total Amount"] = float(item_val_match.group(2).replace(',', ''))
                            records.append(current_line_item)
                            current_line_item = None
                
                # Memory optimization: flush page after processing
                page.flush_cache()
                            
    except Exception as e:
        print(f"Error reading PDF: {e}")
        sys.exit(1)
        
    print(f"Extracted {len(records)} line items.")
    
    if not records:
        print("No records found. Regex might need adjustment.")
        return
        
    # Verification
    total_extracted_val = sum(r["Total Amount"] for r in records if r["Total Amount"] is not None)
    expected_val = wo_summary["Total Order Value (INR)"] or 0.0
    # Note: "Value of Work" usually doesn't include GST. 
    # Let's compare with Value of Work (INR) if possible.
    val_of_work = wo_summary["Value of Work (INR)"] or 0.0
    
    print(f"--- Verification ---")
    print(f"Sum of Line Items: {total_extracted_val:,.2f}")
    print(f"Value of Work (from PDF): {val_of_work:,.2f}")
    diff = abs(total_extracted_val - val_of_work)
    if diff < 1.0: # Account for minor rounding
        print("SUCCESS: Extraction sum matches Value of Work!")
    else:
        print(f"WARNING: Mismatch detected! Difference: {diff:,.2f}")
    print(f"--------------------")

    # Generate Excel using openpyxl
    print(f"Saving to {output_excel}...")
    wb = Workbook()
    
    # Create the Work Order Details sheet first
    ws_details = wb.active
    ws_details.title = "Work Order Details"
    
    ws_details.append(["Work Order Detail", "Value"])
    for key, val in wo_summary.items():
        ws_details.append([key, val])
        
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws_details[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    for row in range(2, len(wo_summary) + 2):
        ws_details[f"A{row}"].font = Font(bold=True)
        if isinstance(ws_details[f"B{row}"].value, (int, float)):
            ws_details[f"B{row}"].number_format = '#,##0.00'
            
    ws_details.column_dimensions['A'].width = 30
    ws_details.column_dimensions['B'].width = 30

    # Now create the line items sheet
    ws = wb.create_sheet(title="Parsed Work Order")
    
    # Define columns to exclude from matrix sheet but include here
    all_headers = list(records[0].keys())
    # Reorder or select specific headers if needed, but we keep them all
    ws.append(all_headers)
    
    for record in records:
        ws.append([record[k] for k in all_headers])
    
    # Formatting
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    ws.freeze_panes = 'A2'
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        if adjusted_width > 50: adjusted_width = 50
        ws.column_dimensions[column].width = adjusted_width
        
    # Highlight Total Amount Column
    total_amt_col_idx = all_headers.index("Total Amount") + 1
    total_amt_col_letter = get_column_letter(total_amt_col_idx)
    highlight_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    
    for row in range(2, len(records) + 2):
        cell = ws[f"{total_amt_col_letter}{row}"]
        cell.fill = highlight_fill
        cell.number_format = '#,##0.00'
        
    # Format other numbers
    site_val_col_idx = all_headers.index("Site Base Value (INR)") + 1
    site_val_col_letter = get_column_letter(site_val_col_idx)
    rate_col_idx = all_headers.index("Unit Rate") + 1
    rate_col_letter = get_column_letter(rate_col_idx)
    
    for row in range(2, len(records) + 2):
        ws[f"{site_val_col_letter}{row}"].number_format = '#,##0.00'
        ws[f"{rate_col_letter}{row}"].number_format = '#,##0.00'
        
    # Merge Site columns for identical sites
    start_row = 2
    site_id_col_idx = all_headers.index("Site ID") + 1
    current_site = None
    
    for row in range(2, len(records) + 3):
        if row <= len(records) + 1:
            site_id = ws.cell(row=row, column=site_id_col_idx).value
        else: site_id = None
            
        if site_id != current_site:
            if current_site is not None:
                end_row = row - 1
                if end_row > start_row:
                    for col_idx in [1, 2, 3, 4]: # Site Index, Site ID, Raw Site ID, Site Base Value
                        ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
                        ws.cell(row=start_row, column=col_idx).alignment = Alignment(horizontal='center', vertical='top')
            current_site = site_id
            start_row = row

    # --- NEW: Site-wise Summary Matrix Sheet ---
    ws_matrix = wb.create_sheet(title="Site-wise Summary")
    
    unique_items = sorted(list(set(r["Item Description"] for r in records)))
    matrix_headers = ['DESCRIPTION', 'SAP ID'] + unique_items
    ws_matrix.append(matrix_headers)
    
    # Group records by Site
    sites_data = {}
    for r in records:
        key = (r["Raw Site ID"], r["Site ID"])
        if key not in sites_data:
            sites_data[key] = {item: 0.0 for item in unique_items}
        sites_data[key][r["Item Description"]] += r["Quantity"]
        
    for (raw_id, sap_id), items in sites_data.items():
        row_data = [raw_id, sap_id]
        for item in unique_items:
            row_data.append(items[item] if items[item] > 0 else None)
        ws_matrix.append(row_data)
        
    # Matrix Sheet Formatting
    for cell in ws_matrix[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    for col in ws_matrix.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws_matrix.column_dimensions[column].width = min(max_length + 2, 50)

    wb.save(output_excel)
    print(f"Success! Saved formatted Excel to {output_excel}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Parse Work Order PDF to Excel")
    parser.add_argument('pdf_path', nargs='?', default='DIGITCOM_630344375.pdf', help='Path to input PDF')
    parser.add_argument('--output', '-o', default='WorkOrder_Summary.xlsx', help='Output Excel file path')
    
    args = parser.parse_args()
    parse_work_order(args.pdf_path, args.output)
