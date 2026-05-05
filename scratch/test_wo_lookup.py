import openpyxl
import os

def get_wo_number(master_path, dc_number):
    try:
        wb = openpyxl.load_workbook(master_path, data_only=True)
        ws = wb.active # Assuming first sheet is the one
        
        # We found:
        # Col 14 (Index 14 if starting at 0, or Col O): WO
        # Col 47 (Index 47 if starting at 0, or Col AV): BILLING FILE
        
        # Let's find the indices dynamically to be safe
        headers = [str(c.value).upper().strip() if c.value else "" for c in ws[2]] # Row 2 has headers
        
        billing_col_idx = None
        wo_col_idx = None
        
        for i, h in enumerate(headers):
            if "BILLING FILE" in h or "DC NUMBER" in h:
                billing_col_idx = i
            if h == "WO":
                wo_col_idx = i
        
        if billing_col_idx is None: billing_col_idx = 47
        if wo_col_idx is None: wo_col_idx = 14
        
        for row in ws.iter_rows(min_row=3, values_only=True): # Data starts at Row 3
            if str(row[billing_col_idx]).strip().upper() == dc_number.upper():
                return str(row[wo_col_idx]).strip()
        
        return "N/A"
    except Exception as e:
        print(f"Error looking up WO: {e}")
        return "N/A"

if __name__ == "__main__":
    master_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/FinaliseBillingFormat/MASTER TRACKER FOR BILLING-AIRFIBER-RJST (1).xlsx'
    dc = "DC0105"
    wo = get_wo_number(master_path, dc)
    print(f"WO for {dc}: {wo}")
