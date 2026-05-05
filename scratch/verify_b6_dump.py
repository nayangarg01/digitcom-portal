import openpyxl

file_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/InputBilling/MIN DUMP-RJST TILL 31 MAR 26.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb['B6 DUMP']
    headers = [str(h).strip() for h in next(ws.iter_rows(max_row=1, values_only=True))]
    enb_col = headers.index('ENB ID')
    common_col = headers.index('COMMON ID')
    
    # Target IDs for Site 1 of DC0111
    target_ids = ['I-RJ-RSNG-ENB-B001', 'I-RJ-RSNG-ENB-V002']
    
    print("\nMatching rows in B6 DUMP for Site 1 ends:")
    for row in ws.iter_rows(min_row=2, values_only=True):
        enb_id = str(row[enb_col]).strip()
        if any(t in enb_id for t in target_ids):
            print(f"ENB ID: {row[enb_col]} | COMMON ID: {row[common_col]}")
            
except Exception as e:
    print(f"Error: {e}")
