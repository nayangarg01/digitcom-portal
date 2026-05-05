import openpyxl

file_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/PerformaInvoiceWork/Performa_Test_A6B6_v3.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws_sum = wb['Summary sheet1']
    print("\n--- A6+B6 Summary Sheet Sample Rows (Row 1-10) ---")
    for row in ws_sum.iter_rows(max_row=10, values_only=True):
        print(row)
        
    ws1 = wb['1']
    print("\n--- A6+B6 Sheet '1' Row 6 (Nature check) ---")
    print(list(ws1.iter_rows(min_row=2, max_row=2, values_only=True))[0])
except Exception as e:
    print(f"Error: {e}")
