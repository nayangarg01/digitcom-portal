import openpyxl

file_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/PerformaInvoiceWork/Performa_Test_v1.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    
    ws1 = wb['1']
    print("\n--- Sheet '1' Sample Rows (Row 1-10) ---")
    for row in ws1.iter_rows(max_row=10, values_only=True):
        print(row)
        
    ws_sum = wb['Summary sheet1']
    print("\n--- Summary Sheet Sample Rows (Row 1-5) ---")
    for row in ws_sum.iter_rows(max_row=5, values_only=True):
        print(row)
except Exception as e:
    print(f"Error: {e}")
