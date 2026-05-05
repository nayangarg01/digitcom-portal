import openpyxl

file_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/InputBilling/DIGITCOM_AIRFIBER_DC0111_JPUR_12-MAR-26 23-MAR-26 & 25-MAR-26_A6+B6.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    if 'Annexture-B6' in wb.sheetnames:
        ws = wb['Annexture-B6']
        print("\n--- Annexture-B6 Headers (Row 2) ---")
        row2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        print(row2)
    else:
        print("Sheet 'Annexture-B6' not found")
except Exception as e:
    print(f"Error: {e}")
