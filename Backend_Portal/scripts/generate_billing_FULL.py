import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.cell.cell import MergedCell
import sys
import argparse
import os
from copy import copy
from openpyxl.drawing.image import Image as OpenpyxlImage

def safe_float(val):
    """
    Safely converts a value to float, handling 'NR', 'NA', or other text by returning 0.0.
    """
    if pd.isna(val) or str(val).strip() == "" or str(val).strip().upper() in ["NR", "NA", "N.A", "-"]:
        return 0.0
    try:
        return float(val)
    except:
        return 0.0

def load_master_data(master_path, dc_number):
    try:
        # Load KM SHEET (first sheet)
        df_full = pd.read_excel(master_path, header=None)
        
        if df_full.empty:
            print("ERROR: Master Tracker file is empty.")
            return None, None
            
        # 1. Discover the 'BILLING FILE' column index (where DC numbers live)
        # Search row 1 (0-indexed) which usually contains headers in this file
        dc_col_idx = 0
        found_header = False
        
        # Safely extract headers from Row 1
        raw_row1 = df_full.iloc[1].tolist() if len(df_full) > 1 else []
        for i, h in enumerate(raw_row1):
            h_str = str(h).upper().strip()
            if "BILLING FILE" in h_str or "DC NUMBER" in h_str:
                dc_col_idx = i
                found_header = True
                break
        
        if not found_header:
            print("WARNING: 'BILLING FILE' column not found in Row 2. Defaulting to column ID search.")
            # Fallback: Search all columns for the DC number if header hit fails
            for c_idx in range(df_full.shape[1]):
                col_sample = df_full.iloc[:, c_idx].astype(str).str.upper().tolist()
                if dc_number.upper() in col_sample:
                    dc_col_idx = c_idx
                    break

        # 2. Filter Sites based on the discovered Billing column
        df_sites = df_full[df_full.iloc[:, dc_col_idx].astype(str).str.strip().str.upper() == dc_number.upper()].copy()
        
        if df_sites.empty:
            print(f"DATA ERROR: No site records found for DC Number '{dc_number}' in column {dc_col_idx}.")
            return None, None
            
        # FIX: Assign proper column names so .upper() works in downstream functions
        # Use the raw header row (Row 1) for column names
        raw_headers = df_full.iloc[1].tolist()
        df_sites.columns = [str(h).strip() for h in raw_headers]
        
        # 3. Discover Item Code mapping (Dual-Key: Check Row 0 AND Row 1)
        # Template uses SAP Codes, Master has them in Row 0 (Index 0)
        code_to_col_idx = {}
        row0 = df_full.iloc[0].tolist()
        row1 = df_full.iloc[1].tolist()
        
        for i in range(len(row0)):
            # Capture SAP Codes (Row 0)
            sap = str(row0[i]).split('.')[0].strip()
            if sap and sap != 'nan':
                code_to_col_idx[sap] = i
                
            # Capture Descriptions (Row 1)
            desc = str(row1[i]).strip()
            if desc and desc != 'nan':
                code_to_col_idx[desc] = i
                
        return df_sites, code_to_col_idx
    except Exception as e:
        print(f"Error loading Master: {e}")
        import traceback
        traceback.print_exc()
        return None, None

def generate_wcc_sheet(df_sites, wb):
    """Injects the filtered sites into the WCC Template sheet using eNBsiteID."""
    if 'WCC' not in wb.sheetnames:
        print("WCC sheet missing.")
        return False
    ws = wb['WCC']
    
    # Locate headers in WCC
    header_row_idx = None
    cols_map = {}
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and 'GIS SECTOR_ID' in str(row):
            header_row_idx = r_idx
            for c_idx, val in enumerate(row, start=1):
                if val: cols_map[str(val).strip()] = c_idx
            break
            
    if not header_row_idx:
        print("Could not locate WCC headers.")
        return False

    start_row = header_row_idx + 1
    base_styles = {}
    for c_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=start_row, column=c_idx)
        base_styles[c_idx] = {
            'font': copy(cell.font), 'border': copy(cell.border), 'fill': copy(cell.fill),
            'number_format': cell.number_format, 'alignment': copy(cell.alignment)
        }

    # Manage row expansion
    if len(df_sites) > 22:
        ws.insert_rows(start_row + 22, amount=(len(df_sites) - 22))
    elif len(df_sites) < 22:
        ws.delete_rows(start_row + len(df_sites), amount=(22 - len(df_sites)))
        
    aktbc_col = next((c for c in df_sites.columns if 'CHRG EXTRA TRANSPORT' in c.upper() or 'AKTBC' == c.upper()), None)

    for i, (_, row) in enumerate(df_sites.iterrows()):
        curr_row = start_row + i
        def get_val(matcher):
            c_name = next((c for c in df_sites.columns if matcher.upper() in c.upper()), None)
            return row[c_name] if c_name else ""
            
        mapping = [
            ('Sr. No', i + 1),
            ('ENB SITE ID', get_val('ENBSITEID')),
            ('PMP SAP ID', get_val('PMP ID')),
            ('GIS SECTOR_ID', get_val('GIS SECTOR')),
            ('No of Sectors', get_val('NO OF SECTOR')),
            ('Tower type', get_val('Tower type')),
            ('JC', get_val('JC')), ('WH', get_val('WH')), ('VEHICLE NO', get_val('VEHICLE NO')),
            ('MIN  NO', get_val('MIN NO')), ('MIN Date', get_val('MIN DATE')), ('Completion Date', get_val('Completion Date')),
            ('REMARKS', "RFS DONE" if pd.notna(get_val('Completion Date')) and str(get_val('Completion Date')) != "" else ""),
            ('ACTUAL KM', safe_float(row[aktbc_col]) if aktbc_col else 0.0),
            ('KM-50', safe_float(get_val('KM-50(for a6+b6-100)'))),
            ('KM IN WO', safe_float(get_val('KM IN WO'))),
            ('A6 in wo', safe_float(get_val('A6 in wo'))),
            ('cpri in wo', safe_float(get_val('cpri in wo'))),
            ('power in wo', safe_float(get_val('power in wo'))),
            ('puff sealant in wo', safe_float(get_val('puff sealant in wo'))),
            ('termination in wo', safe_float(get_val('termination in wo'))),
            ('EXTRA VISIT IN WO', safe_float(get_val('EXTRA VISIT IN WO'))),
            ('Polemount in wo', safe_float(get_val('Polemount in wo')))
        ]
        
        act_km = next((v for k, v in mapping if k == 'ACTUAL KM'), 0.0)
        wo_km = next((v for k, v in mapping if k == 'KM IN WO'), 0.0)
        mapping.append(('GAP', act_km - wo_km))
        mapping.append(('USED KM IN WCC', act_km if act_km <= wo_km else wo_km))

        for c_idx in range(1, ws.max_column + 1):
            c = ws.cell(row=curr_row, column=c_idx)
            styles = base_styles.get(c_idx)
            if styles:
                c.font, c.border, c.fill = copy(styles['font']), copy(styles['border']), copy(styles['fill'])
                c.number_format, c.alignment = styles['number_format'], copy(styles['alignment'])

        for col_name, val in mapping:
            c_target = next((cols_map[k] for k in cols_map if col_name.upper() in k.upper()), None)
            if c_target:
                target_cell = ws.cell(row=curr_row, column=c_target)
                target_cell.value = val.to_pydatetime() if isinstance(val, pd.Timestamp) else val
                
    # Update totals
    last_r = start_row + len(df_sites) - 1
    for sc in ['ACTUAL KM', 'USED KM IN WCC']:
        c_i = next((cols_map[k] for k in cols_map if sc.upper() in k.upper()), None)
        if c_i:
            let = get_column_letter(c_i)
            ws.cell(row=last_r+1, column=c_i).value = f"=SUM({let}{start_row}:{let}{last_r})"
    return True

def generate_jms_sheet(df_sites, code_to_col_idx, wb):
    if 'JMS' not in wb.sheetnames:
        return False
    ws = wb['JMS']
    
    # 1. Coordinate Discovery (Non-Invasive)
    coord = {}
    for r in range(1, 40):
        v = str(ws.cell(row=r, column=2).value).upper() if ws.cell(row=r, column=2).value else ""
        if "COUNT" in v: coord['COUNT_ROW'] = r
        if "SITE ID" in v: coord['SITE_ROW'] = r
        if "SITE TYPE" in v: coord['TOWER_ROW'] = r
        if "SECTORS" in v: coord['SECTOR_ROW'] = r
        # Find TOTAL row to define matrix bottom
        if r > 15 and "TOTAL" in v:
            coord['TOTAL_ROW'] = r
            break
            
    # Defaults if labels not found
    C_ROW = coord.get('COUNT_ROW', 9)
    S_ROW = coord.get('SITE_ROW', 10)
    T_ROW = coord.get('TOWER_ROW', 11)
    SEC_ROW = coord.get('SECTOR_ROW', 12)
    MAX_R = coord.get('TOTAL_ROW', 26)
    START_COL = 4  # Column D
    
    # Identify Template's Summary columns (Total Qty, Rate, Amount)
    # Usually they follow the 22-column matrix
    # Based on DC0105: Matrix is D to Y (Index 4 to 25). AA (27) is Total Qty.
    TOTAL_QTY_COL = 27
    
    # Carry over Column Widths for Summary Columns (Total Qty, Rate, Amount)
    # Col 27, 28, 29 in original are AA, AB, AC
    orig_widths = {
        'qty': ws.column_dimensions[get_column_letter(27)].width,
        'rate': ws.column_dimensions[get_column_letter(28)].width,
        'amt': ws.column_dimensions[get_column_letter(29)].width
    }

    # 2. Matrix Management (Only adding if absolutely necessary)
    num_sites = len(df_sites)
    if num_sites > 22:
        # Insert columns BEFORE Total Quantity to maintain formulas
        ws.insert_cols(27, amount=(num_sites - 22))
        TOTAL_QTY_COL = 27 + (num_sites - 22)
    elif num_sites < 22:
        # We delete columns within the site range only
        ws.delete_cols(START_COL + num_sites, amount=(22 - num_sites))
        TOTAL_QTY_COL = 27 - (22 - num_sites)
    else:
        TOTAL_QTY_COL = 27

    # 3. Data Injection (ONLY TOUCH SITES MATRIX)
    for i, (_, site_row) in enumerate(df_sites.iterrows()):
        curr_col = START_COL + i
        
        # Headers
        ws.cell(row=C_ROW, column=curr_col).value = i + 1
        # Site ID Header at Verdana 35 Bold
        cell_id = ws.cell(row=S_ROW, column=curr_col)
        cell_id.value = str(site_row['eNBsiteID']).strip()
        cell_id.font = Font(name='Verdana', size=35, bold=True)
        cell_id.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
        
        ws.cell(row=T_ROW, column=curr_col).value = str(site_row['Tower type']).strip()
        ws.cell(row=SEC_ROW, column=curr_col).value = site_row['NO OF SECTOR']
        
        # Matrix Values (Items 16 to MAX_R-1)
        for r_idx in range(16, MAX_R):
            item_id_val = ws.cell(row=r_idx, column=1).value
            try:
                item_code = str(int(item_id_val)) if item_id_val else ""
            except:
                item_code = str(item_id_val).strip() if item_id_val else ""
            
            if item_code in code_to_col_idx:
                src_col = code_to_col_idx[item_code]
                val = site_row.iloc[src_col]
                ws.cell(row=r_idx, column=curr_col).value = safe_float(val)
            else:
                # If Item code row is mentioned but not found in tracker, reset to 0
                ws.cell(row=r_idx, column=curr_col).value = 0.0
            
            # Apply Centering to all matrix data cells
            ws.cell(row=r_idx, column=curr_col).alignment = Alignment(horizontal='center', vertical='center')

from copy import copy

def copy_cell_style(src_cell, dst_cell):
    """Clones the exact DNA (Font, Border, Fill, Alignment, etc.) from one cell to another."""
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = copy(src_cell.number_format)
        dst_cell.alignment = copy(src_cell.alignment)

def populate_main_matrix(sheet_name, df_sites, code_to_col_idx, wb, values_only=False):
    """
    Universal Engine for JMS, Abstract, and BOQ. 
    Writes the full site matrix and clones Style DNA from Row 25.
    """
    if sheet_name not in wb.sheetnames: return False
    ws = wb[sheet_name]
    
    # Header Update Logic
    amt_col = 30
    for c in range(4, 50):
        if "AMOUNT" in str(ws.cell(row=12, column=c).value).upper():
            amt_col = c
            break
    
    # 0. Headers (Ensuring exact alignment)
    try:
        wo_number = get_wo_number(os.getenv('MASTER_PATH', 'MASTER.xlsx'), df_sites.iloc[0, 47] if not df_sites.empty else "")
    except:
        wo_number = "P14/630330726"
        
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=amt_col)
    ws.cell(row=1, column=1).value = f"Work Order No : {wo_number}"
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    
    mid = amt_col // 2
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=mid)
    ws.cell(row=2, column=1).value = "Contractor Name: DIGITCOM INDIA TECHNOLOGIES"
    ws.cell(row=2, column=1).alignment = Alignment(horizontal='left', vertical='center')
    
    ws.merge_cells(start_row=2, start_column=mid+1, end_row=2, end_column=amt_col)
    ws.cell(row=2, column=mid+1).value = "Work Order Dated: 03-10-2025"
    ws.cell(row=2, column=mid+1).alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=amt_col)
    ws.cell(row=3, column=1).value = "WO for Airspan A6 and C6 Radios for Airfiber"
    ws.cell(row=3, column=1).alignment = Alignment(horizontal='left', vertical='center')
    
    # Dates
    date_col = next((c for c in df_sites.columns if 'COMPLETION' in c.upper() or 'RFS DATE' in c.upper() or 'DATE' in c.upper()), None)
    min_date = df_sites[date_col].min().strftime('%d-%b-%y').upper() if date_col and not df_sites[date_col].isna().all() else "N/A"
    max_date = df_sites[date_col].max().strftime('%d-%b-%y').upper() if date_col and not df_sites[date_col].isna().all() else "N/A"

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=amt_col)
    ws.cell(row=4, column=1).value = f"Service Done From Date: {min_date}"
    ws.cell(row=4, column=1).alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=amt_col)
    ws.cell(row=5, column=1).value = f"Service Done To Date: {max_date}"
    ws.cell(row=5, column=1).alignment = Alignment(horizontal='center', vertical='center')

    # 0. Safety: Unmerge cells in the data matrix area
    data_merge_ranges = []
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row >= 11 and merged_range.max_row <= 30:
            data_merge_ranges.append(merged_range)
    for m_range in data_merge_ranges:
        ws.unmerge_cells(str(m_range))

    # 1. Discovery & Cleanup
    num_sites = len(df_sites)
    START_COL, C_ROW, S_ROW, T_ROW, SEC_ROW = 4, 11, 12, 13, 14
    MAX_R = 28 # Items 16-27, Total 28
    
    # FIND TEMPLATE TOTALS POSITION (The Anchor)
    orig_total_col = None
    for c in range(START_COL + 1, 60):
        h = str(ws.cell(row=12, column=c).value).upper().strip()
        if "TOTAL QUANTITY" in h: 
            orig_total_col = c
            break
    
    # DYNAMIC SCALING (FIXED FOOTPRINT): Only delete if oversite, Clear if undersite
    if orig_total_col:
        num_template_site_cols = orig_total_col - START_COL
        if num_sites < num_template_site_cols:
            # UNDERSITE: No Delete! Clear unused columns instead to save signatures.
            # This keeps 'Total Quantity' at AA (Column 27)
            for c_clear in range(START_COL + num_sites, orig_total_col):
                for r_clear in range(11, MAX_R):
                    cell = ws.cell(row=r_clear, column=c_clear)
                    cell.value = None
                    cell.fill = PatternFill(fill_type=None)
        elif num_sites > num_template_site_cols:
            # OVERSITE: Insert columns to push totals to the right
            num_to_add = num_sites - num_template_site_cols
            ws.insert_cols(orig_total_col, num_to_add)
    
    # RE-FIND POSITIONS AFTER SCALING
    TOTAL_QTY_COL, rate_col, amt_col = None, None, None
    for c in range(START_COL, 65):
        h = str(ws.cell(row=12, column=c).value).upper().strip()
        if "TOTAL QUANTITY" in h: TOTAL_QTY_COL = c
        elif "RATE AS PER SOW" in h: rate_col = c
        elif "AMOUNT" in h: amt_col = c
    
    if not TOTAL_QTY_COL: TOTAL_QTY_COL = START_COL + num_sites
    if not rate_col: rate_col = TOTAL_QTY_COL + 1
    if not amt_col: amt_col = rate_col + 1

    # 2. Data Injection
    fcl, lcl = get_column_letter(START_COL), get_column_letter(START_COL + num_sites - 1)
    
    # Pre-populate Mandatory Labels for Abstract/BOQ (if missing)
    if sheet_name in ['Abstract', 'BOQ']:
        ws.cell(row=26, column=2).value = "EXTRA VISIT"
        ws.cell(row=27, column=2).value = "POLE MOUNT"
        ws.cell(row=26, column=3).value = "EA"
        ws.cell(row=27, column=3).value = "EA"

    for i, (_, site_row) in enumerate(df_sites.iterrows()):
        curr_col = START_COL + i
        ws.cell(row=C_ROW, column=curr_col).value = i + 1
        
        # PMP ID Authority (Row 12)
        cell_id = ws.cell(row=S_ROW, column=curr_col)
        cell_id.value = str(site_row['PMP ID']).strip()
        cell_id.font = Font(name='Verdana', size=35, bold=True)
        cell_id.alignment = Alignment(text_rotation=90, horizontal='center', vertical='center')
        
        ws.cell(row=T_ROW, column=curr_col).value = str(site_row['Tower type']).strip()
        ws.cell(row=SEC_ROW, column=curr_col).value = site_row['NO OF SECTOR']
        
        for r in range(16, MAX_R):
            item_id_val = ws.cell(row=r, column=1).value
            try: item_code = str(int(item_id_val)) if item_id_val else ""
            except: item_code = str(item_id_val).strip() if item_id_val else ""
            
            # Special bypass for mandatory rows
            if not item_code and r == 26: item_code = "EXTRA VISIT"
            if not item_code and r == 27: item_code = "POLE MOUNT"
            
            val = site_row.iloc[code_to_col_idx[item_code]] if item_code in code_to_col_idx else 0.0
            cell_data = ws.cell(row=r, column=curr_col)
            cell_data.value = safe_float(val)
            cell_data.alignment = Alignment(horizontal='center', vertical='center')

    # 3. STYLE DNA MIRRORING, RATES & SUMMATIONS (Strictly mirror Row 25 down to 27)
    for r in range(16, MAX_R + 1):
        # Mirror Style from Row 25
        for c in range(1, amt_col + 1):
            copy_cell_style(ws.cell(row=25, column=c), ws.cell(row=r, column=c))
            
        if r < MAX_R:
            # ASSIGN SPECIFIC RATES
            item_id_val = ws.cell(row=r, column=1).value
            try: item_code = str(int(item_id_val)) if item_id_val else ""
            except: item_code = str(item_id_val).strip() if item_id_val else ""
            
            if "EXTRA VISIT" in item_code.upper() or r == 26:
                ws.cell(row=r, column=rate_col).value = 1000
            elif "POLE MOUNT" in item_code.upper() or r == 27:
                ws.cell(row=r, column=rate_col).value = 500
            
            # Row-wise Summations
            f_col = get_column_letter(START_COL)
            if r == 27: curr_code = "POLE MOUNT"

            if values_only:
                row_sum = sum([safe_float(df_sites.iloc[i].iloc[code_to_col_idx[curr_code]]) 
                               if curr_code in code_to_col_idx else 0.0 
                               for i in range(num_sites)])
                ws.cell(row=r, column=TOTAL_QTY_COL).value = row_sum
                rate = safe_float(ws.cell(row=r, column=rate_col).value)
                ws.cell(row=r, column=amt_col).value = row_sum * rate
            else:
                ws.cell(row=r, column=TOTAL_QTY_COL).value = f"=SUM({fcl}{r}:{lcl}{r})"
                ws.cell(row=r, column=amt_col).value = f"={get_column_letter(TOTAL_QTY_COL)}{r}*{get_column_letter(rate_col)}{r}"

    # Sectors Total
    ws.cell(row=SEC_ROW, column=TOTAL_QTY_COL).value = f"=SUM({fcl}{SEC_ROW}:{lcl}{SEC_ROW})"

    # Grand Total (Row 28)
    grand_total_row = MAX_R
    for r in range(MAX_R, MAX_R + 10):
        if "TOTAL" in str(ws.cell(row=r, column=2).value).upper():
            grand_total_row = r; break
            
    if values_only:
        v_sum = sum([safe_float(ws.cell(row=rr, column=amt_col).value) for rr in range(16, grand_total_row)])
        ws.cell(row=grand_total_row, column=amt_col).value = v_sum
    else:
        ws.cell(row=grand_total_row, column=amt_col).value = f"=SUM({get_column_letter(amt_col)}16:{get_column_letter(amt_col)}{grand_total_row-1})"

    return True

def generate_annexure_sheet(df_sites, wb, mindump_path=None):
    """
    Builds a 'Clean-Room' Annexure with Size 20 font and Snapshot 2 layout.
    LATEST-DATE LOGIC: For each site, finds the MAX(Date) in MINDUMP and pulls its rows.
    """
    if 'Annexture' not in wb.sheetnames: return False
    ws = wb['Annexture']
    
    # 0. DEEP WIPE: Remove everything below the title to kill ghost formatting
    max_r, max_c = ws.max_row, ws.max_column
    for r in range(2, max_r + 1):
        for c in range(1, max_c + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.font = Font(name='Calibri', size=11)
            cell.alignment = Alignment()

    pmp_ids = df_sites.iloc[:, 1].astype(str).str.strip().tolist()

    try:
        # Priority: 1. Provided Path, 2. Fallback local, 3. Numbers fallback
        if mindump_path and os.path.exists(mindump_path):
            df_dump = pd.read_excel(mindump_path)
        elif os.path.exists('Billing/MINDUMP.xlsx'):
            df_dump = pd.read_excel('Billing/MINDUMP.xlsx')
        else:
            print("ERROR: MINDUMP file not found. Annexure will be empty.")
            df_dump = pd.DataFrame()
            
        if not df_dump.empty:
            df_dump['Site ID'] = df_dump['Site ID'].astype(str).str.strip()
            
            # 1. Per-Site Discovery Loop
            df_all_snapshots = []
            for pid in pmp_ids:
                df_site = df_dump[df_dump['Site ID'] == pid]
                if not df_site.empty:
                    latest_date = df_site['Date'].max()
                    df_snapshot = df_site[df_site['Date'] == latest_date]
                    df_all_snapshots.append(df_snapshot)
            
            if not df_all_snapshots:
                print(f"WARNING: No material found in MINDUMP for sites {pmp_ids}")
                df_filtered = pd.DataFrame()
            else:
                df_filtered = pd.concat(df_all_snapshots)
        else:
            df_filtered = pd.DataFrame()
    except Exception as e:
        print(f"Annexure Error: {e}")
        return False

    if df_filtered.empty:
        # We proceed to make an empty table with zeros later
        pivot = pd.DataFrame(columns=pmp_ids)
    else:
        pivot = df_filtered.pivot_table(index=['SAP Code', 'Material Description'], 
                                       columns='Site ID', values='No. Of Qty', aggfunc='sum').fillna(0)

    # 1. Styles
    medium_side = Side(style='medium', color="000000")
    medium_border = Border(left=medium_side, right=medium_side, top=medium_side, bottom=medium_side)
    F_SIZE = 20
    
    # 2. SITE HEADERS (Row 2, Vertical)
    START_COL = 2
    ws.row_dimensions[2].height = 120
    for i, pmp_id in enumerate(pmp_ids):
        col = START_COL + i
        cell = ws.cell(row=2, column=col)
        cell.value = pmp_id
        cell.font = Font(name='Calibri', size=F_SIZE, bold=True)
        cell.alignment = Alignment(text_rotation=90, horizontal='center', vertical='center', wrap_text=True)
        cell.border = medium_border
        ws.column_dimensions[get_column_letter(col)].width = 7

    num_sites = len(pmp_ids)
    sum_col = START_COL + num_sites
    desc_col = sum_col + 1
    
    # Header Labels for Summary
    header_labels = [(1, "Row Labels"), (sum_col, "GRAND\nTOTAL"), (desc_col, "Material Description")]
    for c, lab in header_labels:
        cell = ws.cell(row=2, column=c)
        cell.value = lab
        cell.font = Font(name='Calibri', size=F_SIZE, bold=True)
        # Use Wrap Text for GRAND TOTAL to keep column narrow
        cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        cell.border = medium_border
        
        if c == desc_col: 
            ws.column_dimensions[get_column_letter(c)].width = 70 # Expanded for long names
        elif c == sum_col:
            ws.column_dimensions[get_column_letter(c)].width = 12

    # 3. Data Infill
    last_r = 2
    for r_idx, (idx_vals, row_vals) in enumerate(pivot.iterrows()):
        curr_row = 3 + r_idx
        last_r = curr_row
        sap_code, mat_desc = idx_vals
        ws.cell(row=curr_row, column=1).value = str(sap_code)
        
        # Calculate Row Sum in Python for RECO handshake
        row_sum = 0
        for i, pmp_id in enumerate(pmp_ids):
            q = float(row_vals.get(pmp_id, 0))
            ws.cell(row=curr_row, column=START_COL+i).value = q
            row_sum += q
        
        # Write Hard Number to Grand Total cell (so RECO can read it)
        ws.cell(row=curr_row, column=sum_col).value = row_sum
        ws.cell(row=curr_row, column=desc_col).value = str(mat_desc)
        
        for c in range(1, desc_col + 1):
            cell = ws.cell(row=curr_row, column=c)
            cell.font = Font(name='Calibri', size=F_SIZE)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = medium_border

    # 4. BOTTOM TOTAL ROW
    total_row = last_r + 1
    ws.cell(row=total_row, column=1).value = "Grand Total"
    ws.cell(row=total_row, column=1).font = Font(name='Calibri', size=F_SIZE, bold=True)
    ws.cell(row=total_row, column=1).border = medium_border
    
    for i in range(num_sites + 1):
        col = START_COL + i
        let = get_column_letter(col)
        ws.cell(row=total_row, column=col).value = f"=SUM({let}3:{let}{last_r})"
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(name='Calibri', size=F_SIZE + 4, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = medium_border
        
    blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for c in range(1, desc_col + 1):
        ws.cell(row=2, column=c).fill = blue_fill
        ws.cell(row=total_row, column=c).fill = blue_fill

    return True

def generate_reco_sheet(df_sites, wb):
    """
    Builds the Reconciliation (RECO) sheet with Live Excel Formulas.
    Data is sourced from the Annexure's Grand Total column.
    """
    if 'Reco' not in wb.sheetnames: return False
    ws = wb['Reco']
    
    # ATTEMPT TO ANCHOR R4G Project (K1 / Row 1, Col 11)
    # If merged, we need to ensure it stays put or is re-merged
    project_val = "R4G Project"
    # We will preserve it in a variable and write it back if scaling shifts it
    
    # 0. DEEP WIPE (Except headers and labels)
    for c in range(3, 100):
        for r in range(10, 35):
            cell_w = ws.cell(row=r, column=c)
            if not isinstance(cell_w, MergedCell):
                cell_w.value = None
            
    # 1. ACQUIRE MATERIALS FROM ANNEXURE
    if 'Annexture' not in wb.sheetnames: return False
    ws_ann = wb['Annexture']
    materials = []
    
    # Find Grand Total and Description Columns
    gt_col, desc_col = 0, 0
    for c in range(1, 45):
        h_val = str(ws_ann.cell(row=2, column=c).value or "").upper()
        if "GRAND" in h_val and "TOTAL" in h_val:
            gt_col = c
        elif "MATERIAL DESCRIPTION" in h_val:
            desc_col = c
            
    if gt_col == 0: return False
    # Fallback for desc_col if not found
    if desc_col == 0: desc_col = 2
    
    for r in range(3, 100):
        sap = str(ws_ann.cell(row=r, column=1).value or "")
        if not sap: break
        # Skip the Grand Total summary row from Annexure
        if "GRAND TOTAL" in sap.upper(): continue
        
        desc = ws_ann.cell(row=r, column=desc_col).value
        qty = ws_ann.cell(row=r, column=gt_col).value
        materials.append({'sap': sap, 'desc': desc, 'qty': qty})
        
    if not materials: return False
    
    # 2. DYNAMIC SCALING (Template baseline is 7 materials: C-I)
    START_COL = 3
    num_materials = len(materials)
    
    # Identify and Unmerge R4G Project to prevent shift conflicts
    project_range = None
    for merged_range in list(ws.merged_cells.ranges):
        min_r, min_c, max_r, max_c = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
        if min_r <= 1 <= max_r: # Check if it covers Row 1 where Project box lives
            cell_val = str(ws.cell(row=min_r, column=min_c).value or "")
            if "PROJECT" in cell_val.upper():
                project_range = (min_r, min_c, max_r, max_c)
                ws.unmerge_cells(merged_range.coord)
                ws.cell(row=min_r, column=min_c).value = None # Clear old pos
    
    if num_materials > 7:
        ws.insert_cols(START_COL + 7, num_materials - 7)
        
    last_mat_col = START_COL + num_materials
    
    # 3. POPULATION & FORMULAS
    for i, mat in enumerate(materials):
        col_idx = START_COL + i
        cl = get_column_letter(col_idx)
        ws.cell(row=10, column=col_idx).value = mat['desc']
        ws.cell(row=10, column=col_idx).alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        ws.cell(row=11, column=col_idx).value = mat['sap']
        ws.cell(row=14, column=col_idx).value = safe_float(mat['qty'])
        ws.cell(row=16, column=col_idx).value = 0
        ws.cell(row=18, column=col_idx).value = f"={cl}14+{cl}16"
        ws.cell(row=20, column=col_idx).value = 0
        ws.cell(row=21, column=col_idx).value = 0
        ws.cell(row=22, column=col_idx).value = 0
        ws.cell(row=23, column=col_idx).value = f"=SUM({cl}21:{cl}22)"
        ws.cell(row=25, column=col_idx).value = f"={cl}18-{cl}23"
        ws.cell(row=28, column=col_idx).value = f"={cl}25"
        ws.cell(row=29, column=col_idx).value = 0
        ws.cell(row=31, column=col_idx).value = f"={cl}28+{cl}29"
        ws.cell(row=33, column=col_idx).value = f"={cl}25-{cl}31"
        for r_styl in [10, 11, 14, 16, 18, 20, 21, 22, 23, 25, 28, 29, 31, 33]:
            copy_cell_style(ws.cell(row=r_styl, column=3), ws.cell(row=r_styl, column=col_idx))

    # 4. PURGE TRAILING & RE-ANCHOR PROJECT BOX
    for c_purge in range(last_mat_col, last_mat_col + 20):
        for r_purge in range(10, 35):
            cell_p = ws.cell(row=r_purge, column=c_purge)
            if not isinstance(cell_p, MergedCell):
                cell_p.value = None
                cell_p.border = Border()
                cell_p.fill = PatternFill(fill_type=None)
    
    # Re-merge Project Box at Absolute Anchor (Relative to material end or original K1)
    # The user SS shows it at the end of the materials.
    target_start_col = last_mat_col + 2
    ws.cell(row=1, column=target_start_col).value = "R4G Project"
    # Create the merge range (7 rows high, 4 columns wide approx)
    new_range = f"{get_column_letter(target_start_col)}1:{get_column_letter(target_start_col+4)}7"
    ws.merge_cells(new_range)
    # Style the merged cell
    top_cell = ws.cell(row=1, column=target_start_col)
    top_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    top_cell.font = Font(bold=True, size=14)
    top_cell.border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

    return True

def get_wo_number(master_path, dc_number):
    """Looks up the WO number from the Master Tracker's 'WO' column based on 'BILLING FILE' matching dc_number."""
    try:
        wb = openpyxl.load_workbook(master_path, data_only=True)
        ws = wb.active 
        
        # Row 2 contains headers
        headers = [str(c.value).upper().strip() if c.value else "" for c in ws[2]]
        
        billing_col_idx = None
        wo_col_idx = None
        
        for i, h in enumerate(headers):
            if "BILLING FILE" in h or "DC NUMBER" in h:
                billing_col_idx = i
            if h == "WO":
                wo_col_idx = i
        
        # Fallbacks to identified indices
        if billing_col_idx is None: billing_col_idx = 47
        if wo_col_idx is None: wo_col_idx = 14
        
        for row in ws.iter_rows(min_row=3, values_only=True):
            if str(row[billing_col_idx]).strip().upper() == dc_number.upper():
                return str(row[wo_col_idx]).strip()
        
        return "N/A"
    except Exception as e:
        print(f"Error looking up WO: {e}")
        return "N/A"

def copy_sheet_between_workbooks(src_ws, dst_wb, sheet_name, index=None):
    """Safely copies a sheet from one workbook to another, including values and styles."""
    if sheet_name in dst_wb.sheetnames:
        del dst_wb[sheet_name]
    
    if index is not None:
        dst_ws = dst_wb.create_sheet(sheet_name, index)
    else:
        dst_ws = dst_wb.create_sheet(sheet_name)
        
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                try:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.alignment = copy(cell.alignment)
                except: pass
                    
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))
        
    for col, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col].width = dim.width
        
    for row, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row].height = dim.height

def inject_main_wcc_from_reference(output_path, df_sites, dc_number, wo_number):
    """Uses the 'Reference-First' approach to ensure stability in Apple Numbers."""
    try:
        ref_path = "/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/FinaliseBillingFormat/DIGITCOM_ AIRFIBER_DC0105_ JDPR_29-JAN-26_A6 (REJECT) 2.xlsx"
        print(f"- Injecting Main WCC from reference: {os.path.basename(ref_path)}")
        
        # 1. Load the reference template as the BASE workbook
        wb_final = openpyxl.load_workbook(ref_path)
        
        # 2. Update Main WCC
        ws_main = wb_final['Main WCC']
        ws_main['D32'] = f"{len(df_sites)} SITES"
        
        date_col = next((c for c in df_sites.columns if 'COMPLETION' in c.upper() or 'RFS DATE' in c.upper()), None)
        date_range = "N/A"
        if date_col:
            dates = pd.to_datetime(df_sites[date_col], errors='coerce')
            min_date = dates.min().strftime('%d-%b-%y').upper() if not dates.isna().all() else "N/A"
            max_date = dates.max().strftime('%d-%b-%y').upper() if not dates.isna().all() else "N/A"
            date_range = f"{min_date} TO {max_date}"
        ws_main['I32'] = date_range
        ws_main['D29'] = wo_number
        
        # 3. Load programmatic sheets from the temp file
        wb_temp = openpyxl.load_workbook(output_path)
        
        # 4. Copy sheets INTO template base
        sheets_to_copy = ['JMS', 'WCC', 'Abstract', 'BOQ', 'Declaration', 'Reco', 'Annexture']
        for sn in wb_temp.sheetnames:
            # Exact match or prefix match, but NEVER match 'Main WCC'
            if sn in sheets_to_copy or any(sn.startswith(base) for base in sheets_to_copy):
                if sn != 'Main WCC':
                    print(f"  - Copying sheet: {sn}")
                    copy_sheet_between_workbooks(wb_temp[sn], wb_final, sn)
        
        wb_final.save(output_path)
        print("- Hybrid Injection COMPLETE")
        
    except Exception as e:
        print(f"Error injecting template: {e}")

def populate_declaration_data(df_sites, wb, dc_number):
    """
    Updates the Declaration sheet with the dynamic site count.
    """
    if 'Declaration' not in wb.sheetnames: return
    ws = wb['Declaration']
    num_sites = len(df_sites)
    
    # Find the cell with "SITES" and replace the number
    import re
    for r in range(1, 30):
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            v = str(cell.value or "")
            if 'SITES' in v.upper():
                new_v = re.sub(r'\d+\s+SITES', f"{num_sites} SITES", v, flags=re.I)
                cell.value = new_v

def main():
    parser = argparse.ArgumentParser(description="Unified Precision Billing Engine")
    parser.add_argument("master_path", help="Path to Master Tracker")
    parser.add_argument("dc_number", help="DC Code (e.g. DC0105)")
    parser.add_argument("--template", help="Path to Master Template", default='Billing/MASTER_JMS_TEMPLATE.xlsx')
    parser.add_argument("--output", help="Path to save the generated file")
    parser.add_argument("--mindump", help="Path to MINDUMP File")
    args = parser.parse_args()

    master_path = args.master_path
    dc_number = args.dc_number
    template_path = args.template
    output_path = args.output if args.output else f"Billing/{dc_number}_Unified_Billing.xlsx"
    mindump_path = args.mindump

    print(f"--- Launching Unified Precision Billing Engine for {dc_number} ---")
    
    df_sites, code_to_col_idx = load_master_data(master_path, dc_number)
    
    if df_sites is not None and not df_sites.empty:
        try:
            # Load from provided or default template path
            wb = openpyxl.load_workbook(template_path)
            
            # Global Header Swap
            target_sheets = wb.sheetnames
            for sheet_name in target_sheets:
                ws_h = wb[sheet_name]
                for r in range(1, 10):
                    for c in range(1, 20):
                        cell = ws_h.cell(row=r, column=c)
                        if "DC-CODE" in str(cell.value):
                            cell.value = str(cell.value).replace("DC-CODE", dc_number)

            # Step 1: Main WCC & WCC
            print("- Step 1: Populating WCC and Main WCC Headers...")
            generate_wcc_sheet(df_sites, wb)
            
            wo_number = get_wo_number(master_path, dc_number)
            inject_main_wcc_from_reference(wb, df_sites, dc_number, wo_number)
            
            # Step 2: JMS
            print("- Step 2: Populating JMS (Style DNA Mirroring)...")
            populate_main_matrix('JMS', df_sites, code_to_col_idx, wb)
            
            # Step 3: Cloning into Abstract & BOQ
            print("- Step 3: Cloning JMS into Abstract & BOQ...")
            jms_ws = wb['JMS']
            for name in ['Abstract', 'BOQ']:
                if name in wb.sheetnames: wb.remove(wb[name])
                new_ws = wb.copy_worksheet(jms_ws)
                new_ws.title = name
                for r in range(1, 10):
                    for c in range(1, 40):
                        cell = new_ws.cell(row=r, column=c)
                        if str(cell.value).upper() == "JMS": cell.value = name.upper()
            
            # Step 4: Declaration
            print("- Step 4: Updating Declaration...")
            populate_declaration_data(df_sites, wb, dc_number)
            
            # Step 5: Annexure
            print("- Step 5: Populating Annexure (LatestSnapshot)...")
            generate_annexure_sheet(df_sites, wb, mindump_path=mindump_path)
            
            # Step 6: RECO
            print("- Step 6: Populating RECO (Pure Logic)...")
            generate_reco_sheet(df_sites, wb)

            # Ensure output directory exists for backend reliability
            output_dir = os.path.dirname(os.path.abspath(output_path))
            if not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
                
            wb.save(output_path)
            
            # Step 7: Hybrid Injection for Stable Main WCC
            wo_number = get_wo_number(master_path, dc_number)
            inject_main_wcc_from_reference(output_path, df_sites, dc_number, wo_number)
            
            print(f"COMPLETE: {output_path}")
        except Exception as e:
            print(f"ERROR: {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)
    else:
        if df_sites is None:
            print(f"CRITICAL ERROR: Failed to load Master Tracker file at {master_path}")
        else:
            print(f"DATA ERROR: No site records found for DC Number '{dc_number}' in the provided Master Tracker.")
        sys.exit(1)

if __name__ == "__main__":
    main()
