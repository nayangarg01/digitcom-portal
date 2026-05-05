import openpyxl

file_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/InputBilling/MIN DUMP-RJST TILL 31 MAR 26.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    print(f"Sheets in MINDUMP: {wb.sheetnames}")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- {sheet_name} First 5 rows ---")
        for row in ws.iter_rows(max_row=5, values_only=True):
            print(row)
except Exception as e:
    print(f"Error: {e}")
