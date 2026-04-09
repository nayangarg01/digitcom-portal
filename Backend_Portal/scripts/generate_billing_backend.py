import pandas as pd
import openpyxl
import sys
import argparse
import os
from copy import copy
from openpyxl.utils import get_column_letter

def load_master_data(master_path, target_billing_file):
    """Loads the Master DPR sheet with Row 1 as Item Codes and Row 2 as Headers."""
    try:
        # Load raw data to capture first row (Row 0) codes
        df_raw = pd.read_excel(master_path, sheet_name='KM SHEET', header=None)
        
        # Row 0 (Index 0): Item codes
        item_codes = df_raw.iloc[0].tolist()
        # Row 1 (Index 1): Actual Column Headers
        col_names = df_raw.iloc[1].tolist()
        
        # Site data starts at Row 2 (Index 2)
        df_sites = df_raw.iloc[2:].copy()
        df_sites.columns = [str(c).strip() for c in col_names]
        
        # Identify the BILLING FILE column index to filter rows
        bill_col = next((c for c in df_sites.columns if 'BILLING FILE' in c.upper()), None)
        if not bill_col:
            print("Error: BILLING FILE column not found in Master DPR Row 2.")
            return None, None
            
        df_filtered = df_sites[df_sites[bill_col].astype(str).str.strip().str.upper() == target_billing_file.upper()].copy()
        
        # Map Item Code (as string) to column index in df_filtered
        code_to_col_idx = {}
        for idx, code in enumerate(item_codes):
            if pd.notna(code):
                # Standardize code to string (e.g. 3367489)
                try:
                    s_code = str(int(code))
                    code_to_col_idx[s_code] = idx
                except:
                    code_to_col_idx[str(code).strip()] = idx
                    
        return df_filtered, code_to_col_idx

    except Exception as e:
        print(f"Error loading master file {master_path}: {e}")
        return None, None

def generate_wcc_sheet(df_sites, wb):
    """Injects the filtered sites into the WCC Template sheet using eNBsiteID."""
    if 'WCC' not in wb.sheetnames:
        print("WCC sheet missing.")
        return False
    ws = wb['WCC']
    
    # Locate headers in WCC
    header_row_idx = None
    cols_map = {}
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and 'GIS SECTOR_ID' in str(row):
            header_row_idx = r_idx
            for c_idx, val in enumerate(row, start=1):
                if val: cols_map[str(val).strip()] = c_idx
            break
            
    if not header_row_idx:
        print("Could not locate WCC headers.")
        return False

    start_row = header_row_idx + 1
    base_styles = {}
    for c_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=start_row, column=c_idx)
        base_styles[c_idx] = {
            'font': copy(cell.font), 'border': copy(cell.border), 'fill': copy(cell.fill),
            'number_format': cell.number_format, 'alignment': copy(cell.alignment)
        }

    # Manage row expansion
    if len(df_sites) > 22:
        ws.insert_rows(start_row + 22, amount=(len(df_sites) - 22))
    elif len(df_sites) < 22:
        ws.delete_rows(start_row + len(df_sites), amount=(22 - len(df_sites)))
        
    aktbc_col = next((c for c in df_sites.columns if 'CHRG EXTRA TRANSPORT' in c.upper() or 'AKTBC' == c.upper()), None)

    for i, (_, row) in enumerate(df_sites.iterrows()):
        curr_row = start_row + i
        def get_val(matcher):
            c_name = next((c for c in df_sites.columns if matcher.upper() in c.upper()), None)
            return row[c_name] if c_name else ""
            
        mapping = [
            ('Sr. No', i + 1),
            ('ENB SITE ID', get_val('ENBSITEID')),
            ('PMP SAP ID', get_val('PMP ID')),
            ('GIS SECTOR_ID', get_val('GIS SECTOR')),
            ('No of Sectors', get_val('NO OF SECTOR')),
            ('Tower type', get_val('Tower type')),
            ('JC', get_val('JC')), ('WH', get_val('WH')), ('VEHICLE NO', get_val('VEHICLE NO')),
            ('MIN  NO', get_val('MIN NO')), ('MIN Date', get_val('MIN DATE')), ('Completion Date', get_val('Completion Date')),
            ('REMARKS', "RFS DONE" if pd.notna(get_val('Completion Date')) and str(get_val('Completion Date')) != "" else ""),
            ('ACTUAL KM', float(row[aktbc_col]) if aktbc_col and pd.notna(row[aktbc_col]) else 0.0),
            ('KM IN WO', float(get_val('KM IN WO')) if pd.notna(get_val('KM IN WO')) and str(get_val('KM IN WO')) != "" else 0.0)
        ]
        
        act_km = next((v for k, v in mapping if k == 'ACTUAL KM'), 0.0)
        wo_km = next((v for k, v in mapping if k == 'KM IN WO'), 0.0)
        mapping.append(('GAP', act_km - wo_km))
        mapping.append(('USED KM IN WCC', act_km if act_km <= wo_km else wo_km))

        for c_idx in range(1, ws.max_column + 1):
            c = ws.cell(row=curr_row, column=c_idx)
            styles = base_styles.get(c_idx)
            if styles:
                c.font, c.border, c.fill = copy(styles['font']), copy(styles['border']), copy(styles['fill'])
                c.number_format, c.alignment = styles['number_format'], copy(styles['alignment'])

        for col_name, val in mapping:
            c_target = next((cols_map[k] for k in cols_map if col_name.upper() in k.upper()), None)
            if c_target:
                target_cell = ws.cell(row=curr_row, column=c_target)
                target_cell.value = val.to_pydatetime() if isinstance(val, pd.Timestamp) else val
                
    # Update totals
    last_r = start_row + len(df_sites) - 1
    for sc in ['ACTUAL KM', 'USED KM IN WCC']:
        c_i = next((cols_map[k] for k in cols_map if sc.upper() in k.upper()), None)
        if c_i:
            let = get_column_letter(c_i)
            ws.cell(row=last_r+1, column=c_i).value = f"=SUM({let}{start_row}:{let}{last_r})"
    return True

def generate_jms_sheet(df_sites, code_to_col_idx, wb):
    """Horizontal JMS injection using exact Item Code mapping and eNBsiteID."""
    if 'JMS' not in wb.sheetnames:
        print("JMS sheet missing.")
        return False
    ws = wb['JMS']
    
    # Dynamic row detection for JMS
    jms_rows = {}
    for r in range(1, 16):
        val = str(ws.cell(row=r, column=2).value).upper().strip() if ws.cell(row=r, column=2).value else ""
        if "COUNT -" in val: jms_rows['COUNT'] = r
        if "SITE ID --" in val: jms_rows['SITE_ID'] = r
        if "SITE TYPE" in val: jms_rows['TOWER'] = r
        if "SECTORS" in val: jms_rows['SECTORS'] = r
    
    SITE_ROW = jms_rows.get('SITE_ID', 10)
    COUNT_ROW = jms_rows.get('COUNT', 9)
    TOWER_ROW = jms_rows.get('TOWER', 11)
    SECTOR_ROW = jms_rows.get('SECTORS', 12)
    START_COL = 4
    
    # Pre-clean matrix to zero (expanded to cover up to Row 55 just in case)
    for r in range(16, 55):
        for i in range(len(df_sites) if len(df_sites) > 22 else 22):
            ws.cell(row=r, column=START_COL + i).value = 0

    # Column Management
    if len(df_sites) > 22:
        ws.insert_cols(START_COL + 22, amount=(len(df_sites) - 22))
    elif len(df_sites) < 22:
        ws.delete_cols(START_COL + len(df_sites), amount=(22 - len(df_sites)))
        
    # Styles copying (Take styles from Row 16, Col 4 as base)
    col_styles = {}
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(row=r, column=START_COL)
        col_styles[r] = {
            'font': copy(cell.font), 'border': copy(cell.border), 'fill': copy(cell.fill),
            'number_format': cell.number_format, 'alignment': copy(cell.alignment)
        }

    # Sites Transposing
    for i, (_, row) in enumerate(df_sites.iterrows()):
        curr_col = START_COL + i
        ws.cell(row=COUNT_ROW, column=curr_col).value = i + 1
        ws.cell(row=SITE_ROW, column=curr_col).value = str(row['eNBsiteID']).strip()
        ws.cell(row=TOWER_ROW, column=curr_col).value = str(row['Tower type']).strip()
        ws.cell(row=SECTOR_ROW, column=curr_col).value = row['NO OF SECTOR']
        
        for r_idx, s in col_styles.items():
            if r_idx < 10: continue
            c = ws.cell(row=r_idx, column=curr_col)
            c.font, c.border, c.fill, c.number_format, c.alignment = copy(s['font']), copy(s['border']), copy(s['fill']), s['number_format'], copy(s['alignment'])

    # Item Row Mapping (Precision)
    for r_idx in range(16, ws.max_row + 1):
        # Column A (Index 1) has Item Code
        cell_val = ws.cell(row=r_idx, column=1).value
        # Standardize to string
        try:
            item_code = str(int(cell_val)) if cell_val else ""
        except:
            item_code = str(cell_val).strip() if cell_val else ""
            
        if item_code in code_to_col_idx:
            # We found the exact column in Master DPR
            target_col_idx = code_to_col_idx[item_code]
            # Write quantities for each site
            for i, (_, row) in enumerate(df_sites.iterrows()):
                val = row.iloc[target_col_idx]
                ws.cell(row=r_idx, column=START_COL + i).value = float(val) if pd.notna(val) else 0.0

    # Summary Formulas
    summary_start_col = None
    for c_idx in range(START_COL + len(df_sites), ws.max_column + 1):
        if "Total Quantity" in str(ws.cell(row=SITE_ROW + 1, column=c_idx).value):
            summary_start_col = c_idx
            break
    if summary_start_col:
        qty_c, rate_c, amt_c = summary_start_col, summary_start_col + 1, summary_start_col + 2
        fcl, lcl = get_column_letter(START_COL), get_column_letter(START_COL + len(df_sites) - 1)
        
        # Scan for Grand Total row
        for r in range(16, ws.max_row + 1):
            # Scan columns for the word "TOTAL" 
            row_vals = [str(ws.cell(row=r, column=c_idx).value).upper() for c_idx in range(1, ws.max_column + 1)]
            if any("TOTAL" in v for v in row_vals):
                ws.cell(row=r, column=amt_c).value = f"=SUM({get_column_letter(amt_c)}16:{get_column_letter(amt_c)}{r-1})"
                break
            
            # Fill row sums
            ws.cell(row=r, column=qty_c).value = f"=SUM({fcl}{r}:{lcl}{r})"
            ws.cell(row=r, column=amt_c).value = f"={get_column_letter(qty_c)}{r}*{get_column_letter(rate_c)}{r}"
    return True

def main():
    parser = argparse.ArgumentParser(description="Precision Unified JIO Billing Sheet Generator.")
    parser.add_argument("master_file", help="Path to MASTERDPR.xlsx")
    parser.add_argument("billing_target", help="DC Code (e.g., DC0105)")
    parser.add_argument("template_path", help="Path to the template file")
    parser.add_argument("output_path", help="Path to save the generated file")
    args = parser.parse_args()

    print(f"--- Launching Precision Mapping for {args.billing_target.upper()} ---")
    df_sites, code_to_col_idx = load_master_data(args.master_file, args.billing_target)
    if df_sites is not None:
        try:
            wb = openpyxl.load_workbook(args.template_path)
            print("- Populating WCC (using eNBsiteID)...")
            generate_wcc_sheet(df_sites, wb)
            print("- Populating JMS (using Item Code Precision Mapping)...")
            generate_jms_sheet(df_sites, code_to_col_idx, wb)
            wb.save(args.output_path)
            print(f"COMPLETE: {args.output_path}")
        except Exception as e:
            print(f"ERROR: {e}")
            import traceback
            traceback.print_exc()
    else:
        print("Data failed to load.")

if __name__ == '__main__':
    main()
