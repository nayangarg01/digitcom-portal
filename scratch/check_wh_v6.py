import openpyxl

generated_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/Billing/DC0111_A6B6_Billing_Final_Test_v6.xlsx'

try:
    wb = openpyxl.load_workbook(generated_path, data_only=True)
    if 'Declaration' in wb.sheetnames:
        ws = wb['Declaration']
        print("\n--- Declaration Sheet Warehouse ---")
        print(f"Row 8: {ws['B8'].value}")
        
    if 'Reco-A6' in wb.sheetnames:
        ws = wb['Reco-A6']
        print("\n--- Reco-A6 Warehouse ---")
        # Warehouse is Row 6 in the meta block
        print(f"Row 6: {ws['B6'].value}")
except Exception as e:
    print(f"Error: {e}")
