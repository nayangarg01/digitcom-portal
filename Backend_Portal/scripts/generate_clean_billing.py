import pandas as pd
import xlsxwriter
import sys
import argparse
import os
import openpyxl
from copy import copy
from openpyxl.drawing.image import Image as OpenpyxlImage

def safe_float(val):
    if pd.isna(val) or str(val).strip() == "" or str(val).strip().upper() in ["NR", "NA", "N.A", "-"]:
        return 0.0
    try:
        return float(val)
    except:
        return 0.0

def format_date(val):
    if pd.isna(val) or str(val).strip() == "":
        return ""
    try:
        dt = pd.to_datetime(val)
        return dt.strftime('%d-%b-%y')
    except:
        return val

def get_warehouse_name(code):
    code = str(code).strip().upper()
    mapping = {
        'JLKD': 'JAIPUR',
        'JLJH': 'JODHPUR',
        'JLJQ': 'SAFEDABAD'
    }
    return mapping.get(code, code if code else 'JODHPUR')

def load_master_data(master_path, dc_number, activity='A6'):
    try:
        sheet_name = 'A6+B6 Billings' if activity == 'A6_B6' else 0
        df_full = pd.read_excel(master_path, sheet_name=sheet_name, header=None)
        if df_full.empty: return None, None
            
        dc_col_idx = 0
        # For A6+B6, headers are usually in row 0, but for safety we check row 1 as well
        header_row = 1 if activity == 'A6' else 0
        raw_headers_list = df_full.iloc[header_row].tolist() if len(df_full) > header_row else []
        
        for i, h in enumerate(raw_headers_list):
            h_str = str(h).upper().strip()
            if "BILLING FILE" in h_str or "DC NUMBER" in h_str:
                dc_col_idx = i
                break

        df_sites = df_full[df_full.iloc[:, dc_col_idx].astype(str).str.strip().str.upper() == dc_number.upper()].copy()
        if df_sites.empty: return None, None
            
        raw_headers = df_full.iloc[header_row].tolist()
        df_sites.columns = [str(h).strip() for h in raw_headers]
        
        code_to_col_idx = {}
        # Matrix mapping (only relevant for A6 for now, or if A6+B6 uses same item codes)
        row0 = df_full.iloc[0].tolist()
        row1 = df_full.iloc[1].tolist()
        for i in range(len(row0)):
            sap = str(row0[i]).split('.')[0].strip()
            if sap and sap != 'nan': code_to_col_idx[sap] = i
            desc = str(row1[i]).strip()
            if desc and desc != 'nan': code_to_col_idx[desc] = i
                
        return df_sites, code_to_col_idx
    except Exception as e:
        print(f"Error loading Master: {e}")
        return None, None

# Template structure for Matrix
TEMPLATE_ITEMS = [
  {"sap": "3367489", "desc": "CHRG BASE RADIO INST&COMS AZIMUTH", "uom": "EA", "rate": 200},
  {"sap": "3367548", "desc": "CHRG LAYING-MULTIMODE FIBER CABLE", "uom": "M", "rate": 5},
  {"sap": "3137158", "desc": "LAYING-2CX2.5SQMM POWER CABLE", "uom": "M", "rate": 12},
  {"sap": "3397253", "desc": "CHRG EXTRA TRANSPORT. > 50 KM (PICKUP", "uom": "KM", "rate": 20},
  {"sap": "3397271", "desc": "CHRG TRANSPORT-UPTO 50KM-SCV-2-3SITE", "uom": "EA", "rate": 1500},
  {"sap": "3367713", "desc": "CHRG ATP-11C - WIFI A6 DEPLOYMENT", "uom": "EA", "rate": 700},
  {"sap": "3367739", "desc": "CHRG ATP-11A - WIFI A6 DEPLOYMENT", "uom": "EA", "rate": 700},
  {"sap": "3317347", "desc": "SITC 1CX6 SQMM CU CBL YY FRLSH", "uom": "M", "rate": 68},
  {"sap": "3383067", "desc": "CHRG APPLY PUFF SEALENT AT CABLE ENTRY", "uom": "EA", "rate": 330},
  {"sap": "3269867", "desc": "TERMINATION OF CABLE 1CX6 SQMM Y", "uom": "M", "rate": 40},
  {"sap": "3397248", "desc": "EXTRA VISIT", "uom": "EA", "rate": 1000},
  {"sap": "3268025", "desc": "INSTALLATION OF POLE MOUNT ON TOWER", "uom": "EA", "rate": 500}
]

TEMPLATE_ITEMS_A6_B6 = [
  {"sap": "3398758", "desc": "ITC A6 & B6 - ONE-SECTOR SITE", "uom": "EA", "rate": 13015},
  {"sap": "3398834", "desc": "ITC A6 & B6 - TWO-SECTOR SITE", "uom": "EA", "rate": 14915},
  {"sap": "3398764", "desc": "ITC A6 & B6 - THREE-SECTOR SITE", "uom": "EA", "rate": 16815},
  {"sap": "3339581", "desc": "CHRG TRANSPORTATION-BEYOND 100 KM-SCV", "uom": "KM", "rate": 20}
]

def get_wo_number(master_path, dc_number):
    """Looks up the WO number from the Master Tracker's 'WO' column based on 'BILLING FILE' matching dc_number."""
    try:
        wb = openpyxl.load_workbook(master_path, data_only=True)
        ws = wb.active 
        
        # Row 2 contains headers
        headers = [str(c.value).upper().strip() if c.value else "" for c in ws[2]]
        
        billing_col_idx = None
        wo_col_idx = None
        
        for i, h in enumerate(headers):
            if "BILLING FILE" in h or "DC NUMBER" in h:
                billing_col_idx = i
            if h == "WO":
                wo_col_idx = i
        
        # Fallbacks to identified indices
        if billing_col_idx is None: billing_col_idx = 47
        if wo_col_idx is None: wo_col_idx = 14
        
        for row in ws.iter_rows(min_row=3, values_only=True):
            if str(row[billing_col_idx]).strip().upper() == dc_number.upper():
                return str(row[wo_col_idx]).strip()
        
        return "N/A"
    except Exception as e:
        print(f"Error looking up WO: {e}")
        return "N/A"

def copy_sheet_between_workbooks(src_ws, dst_wb, sheet_name, index=None):
    """Safely copies a sheet from one workbook to another, including values and styles."""
    if sheet_name in dst_wb.sheetnames:
        del dst_wb[sheet_name]
    
    if index is not None:
        dst_ws = dst_wb.create_sheet(sheet_name, index)
    else:
        dst_ws = dst_wb.create_sheet(sheet_name)
        
    # Copy values and basic styles
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                try:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.alignment = copy(cell.alignment)
                except:
                    pass
                    
    # Merged cells
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))
        
    # Column Dimensions
    for col, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col].width = dim.width
        
    # Row Dimensions
    for row, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row].height = dim.height

def inject_main_wcc_template(output_path, ref_path, df_sites, dc_number, wo_number):
    """Uses the 'Reference-First' approach to ensure stability in Apple Numbers."""
    try:
        print(f"- Injecting Main WCC using Stable Hybrid logic...")
        # 1. Load the reference template as the BASE workbook (this ensures a healthy structure)
        wb_final = openpyxl.load_workbook(ref_path)
        
        # 2. Update Main WCC in the base template
        if 'Main WCC' not in wb_final.sheetnames:
            print("ERROR: 'Main WCC' not found in template.")
            return
            
        ws_main = wb_final['Main WCC']
        ws_main['D32'] = f"{len(df_sites)} SITES"
        
        # Completion Date
        date_col = next((c for c in df_sites.columns if 'COMPLETION' in c.upper() or 'RFS DATE' in c.upper()), None)
        date_range = "N/A"
        if date_col:
            dates = pd.to_datetime(df_sites[date_col], errors='coerce')
            min_date = dates.min().strftime('%d-%b-%y').upper() if not dates.isna().all() else "N/A"
            max_date = dates.max().strftime('%d-%b-%y').upper() if not dates.isna().all() else "N/A"
            date_range = f"{min_date} TO {max_date}"
        ws_main['I32'] = date_range
        ws_main['D29'] = wo_number
        
        # 3. Load the programmatic sheets from the temp XlsxWriter file
        wb_temp = openpyxl.load_workbook(output_path)
        
        # 4. Copy programmatic sheets INTO the template base
        # Using exact matching to avoid overwriting 'Main WCC' because it contains 'WCC'
        sheets_to_copy = ['JMS', 'WCC', 'Abstract', 'BOQ', 'Declaration', 'Reco', 'Annexture']
        for sn in wb_temp.sheetnames:
            if sn in sheets_to_copy or any(sn.startswith(base) for base in sheets_to_copy):
                print(f"  - Copying programmatic sheet: {sn}")
                copy_sheet_between_workbooks(wb_temp[sn], wb_final, sn)
        
        # 5. Save the final file (overwriting the XlsxWriter temp file)
        wb_final.save(output_path)
        print("- Hybrid Generation COMPLETE")
        
    except Exception as e:
        print(f"Error in hybrid generation: {e}")
        import traceback
        traceback.print_exc()

def write_main_wcc(wb, df_sites, dc_number, formats):
    # This is now a placeholder, we will overwrite it with the template using openpyxl
    ws = wb.add_worksheet('Main WCC')
    ws.write('A1', 'TEMPLATE PLACEHOLDER')

def write_wcc(wb, df_sites, dc_number, formats, activity='A6', wo_number='P14/630330726'):
    ws = wb.add_worksheet('WCC')
    
    # Formats with thin borders
    f_title = wb.add_format({'bold': True, 'font_size': 12, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#DCE6F1', 'border': 1})
    f_cert = wb.add_format({'bold': True, 'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'text_wrap': True})
    f_head = wb.add_format({'bold': True, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#DCE6F1', 'border': 1, 'text_wrap': True})
    f_head_yellow = wb.add_format({'bold': True, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFFF00', 'border': 1, 'text_wrap': True})
    f_cell = wb.add_format({'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'border': 1})
    f_date = wb.add_format({'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'num_format': 'dd-mmm-yy'})
    
    if activity == 'A6_B6':
        headers_1 = [
            'Sr. No', 'FT ENB SAP ID', 'FT PMP SAP ID', 'FT GIS SECTOR_ID', 'FB-FT HOP ID', 'No of Sectors', 
            'Tower type ', 'JC', 'WH', 'VEHICLE NO', 'MIN  NO', 'MIN Date', 
            'Completion Date ', 'REMARKS'
        ]
        headers_2 = [
            'ACTUAL KM', 'KM IN WO', 'GAP', 'USED KM IN WCC'
        ]
    else:
        headers_1 = [
            'Sr. No', 'ENB SITE ID', 'PMP SAP ID', 'GIS SECTOR_ID', 'No of Sectors', 
            'Tower type', 'JC', 'WH', 'VEHICLE NO', 'MIN NO', 'MIN Date', 
            'Completion Date', 'REMARKS'
        ]
        headers_2 = [
            'ACTUAL KM', 'KM IN WO', 'GAP', 'USED KM IN WCC'
        ]
    
    # 1. Main Title
    ws.merge_range('C3:P4' if activity == 'A6_B6' else 'C3:O4', 'Work Completion Certificate', formats['title'])
    
    # 2. Certification Text
    cert_text = f"This is to certify that below sites pertaining to WO/WCO No.{wo_number} Dated in 03-10-2025 respect of Digitcom India Technologies  has  been successfully completed in all respect."
    ws.merge_range('C6:P6' if activity == 'A6_B6' else 'C6:O6', cert_text, formats['cert_text'])
    
    # 3. Write Headers
    r_head = 8
    col_idx = 2  # Start at column C (index 2)
    for h in headers_1:
        ws.write(r_head, col_idx, h, formats['header_blue'])
        if 'ID' in h or 'SECTOR' in h:
            ws.set_column(col_idx, col_idx, 22)
        elif 'HOP' in h:
            ws.set_column(col_idx, col_idx, 40)
        elif 'Date' in h or 'REMARKS' in h or 'VEHICLE' in h:
            ws.set_column(col_idx, col_idx, 15)
        else:
            ws.set_column(col_idx, col_idx, 10)
        col_idx += 1
        
    col_idx = 2 + len(headers_1) + 2 # Add some padding columns (Column Q or R)
    start_yellow_col = col_idx
    for h in headers_2:
        ws.write(r_head, col_idx, h, f_head_yellow)
        ws.set_column(col_idx, col_idx, 12)
        col_idx += 1

    def get_val(row, matcher):
        c_name = next((c for c in df_sites.columns if matcher.upper() in c.upper()), None)
        return row[c_name] if c_name else ""
    
    if activity == 'A6_B6':
        aktbc_col = next((c for c in df_sites.columns if 'AKTBC(FT)' in c.upper()), None)
    else:
        aktbc_col = next((c for c in df_sites.columns if 'CHRG EXTRA TRANSPORT' in c.upper() or 'AKTBC' == c.upper()), None)

    r_idx = 9
    total_act = 0
    total_used = 0

    for i, (_, row) in enumerate(df_sites.iterrows()):
        act_km = safe_float(row[aktbc_col]) if aktbc_col else 0.0
        wo_km = safe_float(get_val(row, 'KM IN WO'))
        gap = act_km - wo_km
        # Logic: ACTUAL KM if GAP is negative (ACTUAL < WO), else KM IN WO
        used_km = act_km if gap < 0 else wo_km
        
        total_act += act_km
        total_used += used_km
        
        if activity == 'A6_B6':
            vals_1 = [
                i + 1, get_val(row, 'eNBsiteID'), get_val(row, 'PMP ID'), get_val(row, 'SEC ID'),
                get_val(row, 'FB-FT HOP ID'), safe_float(get_val(row, 'NO OF SECTOR')), get_val(row, 'TOWER'), 
                get_val(row, 'JC'), get_val(row, 'WAREHOUSE'), get_val(row, 'VEHICLE NO'), get_val(row, 'MIN NO'),
                format_date(get_val(row, 'MIN DATE')), format_date(get_val(row, 'RFS DATE')), 
                "RFS DONE" if pd.notna(get_val(row, 'RFS DATE')) and str(get_val(row, 'RFS DATE')) != "" else ""
            ]
        else:
            vals_1 = [
                i + 1, get_val(row, 'ENBSITEID'), get_val(row, 'PMP ID'), get_val(row, 'GIS SECTOR'),
                safe_float(get_val(row, 'NO OF SECTOR')), get_val(row, 'Tower type'), get_val(row, 'JC'),
                get_val(row, 'WH'), get_val(row, 'VEHICLE NO'), get_val(row, 'MIN NO'),
                format_date(get_val(row, 'MIN DATE')), format_date(get_val(row, 'Completion Date')), 
                "RFS DONE" if pd.notna(get_val(row, 'Completion Date')) and str(get_val(row, 'Completion Date')) != "" else ""
            ]
        
        vals_2 = [
            act_km, wo_km, gap, used_km
        ]
        
        for c, val in enumerate(vals_1):
            c_pos = 2 + c
            if isinstance(val, pd.Timestamp):
                ws.write_datetime(r_idx, c_pos, val, formats['date'])
            elif isinstance(val, (int, float)):
                ws.write_number(r_idx, c_pos, val, formats['number'])
            else:
                ws.write(r_idx, c_pos, str(val), formats['cell'])
                
        for c, val in enumerate(vals_2):
            c_pos = start_yellow_col + c
            ws.write_number(r_idx, c_pos, val, formats['number'])
            
        r_idx += 1

    # Totals Row for Yellow Table
    ws.write(r_idx, start_yellow_col, total_act, formats['header_yellow'])
    ws.write(r_idx, start_yellow_col + 1, "", formats['header_yellow'])
    ws.write(r_idx, start_yellow_col + 2, "", formats['header_yellow'])
    ws.write(r_idx, start_yellow_col + 3, total_used, formats['header_yellow'])

    r_sig = r_idx + 2
    ws.write(r_sig, 3, "SIGN:", formats['bold_left'])
    ws.write(r_sig+1, 3, "PROJECT-IN-CHARGE", formats['bold_left'])
    ws.write(r_sig+2, 3, "MR. YUNUS KHAN", formats['bold_left'])
    ws.write(r_sig+3, 3, "DATE:", formats['bold_left'])
    
    ws.write(r_sig, 12 if activity == 'A6' else 13, "SIGN:", formats['bold_left'])
    ws.write(r_sig+1, 12 if activity == 'A6' else 13, "DEPLOYMENT HEAD", formats['bold_left'])
    ws.write(r_sig+2, 12 if activity == 'A6' else 13, "MR. MANISH NAHAR", formats['bold_left'])
    ws.write(r_sig+3, 12 if activity == 'A6' else 13, "DATE:", formats['bold_left'])

def write_matrix_sheet(wb, sheet_name, df_sites, code_to_col_idx, dc_number, formats, include_amounts=True, activity='A6', wo_number='P14/630330726'):
    ws = wb.add_worksheet(sheet_name)
    num_sites = len(df_sites)
    
    # Calculate dates
    date_col = None
    for c in ['Completion Date ', 'Completion Date', 'RFS DATE']:
        if c in df_sites.columns:
            date_col = c
            break
            
    min_date_str = "N/A"
    max_date_str = "N/A"
    if date_col:
        dates = pd.to_datetime(df_sites[date_col], errors='coerce')
        min_date_str = dates.min().strftime('%d-%b-%y').upper() if not dates.isna().all() else "N/A"
        max_date_str = dates.max().strftime('%d-%b-%y').upper() if not dates.isna().all() else "N/A"

    tot_col = 3 + num_sites
    last_col = tot_col + 2 if include_amounts else tot_col
    
    ws.set_column(0, 0, 15)
    ws.set_column(1, 1, 40)
    ws.set_column(2, 2, 8)
    for col in range(3, tot_col):
        ws.set_column(col, col, 5)
    ws.set_column(tot_col, tot_col, 15)
    if include_amounts:
        ws.set_column(tot_col + 1, tot_col + 1, 15)
        ws.set_column(tot_col + 2, tot_col + 2, 15)

    f_title = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#D9D9D9', 'border': 1})
    f_center = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter'})
    f_left = wb.add_format({'bold': True, 'align': 'left', 'valign': 'vcenter'})
    f_right = wb.add_format({'bold': True, 'align': 'right', 'valign': 'vcenter'})
    f_head = wb.add_format({'bold': True, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'bg_color': '#DCE6F1', 'border': 1})
    f_head_vert = wb.add_format({'bold': True, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'rotation': 90, 'bg_color': '#DCE6F1', 'border': 1})
    
    # Title Block
    ws.merge_range(0, 0, 0, last_col, sheet_name, f_title)
    
    # Row 1: Work Order No (Center)
    ws.merge_range(1, 0, 1, last_col, f'Work Order No : {wo_number}', f_center)
    
    # Row 2: Contractor Name (Left) + Work Order Dated (Center)
    mid = last_col // 2
    ws.merge_range(2, 0, 2, mid, 'Contractor Name: DIGITCOM INDIA TECHNOLOGIES', f_left)
    ws.merge_range(2, mid + 1, 2, last_col, 'Work Order Dated: 03-10-2025', f_center)
    
    # Row 3: Project Description (Left)
    ws.merge_range(3, 0, 3, last_col, 'WO for Airspan A6 and C6 Radios for Airfiber' if activity != 'A6_B6' else 'WO for Airspan A6 +B6 Radios for Airfiber', f_left)
    
    # Row 4 & 5: Service Dates (Center)
    ws.merge_range(4, 0, 4, last_col, f'Service Done From Date: {min_date_str}', f_center)
    ws.merge_range(5, 0, 5, last_col, f'Service Done To Date: {max_date_str}', f_center)
    
    # Site Headers
    ws.write(7, 1, 'Count -', formats['bold_right'])
    ws.write(7, 2, '', formats['bold_right'])
    ws.write(8, 0, 'Code', f_head)
    ws.write(8, 1, 'Site ID --', f_head)
    ws.write(8, 2, '', f_head)
    
    ws.set_row(8, 150)  # Increase row height for rotated text
    
    for i, (_, row) in enumerate(df_sites.iterrows()):
        col = 3 + i
        ws.write(7, col, i + 1, formats['number_bold'])
        pmp_id = str(row.get('PMP ID', '')).strip()
        ws.write(8, col, pmp_id, f_head_vert)
        
    ws.write(8, tot_col, 'Total Quantity', f_head)
    if include_amounts:
        ws.write(8, tot_col + 1, 'RATE AS PER SOW', f_head)
        ws.write(8, tot_col + 2, 'AMOUNT', f_head)
        
    # Site Type & Sectors
    ws.write(9, 1, 'Site Type', formats['bold_right'])
    ws.write(9, 2, '', formats['bold_right'])
    for i, (_, row) in enumerate(df_sites.iterrows()):
        tt_col = 'TOWER' if activity == 'A6_B6' else 'Tower type'
        tt = str(row.get(tt_col, '')).strip()
        ws.write(9, 3 + i, tt, formats['number'])
        
    ws.write(10, 1, 'Sectors', formats['bold_right'])
    ws.write(10, 2, '', formats['bold_right'])
    total_sectors = 0
    for i, (_, row) in enumerate(df_sites.iterrows()):
        sec_col = 'NO OF SECTOR' if activity == 'A6_B6' else 'NO OF SECTOR' # Same key actually
        sec = safe_float(row.get(sec_col))
        total_sectors += sec
        ws.write(10, 3 + i, sec, formats['number'])
    ws.write(10, tot_col, total_sectors, formats['number_bold'])
    
    # Data Table Headers
    ws.write(11, 0, 'Item code', f_head)
    ws.write(11, 1, 'Description of Item', f_head)
    ws.write(11, 2, 'UOM', f_head)
    
    r_idx = 12
    items = TEMPLATE_ITEMS_A6_B6 if activity == 'A6_B6' else TEMPLATE_ITEMS
    for item in items:
        ws.write(r_idx, 0, item['sap'], formats['cell'])
        ws.write(r_idx, 1, item['desc'], formats['cell_left'])
        ws.write(r_idx, 2, item['uom'], formats['cell'])
        
        row_sum = 0
        for i, (_, site_row) in enumerate(df_sites.iterrows()):
            col = 3 + i
            sap_code = item['sap']
            
            # Special Mapping Logic
            if activity == 'A6_B6':
                if sap_code == "3339581":
                    val = safe_float(site_row.get('AKTBC(FT)', 0.0))
                else:
                    # For SAP codes like '3398758(ITC A6 & B6 - ONE-SECTOR SITE)', we check for the prefix in columns
                    found_col = next((c for c in df_sites.columns if sap_code in c), None)
                    val = safe_float(site_row[found_col]) if found_col else 0.0
            else:
                val = site_row.iloc[code_to_col_idx[sap_code]] if sap_code in code_to_col_idx else 0.0
                if sap_code == "3397248":
                    val = safe_float(site_row.get('EXTRA VISIT IN WO', 0.0))
                elif sap_code == "3268025":
                    val = safe_float(site_row.get('Polemount in wo', 0.0))
                else:
                    val = safe_float(val)
                
            ws.write(r_idx, col, val, formats['number'])
            row_sum += val
            
        ws.write_formula(r_idx, tot_col, f"=SUM({xlsxwriter.utility.xl_col_to_name(3)}{r_idx+1}:{xlsxwriter.utility.xl_col_to_name(tot_col-1)}{r_idx+1})", formats['number_bold'])
        
        if include_amounts:
            rate = safe_float(item['rate'])
            ws.write(r_idx, tot_col + 1, rate, formats['number'])
            ws.write_formula(r_idx, tot_col + 2, f"={xlsxwriter.utility.xl_col_to_name(tot_col)}{r_idx+1}*{xlsxwriter.utility.xl_col_to_name(tot_col+1)}{r_idx+1}", formats['number_bold'], row_sum * rate)
            
        r_idx += 1
        
    if include_amounts:
        ws.merge_range(r_idx, 0, r_idx, tot_col + 1, "TOTAL", formats['bold_right'])
        ws.write_formula(r_idx, tot_col + 2, f"=SUM({xlsxwriter.utility.xl_col_to_name(tot_col+2)}15:{xlsxwriter.utility.xl_col_to_name(tot_col+2)}{r_idx})", formats['number_bold'])

    r_sig = r_idx + 2
    left_col = 1
    right_col = tot_col + 2 if include_amounts else tot_col
    
    ws.write(r_sig, left_col, "SIGN:", formats['bold_left'])
    ws.write(r_sig+1, left_col, "PROJECT-IN-CHARGE", formats['bold_left'])
    ws.write(r_sig+2, left_col, "MR. YUNUS KHAN", formats['bold_left'])
    ws.write(r_sig+3, left_col, "DATE:", formats['bold_left'])
    
    if sheet_name == 'BOQ':
        mid_col = (left_col + right_col) // 2
        ws.write(r_sig, mid_col, "SIGN:", formats['bold_left'])
        ws.write(r_sig+1, mid_col, "DEPLOYMENT HEAD", formats['bold_left'])
        ws.write(r_sig+2, mid_col, "MR. MANISH NAHAR", formats['bold_left'])
        ws.write(r_sig+3, mid_col, "DATE:", formats['bold_left'])
        
        ws.write(r_sig, right_col, "SIGN:", formats['bold_left'])
        ws.write(r_sig+1, right_col, "RJIO CTO", formats['bold_left'])
        ws.write(r_sig+2, right_col, "MR.RAJEEV KUMAR GUPTA", formats['bold_left'])
        ws.write(r_sig+3, right_col, "DATE:", formats['bold_left'])
    else:
        ws.write(r_sig, right_col, "SIGN:", formats['bold_left'])
        ws.write(r_sig+1, right_col, "DEPLOYMENT HEAD", formats['bold_left'])
        ws.write(r_sig+2, right_col, "MR. MANISH NAHAR", formats['bold_left'])
        ws.write(r_sig+3, right_col, "DATE:", formats['bold_left'])

def write_declaration(wb, df_sites, dc_number, formats, activity='A6'):
    ws = wb.add_worksheet('Declaration')
    
    # Get warehouse name
    wh_code = df_sites['WAREHOUSE'].iloc[0] if 'WAREHOUSE' in df_sites.columns and not df_sites.empty else 'JLJH'
    wh_name = get_warehouse_name(wh_code)
    # Increased widths to prevent Apple Numbers from visually stretching the merged cells
    ws.set_column('A:A', 35)
    ws.set_column('B:C', 30)
    ws.set_column('D:D', 40)
    
    f_title = wb.add_format({'bold': True, 'font_size': 14, 'align': 'center', 'valign': 'vcenter', 'border': 2})
    f_bold = wb.add_format({'bold': True, 'font_size': 10, 'align': 'left', 'valign': 'vcenter', 'border': 1})
    f_bold_center = wb.add_format({'bold': True, 'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'border': 1})
    f_center = wb.add_format({'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'border': 1})
    f_text = wb.add_format({'font_size': 10, 'align': 'left', 'valign': 'top', 'text_wrap': True, 'border': 1})
    f_empty = wb.add_format({'border': 1})
    
    # Row 1
    ws.merge_range('A1:D1', 'DECLARATION STATEMENT', f_title)
    
    # Row 2
    ws.write('A2', 'Name of Contractor', f_bold)
    ws.merge_range('B2:C2', 'DIGITCOM INDIA TECHNOLOGIES', f_bold_center)
    ws.merge_range('D2:D6', '', f_empty) # Logo placeholder
    
    # Row 3 & 4
    ws.merge_range('A3:A4', 'Authorised Signatory', f_bold)
    ws.merge_range('B3:C4', '', f_empty)
    
    # Row 5
    ws.write('A5', 'Vendor Code', f_bold)
    ws.merge_range('B5:C5', '3267708', f_bold_center)
    
    # Row 6
    ws.write('A6', 'Work Order No:', f_bold)
    ws.merge_range('B6:C6', 'P14/630330726', f_bold_center)
    
    # Row 7
    ws.write('A7', 'SAP ID/WBS :', f_bold)
    ws.merge_range('B7:D7', 'As per Annexture', f_bold_center)
    
    # Row 8
    ws.write('A8', 'Warehouse Location', f_bold)
    ws.merge_range('B8:D8', wh_name, f_bold_center)
    
    # Row 9
    ws.merge_range('A9:D9', 'Declaration', f_bold_center)
    
    # Get max date
    date_col = 'Completion Date ' if 'Completion Date ' in df_sites.columns else 'Completion Date'
    max_date_str = "30.03.2026"
    if date_col in df_sites.columns:
        dates = pd.to_datetime(df_sites[date_col], errors='coerce')
        if not dates.isna().all():
            max_date_str = dates.max().strftime('%d.%m.%Y')
            
    cert_text = f"We hereby certify that this Material Reconcilation Statement as on {max_date_str} attached herein is certified and justified by Bills submited by Contractor for given Work Done on given site as per Work Order Issued."
    
    # Row 10-12
    ws.merge_range('A10:D12', cert_text, f_text)
    ws.set_row(9, 30)
    ws.set_row(10, 30)
    ws.set_row(11, 30)
    
    # Row 13
    activity_label = "A6+B6" if activity == "A6_B6" else "A6"
    ws.merge_range('A13:D13', f'{len(df_sites)} SITES({activity_label})', f_bold_center)
    
    # Row 14 (blank narrow with vertical split)
    ws.write('A14', '', f_empty)
    ws.merge_range('B14:D14', '', f_empty)
    ws.set_row(13, 8)
    
    # Row 15 (Headers)
    ws.write('A15', 'Warehouse Incharge', f_bold_center)
    ws.write('B15', 'Circle Project Head', f_bold_center)
    ws.write('C15', 'Contractor', f_bold_center)
    ws.write('D15', 'Remark', f_bold_center)
    
    # Row 16-19 (Signatures row 1)
    ws.merge_range('A16:A19', '', f_empty)
    ws.merge_range('B16:B19', '', f_empty)
    ws.merge_range('C16:C19', '', f_empty)
    ws.merge_range('D16:D23', '', f_empty) # Remarks spans down
    
    # Row 20
    ws.write('A20', 'Checked By Material Coordinator', f_bold)
    ws.merge_range('B20:C20', 'Through FC&A', f_bold_center)
    
    # Row 21-23 (Signatures row 2)
    ws.merge_range('A21:A23', '', f_empty)
    ws.merge_range('B21:C23', '', f_empty)
    
    # Give signature rows height
    ws.set_row(15, 20)
    ws.set_row(16, 20)
    ws.set_row(17, 20)
    ws.set_row(18, 20)
    ws.set_row(20, 20)
    ws.set_row(21, 20)
    ws.set_row(22, 20)

def write_annexure_and_reco(wb, df_sites, dc_number, formats, mindump_path, activity='A6'):
    if activity == 'A6_B6':
        create_annexure_reco_pair(wb, df_sites, formats, mindump_path, "A6", "Annexture-A6", "Reco-A6")
        create_annexure_reco_pair(wb, df_sites, formats, mindump_path, "B6", "Annexture-B6", "Reco-B6")
    else:
        create_annexure_reco_pair(wb, df_sites, formats, mindump_path, "A6", "Annexture", "Reco")

def create_annexure_reco_pair(wb, df_sites, formats, mindump_path, sub_activity, ann_name, rec_name):
    ws_ann = wb.add_worksheet(ann_name)
    ws_rec = wb.add_worksheet(rec_name)
    
def create_annexure_reco_pair(wb, df_sites, formats, mindump_path, sub_activity, ann_name, rec_name):
    ws_ann = wb.add_worksheet(ann_name)
    ws_rec = wb.add_worksheet(rec_name)
    
    # Get warehouse name
    wh_code = df_sites['WAREHOUSE'].iloc[0] if 'WAREHOUSE' in df_sites.columns and not df_sites.empty else 'JLJH'
    wh_name = get_warehouse_name(wh_code)
    
    # Base Site List
    billing_pmp_ids = df_sites['PMP ID'].astype(str).str.strip().tolist()
    
    # If B6, we need a different set of column headers (Far End + Near End PMPs)
    ann_pmp_ids = []
    
    # Process mindump early to build headers if B6
    try:
        sheet_to_read = f"{sub_activity} DUMP"
        try:
            df_mindump = pd.read_excel(mindump_path, sheet_name=sheet_to_read)
        except:
            df_mindump = pd.read_excel(mindump_path)
            
        if sub_activity == 'B6':
            # Logic for B6: Split HOP ID and find matching PMP IDs in dump
            for _, site_row in df_sites.iterrows():
                hop_id = str(site_row.get('FB-FT HOP ID', '')).strip()
                if not hop_id or hop_id.upper() == 'NONE': continue
                
                # Remove _A6 suffix
                clean_hop = hop_id.replace('_A6', '')
                
                # Split logic: Standard pattern is ID1-ID2 where each starts with I-RJ-
                # We split by '-I-RJ-' to keep it safe
                ends = []
                if '-I-RJ-' in clean_hop:
                    parts = clean_hop.split('-I-RJ-')
                    ends = [parts[0], 'I-RJ-' + parts[1]]
                else:
                    ends = [clean_hop]
                
                matched_pmp_for_this_site = []
                for end in ends:
                    if not end: continue
                    # Search in ENB ID, Site ID, or DWG columns
                    mask = (
                        df_mindump['ENB ID'].astype(str).str.contains(end, na=False) |
                        df_mindump['Site ID'].astype(str).str.contains(end, na=False) |
                        df_mindump['DWG'].astype(str).str.contains(end, na=False)
                    )
                    matched_rows = df_mindump[mask]
                    for _, m_row in matched_rows.iterrows():
                        # Try COMMON ID first, then Site ID as fallback for PMP ID
                        pmp = str(m_row.get('COMMON ID', '')).strip()
                        if not pmp or pmp.lower() in ['nan', 'none']:
                            pmp = str(m_row.get('Site ID', '')).strip()
                        
                        if pmp and pmp.lower() not in ['nan', 'none']:
                            if pmp not in matched_pmp_for_this_site:
                                matched_pmp_for_this_site.append(pmp)
                
                # Add to headers in order
                for p in matched_pmp_for_this_site:
                    if p not in ann_pmp_ids:
                        ann_pmp_ids.append(p)
        else:
            # A6 Case: Headers are just the PMP IDs from the billing tracker
            ann_pmp_ids = billing_pmp_ids
            
        # Match Site Logic for Pivot
        def match_site(row):
            if sub_activity == 'B6':
                cid = str(row.get('COMMON ID', '')).strip()
                if not cid or cid.lower() in ['nan', 'none']:
                    cid = str(row.get('Site ID', '')).strip()
                if cid in ann_pmp_ids:
                    return cid
            else:
                wbs = str(row.get('WBS ID', ''))
                site_id = str(row.get('Site ID', ''))
                for pid in ann_pmp_ids:
                    if pid in wbs or pid in site_id:
                        return pid
            return None
            
        df_mindump['Matched_PMP'] = df_mindump.apply(match_site, axis=1)
        df_filtered = df_mindump[df_mindump['Matched_PMP'].notna()]
        
        if not df_filtered.empty:
            pt = pd.pivot_table(df_filtered, values='No. Of Qty', index=['SAP Code', 'Material Description'], columns='Matched_PMP', aggfunc='sum', fill_value=0)
            for pid in ann_pmp_ids:
                if pid not in pt.columns:
                    pt[pid] = 0
            pt = pt[ann_pmp_ids]
            pt = pt.sort_index()
        else:
            pt = pd.DataFrame(columns=ann_pmp_ids)
            
    except Exception as e:
        print(f"Warning: Could not process MINDUMP for {ann_name}: {e}")
        import traceback
        traceback.print_exc()
        ann_pmp_ids = billing_pmp_ids if not ann_pmp_ids else ann_pmp_ids
        pt = pd.DataFrame(columns=ann_pmp_ids)

    # Setup Annexure Sheet with ann_pmp_ids
    num_cols = len(ann_pmp_ids)
    tot_col = num_cols + 1
    desc_col = tot_col + 1
    
    ws_ann.set_column(0, 0, 15)
    for col in range(1, tot_col):
        ws_ann.set_column(col, col, 5)
    ws_ann.set_column(tot_col, tot_col, 10)
    ws_ann.set_column(desc_col, desc_col, 40)
    
    f_title = wb.add_format({'bold': True, 'font_size': 14, 'align': 'center', 'valign': 'vcenter', 'border': 1})
    f_head = wb.add_format({'bold': True, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'border': 1, 'bg_color': '#DCE6F1'})
    f_head_vert = wb.add_format({'bold': True, 'font_size': 8, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True, 'rotation': 90, 'border': 1, 'bg_color': '#DCE6F1'})
    f_cell = wb.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})
    f_cell_bold = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
    f_sum_row = wb.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#DCE6F1'})
    f_sum_row_bold = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#DCE6F1'})
    f_empty = wb.add_format({'border': 1})
    
    ws_ann.merge_range(0, 0, 0, tot_col, f'Annexture-{sub_activity}', f_title)
    ws_ann.write(0, desc_col, '', f_empty)
    
    ws_ann.set_row(1, 120)
    ws_ann.write(1, 0, 'Row Labels', f_head)
    for i, pid in enumerate(ann_pmp_ids):
        ws_ann.write(1, i + 1, pid, f_head_vert)
    ws_ann.write(1, tot_col, 'Grand Total', f_head)
    ws_ann.write(1, desc_col, '', f_empty)
    
    r_idx = 2
    col_sums = {pid: 0 for pid in ann_pmp_ids}
    grand_total_sum = 0
    reco_items = []
    
    for (sap_code, desc), row in pt.iterrows():
        ws_ann.write(r_idx, 0, str(sap_code), f_cell_bold)
        row_total = 0
        for i, pid in enumerate(ann_pmp_ids):
            val = safe_float(row[pid])
            ws_ann.write(r_idx, i + 1, val, f_cell)
            col_sums[pid] += val
            row_total += val
            
        ws_ann.write(r_idx, tot_col, row_total, f_cell)
        grand_total_sum += row_total
        ws_ann.write(r_idx, desc_col, str(desc), f_cell_bold)
        reco_items.append({"sap": str(sap_code), "desc": str(desc), "annexure_row": r_idx + 1})
        r_idx += 1
        
    # Bottom Grand Total Row
    ws_ann.write(r_idx, 0, 'Grand Total', f_sum_row_bold)
    for i, pid in enumerate(ann_pmp_ids):
        ws_ann.write(r_idx, i + 1, col_sums[pid], f_sum_row)
    ws_ann.write(r_idx, tot_col, grand_total_sum, f_sum_row)
    ws_ann.write(r_idx, desc_col, '', f_empty)
    
    # --- RECO SHEET ---
    # Calculate dates
    date_col = None
    for c in ['Completion Date ', 'Completion Date', 'RFS DATE']:
        if c in df_sites.columns:
            date_col = c
            break
            
    min_date_str = "N/A"
    max_date_str = "N/A"
    if date_col:
        dates = pd.to_datetime(df_sites[date_col], errors='coerce')
        min_date_str = dates.min().strftime('%d-%b-%y').upper() if not dates.isna().all() else "N/A"
        max_date_str = dates.max().strftime('%d-%b-%y').upper() if not dates.isna().all() else "N/A"

    num_items = len(reco_items)
    last_col = 1 + num_items
    
    # Setup columns
    ws_rec.set_column(0, 0, 5)   # Col A: Index
    ws_rec.set_column(1, 1, 45)  # Col B: Description Label
    for i in range(num_items):
        ws_rec.set_column(2 + i, 2 + i, 12)
        
    f_reco_header_label = wb.add_format({'bold': True, 'align': 'left', 'valign': 'vcenter', 'border': 1})
    f_reco_header_val = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
    f_r4g_box = wb.add_format({'bold': True, 'font_size': 16, 'align': 'center', 'valign': 'vcenter', 'border': 2})
    f_reco_title = wb.add_format({'bold': True, 'align': 'left', 'valign': 'vcenter', 'border': 1, 'bg_color': '#DCE6F1'})
    
    # Standard Table Formats
    f_reco_item_head = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'text_wrap': True, 'font_size': 9, 'bg_color': '#D9D9D9'})
    f_reco_cell = wb.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'num_format': '0'})
    f_reco_cell_bold = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'num_format': '0'})
    f_reco_label = wb.add_format({'align': 'left', 'valign': 'vcenter', 'border': 1, 'text_wrap': True})
    f_reco_label_bold = wb.add_format({'bold': True, 'align': 'left', 'valign': 'vcenter', 'border': 1, 'text_wrap': True})
    f_reco_index = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1})
    
    # Specific Color Formats
    f_blue_head = wb.add_format({'bold': True, 'align': 'left', 'valign': 'vcenter', 'border': 1, 'bg_color': '#BDD7EE', 'text_wrap': True})
    f_blue_index = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#BDD7EE'})
    
    f_yellow_label = wb.add_format({'bold': True, 'align': 'left', 'valign': 'vcenter', 'border': 1, 'bg_color': '#FFFF00', 'text_wrap': True})
    f_yellow_cell = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#FFFF00', 'num_format': '0'})
    f_yellow_index = wb.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#FFFF00'})
    
    f_empty = wb.add_format({'border': 1})
    
    # Metadata Header Block (Rows 0-6)
    meta_labels = [
        ('Name of Contractor', 'DIGITCOM INDIA TECHNOLOGIES'),
        ('WO No', 'P14/630330726'),
        ('Site ID', 'As per attached Detail'),
        ('WBS Code', 'As per attached Detail'),
        ('Site Name', 'As per attached Detail'),
        ('Warehouse Location', wh_name),
        ('Time Period', f'{min_date_str} TO {max_date_str}')
    ]
    
    for r, (label, val) in enumerate(meta_labels):
        ws_rec.write(r, 0, label, f_reco_header_label)
        ws_rec.merge_range(r, 1, r, 5, val, f_reco_header_label)
        
    # R4G Project Box (Merged Rows 0-6, starting from Col 6 to last_col)
    if last_col >= 6:
        ws_rec.merge_range(0, 6, 6, last_col, 'R4G Project', f_r4g_box)
    else:
        ws_rec.write(0, last_col + 1, 'R4G Project', f_r4g_box)
        
    # Main Title Row (Row 8)
    ws_rec.merge_range(8, 0, 8, last_col, 'Material Reconciliation Statement', f_reco_title)
    
    # Items Headers (Rows 10-11)
    ws_rec.set_row(10, 60)
    ws_rec.write(10, 1, 'Material Description', f_reco_item_head)
    ws_rec.write(11, 1, 'Material Code', f_reco_item_head)
    ws_rec.write(10, 0, '', f_reco_item_head)
    ws_rec.write(11, 0, '', f_reco_item_head)
    
    for i, item in enumerate(reco_items):
        col = 2 + i
        # Keep numeric item headers standard (White) instead of Gray/Blue
        ws_rec.write(10, col, item['desc'], f_reco_cell_bold)
        ws_rec.write(11, col, item['sap'], f_reco_cell_bold)
        
    # Section A
    ws_rec.write(12, 0, 'A.', f_blue_index)
    ws_rec.write(12, 1, 'Receipt details as confirmed by Contractor:', f_blue_head)
    for i in range(num_items): ws_rec.write(12, 2 + i, '', f_empty)
    
    ws_rec.write(14, 0, '1', f_reco_index)
    ws_rec.write(14, 1, 'Received directly from Warehouse', f_yellow_label)
    for i, item in enumerate(reco_items):
        col = 2 + i
        annex_col_str = xlsxwriter.utility.xl_col_to_name(tot_col)
        ws_rec.write_formula(14, col, f"='{ann_name}'!{annex_col_str}{item['annexure_row']}", f_reco_cell)
        
    ws_rec.write(16, 0, '2', f_reco_index)
    ws_rec.write(16, 1, 'Material received from other contractors', f_reco_label)
    for i in range(num_items): ws_rec.write(16, 2 + i, 0, f_reco_cell)
        
    ws_rec.write(18, 0, '', f_reco_index)
    ws_rec.write(18, 1, 'Total (1+2)', f_yellow_label)
    for i in range(num_items):
        col_str = xlsxwriter.utility.xl_col_to_name(2 + i)
        ws_rec.write_formula(18, 2 + i, f"={col_str}15+{col_str}17", f_reco_cell_bold)
        
    # Section B
    ws_rec.write(20, 0, 'B.', f_blue_index)
    ws_rec.write(20, 1, 'Material Transferred to other contractors / returned to Warehouse', f_blue_head)
    for i in range(num_items): ws_rec.write(20, 2 + i, '', f_empty)
    
    ws_rec.write(21, 0, '1', f_reco_index)
    ws_rec.write(21, 1, 'Material Transferred to other contractors', f_reco_label)
    for i in range(num_items): ws_rec.write(21, 2 + i, 0, f_reco_cell)
        
    ws_rec.write(22, 0, '2', f_reco_index)
    ws_rec.write(22, 1, 'Material Returned to Warehouse (in line with MRN Guidelines)', f_reco_label)
    for i in range(num_items): ws_rec.write(22, 2 + i, 0, f_reco_cell)
        
    ws_rec.write(23, 0, '', f_reco_index)
    ws_rec.write(23, 1, 'Total', f_reco_label_bold)
    for i in range(num_items):
        col_str = xlsxwriter.utility.xl_col_to_name(2 + i)
        ws_rec.write_formula(23, 2 + i, f"={col_str}22+{col_str}23", f_reco_cell_bold)
        
    # Section C
    ws_rec.write(25, 0, 'C', f_yellow_index)
    ws_rec.write(25, 1, 'Balance ( A  -  B )', f_yellow_label)
    for i in range(num_items):
        col_str = xlsxwriter.utility.xl_col_to_name(2 + i)
        ws_rec.write_formula(25, 2 + i, f"={col_str}19-{col_str}24", f_reco_cell_bold)
        
    # Section D
    ws_rec.write(27, 0, 'D', f_reco_index)
    ws_rec.write(27, 1, 'CONSUMPTION', f_reco_label_bold)
    for i in range(num_items): ws_rec.write(27, 2 + i, '', f_empty)
    
    ws_rec.write(28, 0, '1', f_reco_index)
    ws_rec.write(28, 1, 'Material Consumed', f_yellow_label)
    for i in range(num_items):
        col_str = xlsxwriter.utility.xl_col_to_name(2 + i)
        annex_col_str = xlsxwriter.utility.xl_col_to_name(tot_col)
        ws_rec.write_formula(28, 2 + i, f"='{ann_name}'!{annex_col_str}{reco_items[i]['annexure_row']}", f_reco_cell)
        
    ws_rec.write(29, 0, '2', f_reco_index)
    ws_rec.write(29, 1, 'Wastage (max as per WO norms)', f_reco_label)
    for i in range(num_items): ws_rec.write(29, 2 + i, 0, f_reco_cell)
        
    ws_rec.write(31, 0, '', f_reco_index)
    ws_rec.write(31, 1, 'Total -Actual consumption (1+2)', f_yellow_label)
    for i in range(num_items):
        col_str = xlsxwriter.utility.xl_col_to_name(2 + i)
        ws_rec.write_formula(31, 2 + i, f"={col_str}29+{col_str}30", f_reco_cell_bold)
        
    # Section E
    ws_rec.set_row(33, 40)
    ws_rec.write(33, 0, 'E.', f_yellow_index)
    ws_rec.write(33, 1, 'Excess consumption for which cost to be recovered from the Contractor (C-D)', f_yellow_label)
    for i in range(num_items):
        col_str = xlsxwriter.utility.xl_col_to_name(2 + i)
        ws_rec.write_formula(33, 2 + i, f"={col_str}26-{col_str}32", f_reco_cell_bold)

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("master_path")
    parser.add_argument("dc_number")
    parser.add_argument("--output", default=None)
    parser.add_argument("--mindump", default=None)
    parser.add_argument("--activity", default='A6', choices=['A6', 'A6_B6'])
    # Ignored legacy flags
    parser.add_argument("--template", default=None) 
    args = parser.parse_args()

    dc_number = args.dc_number
    output_path = args.output if args.output else f"Billing/{dc_number}_Unified_Billing.xlsx"
    activity = args.activity
    
    print(f"--- Generating Clean Billing Workbook for {dc_number} ({activity}) ---")
    df_sites, code_to_col_idx = load_master_data(args.master_path, dc_number, activity=activity)
    
    if df_sites is not None and not df_sites.empty:
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        
        with xlsxwriter.Workbook(output_path, {'nan_inf_to_errors': True}) as wb:
            formats = {
                'title': wb.add_format({'bold': True, 'font_size': 12, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#DCE6F1', 'border': 1}),
                'cert_text': wb.add_format({'bold': True, 'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'border': 1}),
                'header': wb.add_format({'bold': True, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#DCE6F1', 'border': 1, 'text_wrap': True}),
                'header_blue': wb.add_format({'bold': True, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#DCE6F1', 'border': 1, 'text_wrap': True}),
                'header_yellow': wb.add_format({'bold': True, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFFF00', 'border': 1, 'text_wrap': True}),
                'header_vertical': wb.add_format({'bold': True, 'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'rotation': 90, 'bg_color': '#DCE6F1', 'border': 1}),
                'cell': wb.add_format({'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'border': 1}),
                'cell_left': wb.add_format({'font_size': 9, 'align': 'left', 'valign': 'vcenter', 'border': 1}),
                'number': wb.add_format({'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'num_format': '0'}),
                'number_bold': wb.add_format({'font_size': 9, 'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'num_format': '0', 'bg_color': '#F2F2F2'}),
                'bold_right': wb.add_format({'font_size': 9, 'bold': True, 'align': 'right', 'valign': 'vcenter'}),
                'bold_left': wb.add_format({'font_size': 9, 'bold': True, 'align': 'left', 'valign': 'vcenter'}),
                'date': wb.add_format({'font_size': 9, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'num_format': 'dd-mmm-yy'})
            }
            
            wo_number = get_wo_number(args.master_path, dc_number)
            
            write_main_wcc(wb, df_sites, dc_number, formats) 
            write_wcc(wb, df_sites, dc_number, formats, activity=activity, wo_number=wo_number)
            write_matrix_sheet(wb, 'JMS', df_sites, code_to_col_idx, dc_number, formats, include_amounts=True, activity=activity, wo_number=wo_number)
            write_matrix_sheet(wb, 'Abstract', df_sites, code_to_col_idx, dc_number, formats, include_amounts=True, activity=activity, wo_number=wo_number)
            write_matrix_sheet(wb, 'BOQ', df_sites, code_to_col_idx, dc_number, formats, include_amounts=True, activity=activity, wo_number=wo_number)
            write_declaration(wb, df_sites, dc_number, formats, activity=activity)
            write_annexure_and_reco(wb, df_sites, dc_number, formats, args.mindump, activity=activity)

        # AFTER CLOSING WORKBOOK (XlsxWriter), use the Stable Hybrid logic
        script_dir = os.path.dirname(os.path.abspath(__file__))
        ref_template = os.path.join(script_dir, '..', 'templates', 'billing_template.xlsx')
        
        master_tracker = args.master_path
        
        wo_number = get_wo_number(master_tracker, dc_number)
        inject_main_wcc_template(output_path, ref_template, df_sites, dc_number, wo_number)
        
        print(f"COMPLETE: {output_path}")
    else:
        print("ERROR: No valid data found for DC Number.")

if __name__ == "__main__":
    main()
