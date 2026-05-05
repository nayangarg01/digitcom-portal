import openpyxl

def compare_summaries(orig_path, test_path):
    print(f"Comparing {orig_path} vs {test_path}")
    wb_orig = openpyxl.load_workbook(orig_path, data_only=True)
    ws_orig = wb_orig['Sheet1']
    
    wb_test = openpyxl.load_workbook(test_path, data_only=True)
    ws_test = wb_test['Sheet1']
    
    orig_data = {}
    for r in range(4, ws_orig.max_row):
        desc = ws_orig.cell(row=r, column=2).value
        amt = ws_orig.cell(row=r, column=6).value
        if desc and 'Total' not in desc:
            orig_data[desc] = amt
            
    test_data = {}
    for r in range(4, ws_test.max_row):
        desc = ws_test.cell(row=r, column=2).value
        amt = ws_test.cell(row=r, column=6).value
        if desc and 'Total' not in desc:
            test_data[desc] = amt
            
    all_descs = set(orig_data.keys()) | set(test_data.keys())
    for desc in sorted(all_descs):
        o_amt = orig_data.get(desc, 0)
        t_amt = test_data.get(desc, 0)
        if o_amt != t_amt:
            print(f"Diff in '{desc}': Orig={o_amt}, Test={t_amt}, Diff={o_amt - t_amt}")

orig = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/PerformaInvoiceWork/48/MATERIAL_REPORT_FOR_PERFORMA_INVOICE_NO_048_DIGITCOM.xlsx'
test = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/scratch/MATERIAL_REPORT_048_TEST.xlsx'
compare_summaries(orig, test)
