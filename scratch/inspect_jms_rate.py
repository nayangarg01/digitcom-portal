import openpyxl
import os

def inspect_jms_rate(file_path):
    print(f"\n--- Searching for RATE in: {file_path} ---")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['JMS']
    
    # Check item header row (R15 or search for it)
    header_row = -1
    for r in range(1, 40):
        if "Description of Item" in str(ws.cell(row=r, column=2).value):
            header_row = r
            break
    
    if header_row == -1:
        print("Header row not found!")
        return
        
    print(f"Header row: {header_row}")
    for c in range(1, ws.max_column + 1):
        val = str(ws.cell(row=header_row, column=c).value).strip()
        if val and val != 'None':
            print(f"Col {c:2}: {val}")
        if 'RATE' in val.upper():
            print(f"*** FOUND RATE AT COL {c} ***")

jms_ref = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/FinaliseBillingFormat/DIGITCOM_ AIRFIBER_DC0105_ JDPR_29-JAN-26_A6 (REJECT) 2.xlsx'
inspect_jms_rate(jms_ref)
