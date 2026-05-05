import openpyxl

file_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/InputBilling/MIN DUMP-RJST TILL 31 MAR 26.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb['B6 DUMP']
    headers = [str(h).strip() for h in next(ws.iter_rows(max_row=1, values_only=True))]
    enb_col = headers.index('ENB ID')
    
    print("\nSample ENB IDs from B6 DUMP:")
    for row in ws.iter_rows(min_row=2, max_row=20, values_only=True):
        print(row[enb_col])
            
except Exception as e:
    print(f"Error: {e}")
