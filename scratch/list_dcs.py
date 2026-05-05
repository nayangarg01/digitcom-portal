import openpyxl

file_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/InputBilling/MASTER TRACKER FOR BILLING-AIRFIBER-RJST (1).xlsx'
try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    if 'A6+B6 Billings' in wb.sheetnames:
        ws = wb['A6+B6 Billings']
        # Find DC column (BILLING FILE)
        headers = next(ws.iter_rows(max_row=1, values_only=True))
        dc_col_idx = None
        for i, h in enumerate(headers):
            if h and ("BILLING FILE" in str(h).upper() or "DC NUMBER" in str(h).upper()):
                dc_col_idx = i
                break
        
        if dc_col_idx is not None:
            dcs = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = row[dc_col_idx]
                if val:
                    dcs.add(str(val).strip())
            print(f"Available DCs in A6+B6 Billings: {sorted(list(dcs))}")
        else:
            print("DC column not found in headers")
    else:
        print("Sheet 'A6+B6 Billings' not found")
except Exception as e:
    print(f"Error: {e}")
