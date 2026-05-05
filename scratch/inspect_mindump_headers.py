import openpyxl

file_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/InputBilling/MIN DUMP-RJST TILL 31 MAR 26.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = next(ws.iter_rows(max_row=1, values_only=True))
        print(f"Sheet: {sheet_name} | Headers: {headers}")
except Exception as e:
    print(f"Error: {e}")
