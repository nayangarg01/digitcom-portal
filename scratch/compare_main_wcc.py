import openpyxl

ref_file = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/FinaliseBillingFormat/DIGITCOM_ AIRFIBER_DC0105_ JDPR_29-JAN-26_A6 (REJECT) 2.xlsx'
auto_file = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/FinaliseBillingFormat/DC0105_Clean_Billing_1777900012005.xlsx'

try:
    print("--- Reference Manual File (Main WCC) ---")
    wb_ref = openpyxl.load_workbook(ref_file, data_only=True)
    ws_ref = wb_ref['Main WCC']
    for r in range(1, 10):
        row_vals = [ws_ref.cell(row=r, column=c).value for c in range(1, 10)]
        print(f"Row {r}: {row_vals}")
        
    print("\n--- Automated File (Main WCC) ---")
    wb_auto = openpyxl.load_workbook(auto_file, data_only=True)
    ws_auto = wb_auto['Main WCC']
    for r in range(1, 10):
        row_vals = [ws_auto.cell(row=r, column=r).value for r in range(1, 1)] # Wait, typo
        row_vals = [ws_auto.cell(row=r, column=c).value for c in range(1, 10)]
        print(f"Row {r}: {row_vals}")
except Exception as e:
    print(f"Error: {e}")
