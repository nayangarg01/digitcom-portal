import openpyxl
import os
from copy import copy

def get_wo_number(master_path, dc_number):
    try:
        wb = openpyxl.load_workbook(master_path, data_only=True)
        ws = wb.active 
        headers = [str(c.value).upper().strip() if c.value else "" for c in ws[2]]
        billing_col_idx, wo_col_idx = 47, 14
        for i, h in enumerate(headers):
            if "BILLING FILE" in h or "DC NUMBER" in h: billing_col_idx = i
            if h == "WO": wo_col_idx = i
        for row in ws.iter_rows(min_row=3, values_only=True):
            if str(row[billing_col_idx]).strip().upper() == dc_number.upper():
                return str(row[wo_col_idx]).strip()
        return "N/A"
    except: return "N/A"

def inject_main_wcc_template(output_path, ref_path, dc_number, wo_number, final_path):
    print(f"- Injecting Main WCC from template: {os.path.basename(ref_path)}")
    wb_new = openpyxl.load_workbook(output_path)
    wb_ref = openpyxl.load_workbook(ref_path)
    src_ws = wb_ref['Main WCC']
    if 'Main WCC' in wb_new.sheetnames: del wb_new['Main WCC']
    dst_ws = wb_new.create_sheet('Main WCC', 0)
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font); new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill); new_cell.number_format = copy(cell.number_format)
                new_cell.alignment = copy(cell.alignment)
    for merged_range in src_ws.merged_cells.ranges: dst_ws.merge_cells(str(merged_range))
    for col, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col].width = dim.width
        dst_ws.column_dimensions[col].hidden = dim.hidden
    for row, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row].height = dim.height
        dst_ws.row_dimensions[row].hidden = dim.hidden
    
    dst_ws['D32'] = "22 SITES"; dst_ws['I32'] = "30-JAN-26 TO 04-FEB-26"; dst_ws['D29'] = wo_number
    wb_new.save(final_path)
    print(f"- COMPLETE. Saved to {final_path}")

if __name__ == "__main__":
    ref = "/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/FinaliseBillingFormat/DIGITCOM_ AIRFIBER_DC0105_ JDPR_29-JAN-26_A6 (REJECT) 2.xlsx"
    # Use the file I created in the previous failed run if it exists, or the original
    auto = "/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/FinaliseBillingFormat/DC0105_VERIFIED_OUTPUT.xlsx"
    master = "/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/FinaliseBillingFormat/MASTER TRACKER FOR BILLING-AIRFIBER-RJST (1).xlsx"
    output = "/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/FinaliseBillingFormat/DC0105_VERIFIED_FINAL.xlsx"
    
    wo = get_wo_number(master, "DC0105")
    inject_main_wcc_template(auto, ref, "DC0105", wo, output)
