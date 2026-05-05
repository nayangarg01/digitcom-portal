import openpyxl

generated_path = '/Users/nayangarg/Desktop/DigitcomWebsiteRenovation/Old_Codebase_renovated_v6.1/Billing/DC0111_A6B6_Billing_Test_Full_v4.xlsx'

try:
    wb = openpyxl.load_workbook(generated_path, data_only=True)
    print(f"Sheets in generated file: {wb.sheetnames}")
    
    if 'JMS' in wb.sheetnames:
        ws = wb['JMS']
        print("\n--- JMS Header Area (Row 8-13) ---")
        for row in ws.iter_rows(min_row=8, max_row=13, values_only=True):
            print(row)
        print("\n--- JMS Data Area (Row 15-20) ---")
        for row in ws.iter_rows(min_row=15, max_row=20, values_only=True):
            if any(x is not None for x in row):
                print(row)
except Exception as e:
    print(f"Error: {e}")
